# -*- coding: utf-8 -*-

import copy
import json
import os
import re
import threading
import traceback
import uuid
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from tkinter import filedialog, messagebox

from gk_pilepro.gk_core import (
    last_run_dir,
    load_formula_profiles,
    report_runtime_error_to_admin,
    save_formula_profiles,
)

from gk_pilepro.gk_excel import (
    apply_row_formulas_from_template,
    auto_map_columns,
    auto_mapping_to_excel_columns,
    call_gemini,
    capture_formula_columns,
    convert_excel_value,
    copy_row_dimension,
    copy_style_row,
    ensure_no_column_in_mapping,
    find_blank_rows_before_total,
    find_header_row_smart,
    find_last_data_row_before_total,
    find_no_column_smart,
    find_total_row,
    force_workbook_recalculate,
    get_headers_smart,
    get_stt_value,
    is_formula_value,
    is_no_header,
    is_total_marker_text,
    norm,
    normalize_numeric_like_text,
    normalize_vietnam_date,
    row_has_grey_background,
    select_longest_stt_chain,
    update_total_formulas,
)
from gk_pilepro.ui.gk_ocr_ui import (
    _format_elapsed,
    _show_ocr_done_notification,
    _write_ocr_timing_log,
    ensure_table_image_metadata,
)

def _apply_rows_insert_before_total_chot(self, wb):

    """

    CHỐT:

    - Luôn insert dòng mới ngay trước dòng TỔNG.

    - Điền dữ liệu vào dòng vừa insert.

    - Dòng TỔNG và toàn bộ phần sau TỔNG bị đẩy xuống.

    - Không ghi vào vùng trắng/merged có sẵn.

    - SUM lại các cột cần tổng.

    """

    if not self.sheet_var.get():

        raise ValueError("Bạn chưa chọn sheet.")



    table = self.table_editor.get_current_table()

    if not table:

        raise ValueError("Chưa có dữ liệu từ ảnh.")



    ws = wb[self.sheet_var.get()]

    header_row = find_header_row_smart(ws)

    excel_headers = get_headers_smart(ws, header_row)



    self.header_row = header_row

    self.excel_headers = excel_headers



    total_row = find_total_row(ws, header_row)

    if not total_row:

        raise ValueError("Không tìm thấy dòng TỔNG/TOTAL trong Excel.")



    no_col = find_no_column_smart(ws, excel_headers, header_row, total_row)

    if not no_col:

        raise ValueError("Không tìm thấy cột STT/No trong Excel.")



    rows = table.get("rows", [])

    if not rows:

        raise ValueError("Không có dòng dữ liệu để nhập.")



    mapping = self.mapping_editor.get_mapping()

    if not mapping:

        mapping = auto_mapping_to_excel_columns(table["columns"], excel_headers)

        try:

            auto_idx = auto_map_columns(table["columns"], excel_headers)

            auto_idx = ensure_no_column_in_mapping(table["columns"], auto_idx, excel_headers)

            self.mapping_editor.set_mapping(table["columns"], excel_headers, auto_idx)

        except Exception:

            pass



    if not mapping:

        raise ValueError("Chưa có mapping cột.")



    # Lấy STT lớn nhất trước dòng TỔNG

    stt_nums = []

    memo = {}



    for r in range(header_row + 1, total_row):

        try:

            if row_has_grey_background(ws, r):

                continue

        except Exception:

            pass



        row_text = " ".join(str(ws.cell(r, c).value or "") for c in range(1, ws.max_column + 1))

        try:

            if is_total_marker_text(row_text):

                continue

        except Exception:

            pass



        n = None

        try:

            n = get_stt_value(ws, r, no_col, memo)

        except Exception:

            n = None



        if not isinstance(n, int):

            s = str(ws.cell(r, no_col).value or "").strip()

            if s.isdigit():

                n = int(s)



        if isinstance(n, int):

            stt_nums.append((r, n))



    if stt_nums:

        first_data_row = sorted(stt_nums, key=lambda x: x[0])[0][0]

        style_row, last_no = sorted(stt_nums, key=lambda x: (x[1], x[0]))[-1]

    else:

        best = select_longest_stt_chain(ws, no_col, header_row, total_row)

        if best:

            first_data_row = best[0][0]

            style_row = best[-1][0]

            last_no = best[-1][1]

        else:

            first_data_row = header_row + 1

            style_row = total_row - 1

            last_no = 0



    # Nhận diện cột cần SUM dựa trên dòng TỔNG mẫu trước khi insert

    sum_columns = []

    for c in range(1, ws.max_column + 1):

        if c == no_col:

            continue



        total_val = ws.cell(total_row, c).value



        if is_formula_value(total_val):

            sum_columns.append(c)

            continue



        if isinstance(total_val, (int, float)):

            for rr in range(first_data_row, total_row):

                vv = ws.cell(rr, c).value

                if isinstance(vv, (int, float)):

                    sum_columns.append(c)

                    break



    # CHỖ QUAN TRỌNG NHẤT:

    # Insert ngay trước dòng TỔNG, không ghi vào dòng trắng sẵn có

    insert_at = total_row

    row_count = len(rows)



    ws.insert_rows(insert_at, amount=row_count)



    # Sau insert, dòng TỔNG mới bị đẩy xuống

    total_row_after = total_row + row_count

    target_rows = list(range(insert_at, insert_at + row_count))
    effective_first_data_row = min(first_data_row, min(target_rows))
    data_last_row = max(target_rows)



    for i, data_row in enumerate(rows):

        dst_row = target_rows[i]



        # Copy style từ dòng dữ liệu mẫu

        copy_style_row(ws, style_row, dst_row, ws.max_column)



        try:

            copy_row_dimension(ws, style_row, dst_row)

        except Exception:

            pass



        try:

            apply_row_formulas_from_template(ws, style_row, dst_row)

        except Exception:

            pass



        # STT nối tiếp

        self._safe_set_cell_value(ws, dst_row, no_col, last_no + i + 1)



        # Điền dữ liệu theo mapping

        for src_idx, excel_col in enumerate(mapping):

            if excel_col is None:

                continue

            if excel_col == no_col:

                continue



            # Ô công thức thì giữ công thức

            if is_formula_value(ws.cell(dst_row, excel_col).value):

                continue



            val = data_row[src_idx] if src_idx < len(data_row) else ""

            self._safe_set_cell_value(ws, dst_row, excel_col, convert_excel_value(val))



    # SUM lại các cột cần tổng

    sum_first_row = first_data_row

    sum_last_row = target_rows[-1]



    for c in sorted(set(sum_columns)):

        letter = get_column_letter(c)

        ws.cell(total_row_after, c).value = f"=SUM({letter}{sum_first_row}:{letter}{sum_last_row})"



    updated_total_cols = []
    try:
        updated_total_cols = update_total_formulas(
            ws,
            total_row_after,
            effective_first_data_row,
            data_last_row,
            excel_headers=excel_headers,
            no_col=no_col,
        )
    except Exception:
        updated_total_cols = []

    force_workbook_recalculate(wb)



    out = last_run_dir()

    out.mkdir(exist_ok=True)



    logic = {

        "rule": "CHOT_INSERT_TRUOC_TONG",

        "sheet": ws.title,

        "header_row": header_row,

        "total_row_before": total_row,

        "insert_at": insert_at,

        "rows_added": row_count,

        "data_rows": target_rows,

        "total_row_after": total_row_after,

        "last_stt_before": last_no,

        "new_stt_start": last_no + 1,

        "new_stt_end": last_no + row_count,

        "sum_first_row": sum_first_row,

        "sum_last_row": sum_last_row,

        "sum_columns": sorted(set(sum_columns)),

    }



    (out / "chot_insert_truoc_tong_logic.json").write_text(

        json.dumps(logic, ensure_ascii=False, indent=2),

        encoding="utf-8"

    )



    return {

        "sheet": ws.title,

        "header_row": header_row,

        "last_stt_before": last_no,

        "start_fill_row": insert_at,

        "next_stt_start": last_no + 1,

        "rows_added": row_count,

        "total_row_after": total_row_after,

        "sum_first_row": sum_first_row,

        "sum_last_row": sum_last_row,

        "sum_columns_count": len(set(sum_columns)),

    }









# =========================

# V22.9 FINAL OVERRIDE - GIỮ TOÀN BỘ CHỨC NĂNG CŨ, CHỈ CHỐT LOGIC XUẤT EXCEL

# =========================



def _v229_is_formula(v):

    try:

        return is_formula_value(v)

    except Exception:

        return isinstance(v, str) and v.startswith("=")





def _v229_merged_ranges_shift_for_insert(ws, insert_at, amount):

    """

    openpyxl insert_rows không tự dịch merged ranges.

    Hàm này lưu merge, unmerge, insert, rồi merge lại đúng vị trí.

    Nhờ vậy dòng TỔNG + chữ ký + form dưới TỔNG được đẩy xuống đúng, không vỡ layout.

    """

    old_ranges = []

    try:

        for rng in list(ws.merged_cells.ranges):

            old_ranges.append((rng.min_row, rng.min_col, rng.max_row, rng.max_col))

        for rng in list(ws.merged_cells.ranges):

            try:

                ws.unmerge_cells(str(rng))

            except Exception:

                pass

    except Exception:

        old_ranges = []



    ws.insert_rows(insert_at, amount=amount)



    for min_row, min_col, max_row, max_col in old_ranges:

        if min_row >= insert_at:

            min_row += amount

            max_row += amount

        elif min_row < insert_at <= max_row:

            max_row += amount

        try:

            ws.merge_cells(

                start_row=min_row, start_column=min_col,

                end_row=max_row, end_column=max_col

            )

        except Exception:

            pass





def _v229_find_stt_before_total(ws, no_col, header_row, total_row):

    """

    Lấy STT lớn nhất thật sự trước dòng TỔNG.

    Không lấy vùng sau TỔNG/chữ ký. Đọc được STT số và STT công thức =A16+1.

    """

    nums = []

    memo = {}

    for r in range(header_row + 1, total_row):

        try:

            if row_has_grey_background(ws, r):

                continue

        except Exception:

            pass

        try:

            row_text = " ".join(str(ws.cell(r, c).value or "") for c in range(1, ws.max_column + 1))

            if is_total_marker_text(row_text):

                continue

        except Exception:

            pass



        n = None

        try:

            n = get_stt_value(ws, r, no_col, memo)

        except Exception:

            n = None

        if not isinstance(n, int):

            s = str(ws.cell(r, no_col).value or "").strip()

            if s.isdigit():

                n = int(s)

            else:

                try:

                    f = float(s.replace(",", "."))

                    if f.is_integer():

                        n = int(f)

                except Exception:

                    pass

        if isinstance(n, int):

            nums.append((r, n))



    if nums:

        first_row = sorted(nums, key=lambda x: x[0])[0][0]

        style_row, last_no = sorted(nums, key=lambda x: (x[1], x[0]))[-1]

        return first_row, style_row, last_no



    # fallback giữ chức năng cũ

    best = None

    try:

        best = select_longest_stt_chain(ws, no_col, header_row, total_row)

    except Exception:

        best = None

    if best:

        return best[0][0], best[-1][0], best[-1][1]



    last_row = None

    try:

        last_row = find_last_data_row_before_total(ws, header_row, total_row, None, no_col)

    except Exception:

        last_row = None

    return header_row + 1, (last_row or total_row - 1), 0





def _v229_capture_sum_columns(ws, total_row, first_data_row, last_data_row, no_col):

    """

    Chỉ SUM các cột mà file Excel mẫu cần tổng:

    - ô dòng TỔNG có công thức

    - hoặc ô dòng TỔNG là số và phía trên có dữ liệu số

    Không SUM cột STT.

    """

    cols = set()

    def _is_contract_like_header(col_idx):
        try:
            header_row = find_header_row_smart(ws)
            headers = get_headers_smart(ws, header_row)
            header_by_col = {int(c): str(v or "") for c, v in headers}
            n = norm(header_by_col.get(col_idx, ""))
            return any(tok in n for tok in ["hop dong", "contract", "contract no", "so hop dong", "hd"])
        except Exception:
            return False

    for c in range(1, ws.max_column + 1):

        if c == no_col:

            continue

        if _is_contract_like_header(c):
            continue

        v = ws.cell(total_row, c).value

        if _v229_is_formula(v):

            cols.add(c)

            continue

        is_num_total = False

        if isinstance(v, (int, float)):

            is_num_total = True

        else:

            sv = str(v or "").strip().replace(".", "").replace(",", ".")

            try:

                if sv != "":

                    float(sv)

                    is_num_total = True

            except Exception:

                is_num_total = False

        if is_num_total:

            for rr in range(first_data_row, max(first_data_row, last_data_row) + 1):

                vv = ws.cell(rr, c).value

                if isinstance(vv, (int, float)):

                    cols.add(c)

                    break

                svv = str(vv or "").strip().replace(".", "").replace(",", ".")

                try:

                    if svv != "":

                        float(svv)

                        cols.add(c)

                        break

                except Exception:

                    pass

    return sorted(cols)





def _v229_normalize_mapping_to_excel_columns(source_cols, mapping, excel_headers):

    """

    Chấp nhận cả mapping dạng index combobox và dạng cột Excel thật.

    Trả về list source_idx -> excel column number.

    """

    out = []

    header_cols = [c for c, _ in excel_headers]

    max_col = max(header_cols) if header_cols else 0

    for m in (mapping or []):

        if m is None:

            out.append(None)

            continue

        try:

            mi = int(m)

        except Exception:

            out.append(None)

            continue

        # Nếu mi là index trong excel_headers

        if 0 <= mi < len(excel_headers) and mi not in header_cols:

            out.append(excel_headers[mi][0])

        # Nếu mi là cột Excel thật

        elif 1 <= mi <= max_col:

            out.append(mi)

        elif 0 <= mi < len(excel_headers):

            out.append(excel_headers[mi][0])

        else:

            out.append(None)



    # kéo dài mapping nếu thiếu

    while len(out) < len(source_cols):

        out.append(None)



    # STT/No trong ảnh luôn bỏ qua, tool tự nối STT theo Excel

    try:

        for i, src in enumerate(source_cols):

            if is_no_header(src):

                out[i] = None

    except Exception:

        pass

    return out[:len(source_cols)]





def _v229_table_max_col(no_col=None, mapping=None, sum_columns=None, excel_headers=None):
    cols = []
    for source in (mapping or []):
        try:
            if source:
                cols.append(int(source))
        except Exception:
            pass
    for source in (sum_columns or []):
        try:
            if source:
                cols.append(int(source))
        except Exception:
            pass
    for item in (excel_headers or []):
        try:
            cols.append(int(item[0]))
        except Exception:
            pass
    try:
        if no_col:
            cols.append(int(no_col))
    except Exception:
        pass
    return max(cols) if cols else None


def _v229_safe_copy_row(ws, src_row, dst_row, max_col=None):
    try:
        max_col = int(max_col or 0)
    except Exception:
        max_col = 0
    if max_col <= 0:
        max_col = min(int(ws.max_column or 1), 32)

    try:

        copy_style_row(ws, src_row, dst_row, max_col)

    except Exception:

        pass

    try:

        copy_row_dimension(ws, src_row, dst_row)

    except Exception:

        pass

    try:

        apply_row_formulas_from_template(ws, src_row, dst_row)

    except Exception:

        pass





def _v229_filter_data_rows(rows):

    """Bỏ dòng rỗng hoàn toàn trong preview, không đụng dòng có dữ liệu."""

    out = []

    for r in rows or []:

        vals = list(r) if isinstance(r, (list, tuple)) else [r]

        if any(str(x or "").strip() for x in vals):

            out.append(vals)

    return out





def postprocess_to_hop_coc_d1_d2(tables):

    """

    Bổ sung/ghi đè nhẹ rule Tổ hợp cọc:

    - Nếu AI đã trả D1/D2/D3... thì giữ nguyên.

    - Nếu còn cột 'Tổ hợp cọc' nhưng chưa có D1/D2/D3... thì tách giá trị thành D1..D6.

    - Nếu ô tổ hợp chứa '6 10', '6|10', '6 + 10 + 14' thì tách thành nhiều cột D1, D2, D3...

    """

    if not tables:

        return tables



    def _n(x):

        try:

            return norm(x)

        except Exception:

            return str(x or "").lower().strip()



    def _looks_number(x):

        s = str(x or "").strip().replace(",", ".")

        if s.startswith("+") or s.startswith("-"):

            s = s[1:]

        return bool(re.fullmatch(r"\d+(?:\.\d+)?", s))



    def _extract_parts(value):

        text = str(value or "").strip()

        if not text:

            return []

        parts = re.findall(r"[-+]?\d+(?:[,.]\d+)?", text)

        if len(parts) >= 2:

            return parts[:6]

        if any(sep in text for sep in ("|", "+", "/", "\\", ",", ";", " ")):

            tokens = [p for p in re.split(r"[|+/\\,;\s]+", text) if p]

            if len(tokens) >= 2:

                return tokens[:6]

        return parts[:1] if parts else []



    for t in tables:

        cols = list(t.get("columns", []))

        rows = [list(r) for r in t.get("rows", [])]

        if not cols:

            continue

        ncols = [_n(c) for c in cols]

        if any(x in {"d1", "đ1", "1st"} for x in ncols) and any(x in {"d2", "đ2", "2nd"} for x in ncols):

            t["rows"] = rows

            continue

        idx = None

        for i, c in enumerate(ncols):

            if "to hop coc" in c or c == "to hop" or "pile combination" in c:

                idx = i

                break

        if idx is None:

            t["rows"] = rows

            continue



        has_any_d = any(x in {"d1", "đ1", "1st", "d2", "đ2", "2nd", "d3", "đ3", "3rd", "d4", "đ4", "4th", "d5", "đ5", "5th", "d6", "đ6", "6th"} for x in ncols)

        if has_any_d:

            t["rows"] = rows

            continue



        # Lấy số lượng cột cần tách theo dữ liệu thật, tối đa D6

        max_parts = 0

        for r in rows[:20]:

            if idx < len(r):

                parts = _extract_parts(r[idx])

                if len(parts) > max_parts:

                    max_parts = len(parts)

            if max_parts >= 6:

                break

        if max_parts < 2:

            # Nếu dữ liệu không đủ rõ thì vẫn thử xem có chuỗi số ở các ô kế bên không

            for r in rows[:20]:

                seq = []

                for j in range(idx, min(len(r), idx + 6)):

                    val = str(r[j] or "").strip()

                    if not val:

                        continue

                    if _looks_number(val):

                        seq.append(val)

                    else:

                        break

                if len(seq) > max_parts:

                    max_parts = len(seq)

                if max_parts >= 2:

                    break

        if max_parts < 2:

            t["rows"] = rows

            continue

        max_parts = min(6, max_parts)



        new_cols = cols[:]

        new_cols[idx] = "D1"

        for offset in range(1, max_parts):

            new_cols.insert(idx + offset, f"D{offset + 1}")



        new_rows = []

        for r in rows:

            rr = list(r)

            parts = _extract_parts(rr[idx] if idx < len(rr) else "")

            if not parts:

                parts = []

                for j in range(idx, min(len(rr), idx + max_parts)):

                    val = str(rr[j] or "").strip()

                    if val and _looks_number(val):

                        parts.append(val)

                    elif j > idx:

                        break

            parts = parts[:max_parts]

            rr[idx] = parts[0] if len(parts) >= 1 else ""

            for offset in range(1, max_parts):

                insert_at = idx + offset

                rr.insert(insert_at, parts[offset] if offset < len(parts) else "")

            if len(rr) < len(new_cols):

                rr += [""] * (len(new_cols) - len(rr))

            new_rows.append(rr[:len(new_cols)])



        t["columns"] = new_cols

        t["rows"] = new_rows

        t["title"] = t.get("title") or f"Bảng đã tách Tổ hợp cọc D1-D{max_parts}"

    return tables


def merge_ocr_tables_for_continuous_read(tables):

    """

    Gộp các bảng OCR cùng cấu trúc thành một bảng liên tục để preview, mapping,

    ghi Excel và tổng hợp ngày đều dùng đủ dữ liệu đã đọc.

    """

    if not tables:

        return tables

    def _clean_col(name):

        try:

            return norm(name)

        except Exception:

            return str(name or "").strip().lower()

    def _signature(columns):

        return tuple(_clean_col(c) for c in (columns or []))

    def _is_stt_col(name):

        n = _clean_col(name)

        return n in {"stt", "no", "so thu tu", "tt", "no."} or n.startswith("stt ")

    groups = []

    passthrough = []

    for table in tables or []:

        if not isinstance(table, dict):

            passthrough.append(table)

            continue

        cols = list(table.get("columns") or [])

        rows = [list(r) if isinstance(r, (list, tuple)) else [r] for r in (table.get("rows") or [])]

        if not cols or not rows:

            passthrough.append(table)

            continue

        # Bảng key-value thường là phần thông tin phiếu, không phải dòng dữ liệu.

        if len(cols) <= 2 and {_clean_col(c) for c in cols} <= {"truong", "gia tri", "field", "value"}:

            passthrough.append(table)

            continue

        sig = _signature(cols)

        found = None

        for group in groups:

            if group["signature"] == sig:

                found = group

                break

        if found is None:

            found = {

                "signature": sig,

                "columns": cols,

                "rows": [],

                "titles": [],

                "row_source_indexes": [],

                "row_bboxes": [],

                "cell_bboxes": [],

            }

            groups.append(found)

        title = str(table.get("title") or "").strip()

        if title:

            found["titles"].append(title)

        width = len(found["columns"])
        row_source_indexes = list(table.get("_row_source_indexes") or [])
        table_source_index = table.get("_source_image_index")
        row_bboxes = list(table.get("_row_bboxes") or [])
        cell_bboxes = list(table.get("_cell_bboxes") or [])

        for row_idx, row in enumerate(rows):

            rr = list(row)

            if len(rr) < width:

                rr += [""] * (width - len(rr))

            found["rows"].append(rr[:width])
            source_index = (
                row_source_indexes[row_idx]
                if row_idx < len(row_source_indexes)
                else table_source_index
            )
            found["row_source_indexes"].append(source_index)
            found["row_bboxes"].append(
                row_bboxes[row_idx] if row_idx < len(row_bboxes) else None
            )
            found["cell_bboxes"].append(
                cell_bboxes[row_idx] if row_idx < len(cell_bboxes) else []
            )

    merged = []

    for group in groups:

        cols = list(group["columns"])

        rows = [list(r) for r in group["rows"]]

        stt_idx = next((i for i, name in enumerate(cols) if _is_stt_col(name)), None)

        if stt_idx is not None:

            next_no = 1

            for row in rows:

                if any(str(v or "").strip() for i, v in enumerate(row) if i != stt_idx):

                    row[stt_idx] = str(next_no)

                    next_no += 1

                else:

                    row[stt_idx] = ""

        title = "Bảng dữ liệu đã gộp"

        if len(group["titles"]) == 1:

            title = group["titles"][0]

        elif group["titles"]:

            title = f"Bảng dữ liệu đã gộp ({len(group['titles'])} phần)"

        merged.append(
            {
                "title": title,
                "columns": cols,
                "rows": rows,
                "_row_source_indexes": list(group["row_source_indexes"]),
                "_row_bboxes": list(group["row_bboxes"]),
                "_cell_bboxes": list(group["cell_bboxes"]),
            }
        )

    merged.sort(key=lambda t: (len(t.get("columns") or []), len(t.get("rows") or [])), reverse=True)

    return merged + passthrough


def _v229_apply_rows_to_workbook(self, wb):

    """

    Logic chốt:

    1. Tìm dòng TỔNG chính.

    2. Insert đúng số dòng mới ngay trước dòng TỔNG.

    3. Dòng TỔNG + toàn bộ phần sau TỔNG + merged ranges được đẩy xuống.

    4. Điền dữ liệu vào đúng dòng vừa insert.

    5. Copy style/công thức dòng mẫu; SUM lại đúng cột cần SUM theo file mẫu.

    """

    if not self.sheet_var.get():

        raise ValueError("Bạn chưa chọn sheet.")

    table = self.table_editor.get_current_table()

    if not table:

        raise ValueError("Chưa có dữ liệu từ ảnh.")



    # Đảm bảo bảng nguồn đã được xử lý Tổ hợp cọc

    fixed_tables = postprocess_to_hop_coc_d1_d2([table])

    table = fixed_tables[0] if fixed_tables else table



    ws = wb[self.sheet_var.get()]

    header_row = find_header_row_smart(ws)

    excel_headers = get_headers_smart(ws, header_row)

    self.header_row = header_row

    self.excel_headers = excel_headers



    total_row = find_total_row(ws, header_row)

    if not total_row:

        raise ValueError("Không tìm thấy dòng TỔNG/TOTAL trong Excel.")



    no_col = find_no_column_smart(ws, excel_headers, header_row, total_row)

    if not no_col:

        raise ValueError("Không tìm thấy cột STT/No trong Excel.")



    source_cols = list(table.get("columns", []))

    rows = _v229_filter_data_rows(table.get("rows", []))

    if not rows:

        raise ValueError("Không có dòng dữ liệu để nhập.")



    raw_mapping = None

    try:

        raw_mapping = self.mapping_editor.get_mapping()

    except Exception:

        raw_mapping = None

    if not raw_mapping or all(x is None for x in raw_mapping):

        raw_mapping = auto_mapping_to_excel_columns(source_cols, excel_headers)

        try:

            auto_idx = auto_map_columns(source_cols, excel_headers)

            auto_idx = ensure_no_column_in_mapping(source_cols, auto_idx, excel_headers)

            self.mapping_editor.set_mapping(source_cols, excel_headers, auto_idx)

        except Exception:

            pass



    mapping = _v229_normalize_mapping_to_excel_columns(source_cols, raw_mapping, excel_headers)

    if not mapping or all(x is None for x in mapping):

        raise ValueError("Chưa có mapping cột hoặc mapping đang bỏ qua toàn bộ cột.")



    first_data_row, style_row, last_no = _v229_find_stt_before_total(ws, no_col, header_row, total_row)

    sum_columns = _v229_capture_sum_columns(ws, total_row, first_data_row, total_row - 1, no_col)
    table_max_col = _v229_table_max_col(no_col, mapping, sum_columns, excel_headers)



    insert_at = total_row

    row_count = len(rows)



    # Chèn dòng và dịch merged ranges để giữ nguyên form phía dưới

    _v229_merged_ranges_shift_for_insert(ws, insert_at, row_count)

    total_row_after = total_row + row_count

    target_rows = list(range(insert_at, insert_at + row_count))
    effective_first_data_row = min(first_data_row, min(target_rows))
    data_last_row = max(target_rows)



    for i, data_row in enumerate(rows):

        dst_row = target_rows[i]

        _v229_safe_copy_row(ws, style_row, dst_row, table_max_col)



        # STT tự nối tiếp theo Excel, không lấy STT từ ảnh

        self._safe_set_cell_value(ws, dst_row, no_col, last_no + i + 1)



        for src_idx, excel_col in enumerate(mapping):

            if excel_col is None:

                continue

            if excel_col == no_col:

                continue

            # Nếu dòng mẫu đã copy công thức vào ô này thì giữ công thức

            try:

                if _v229_is_formula(ws.cell(dst_row, excel_col).value):

                    continue

            except Exception:

                pass



            val = data_row[src_idx] if src_idx < len(data_row) else ""

            self._safe_set_cell_value(ws, dst_row, excel_col, convert_excel_value(val))



    # SUM lại đúng các cột cần tổng theo file mẫu

    sum_first_row = first_data_row

    sum_last_row = target_rows[-1]

    for c in sorted(set(sum_columns)):

        if c == no_col:

            continue

        try:

            letter = get_column_letter(c)

            ws.cell(total_row_after, c).value = f"=SUM({letter}{sum_first_row}:{letter}{sum_last_row})"

        except Exception:

            pass



    updated_total_cols = []
    try:
        updated_total_cols = update_total_formulas(
            ws,
            total_row_after,
            effective_first_data_row,
            data_last_row,
            excel_headers=excel_headers,
            no_col=no_col,
        )
    except Exception:
        updated_total_cols = []

    cleared_non_total_formula_cols = _v231_clear_non_total_measure_formulas(
        ws,
        total_row_after,
        excel_headers=excel_headers,
        no_col=no_col,
    )

    force_workbook_recalculate(wb)



    out = last_run_dir()

    out.mkdir(exist_ok=True)

    logic = {

        "rule": "V22.9_FINAL_INSERT_BEFORE_TOTAL_KEEP_ALL_FEATURES",

        "sheet": ws.title,

        "header_row": header_row,

        "total_row_before": total_row,

        "insert_at": insert_at,

        "rows_added": row_count,

        "data_rows": target_rows,

        "total_row_after": total_row_after,

        "last_stt_before": last_no,

        "new_stt_start": last_no + 1,

        "new_stt_end": last_no + row_count,

        "sum_first_row": sum_first_row,

        "sum_last_row": sum_last_row,

        "sum_columns": sorted(set(sum_columns)),

        "mapping": [

            {

                "source": source_cols[i] if i < len(source_cols) else "",

                "excel_col": c,

                "excel_letter": get_column_letter(c) if c else None,

            }

            for i, c in enumerate(mapping)

        ],

    }

    (out / "v22_9_final_apply_logic.json").write_text(

        json.dumps(logic, ensure_ascii=False, indent=2),

        encoding="utf-8"

    )



    return {

        "sheet": ws.title,

        "header_row": header_row,

        "last_stt_before": last_no,

        "start_fill_row": insert_at,

        "next_stt_start": last_no + 1,

        "rows_added": row_count,

        "total_row_after": total_row_after,

        "sum_first_row": sum_first_row,

        "sum_last_row": sum_last_row,

        "sum_columns_count": len(set(sum_columns)),

        "logic_file": str(out / "v22_9_final_apply_logic.json"),

    }





def _v229_preview_excel(self):

    """Preview không ghi đè file cũ, tránh PermissionError khi Excel đang mở."""

    if not self.excel_path:

        messagebox.showwarning("Thiếu Excel", "Bạn chưa chọn file Excel.")

        return

    try:

        from datetime import datetime

        wb = load_workbook(self.excel_path)

        info = self._apply_rows_to_workbook(wb)



        out = last_run_dir()

        out.mkdir(exist_ok=True)

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        preview_path = out / f"preview_sau_khi_ghep_{stamp}.xlsx"

        force_workbook_recalculate(wb)

        wb.save(preview_path)



        try:

            os.startfile(str(preview_path))

        except Exception:

            pass



        self._set_status(f"Đã tạo file xem trước: {preview_path}", "success")

        messagebox.showinfo(

            "Đã tạo xem trước",

            f"Đã tạo file preview để kiểm tra trước khi lưu thật.\n\n"

            f"Sheet: {info['sheet']}\n"

            f"Bắt đầu chèn từ dòng: {info['start_fill_row']}\n"

            f"Số dòng thêm: {info['rows_added']}\n"

            f"Dòng TỔNG sau khi đẩy xuống: {info['total_row_after']}\n\n"

            f"File preview:\n{preview_path}\n\n"

            f"Nếu đúng thì quay lại tool bấm: Điền tiếp vào Excel."

        )

    except Exception as exc:

        error_path, _log_path = report_runtime_error_to_admin(
            "preview_excel",
            exc,
            {
                "excel_path": self.excel_path,
                "sheet": self.sheet_var.get() if hasattr(self, "sheet_var") else "",
            },
            error_file_name="last_error_preview.txt",
            notify_server=True,
        )

        detail = str(exc).strip() or exc.__class__.__name__

        messagebox.showerror(
            "Lỗi xem trước",
            f"Xem trước Excel bị lỗi:\n{detail}\n\nĐã ghi log và gửi cho Admin nếu server đang bật.\nFile lỗi: {error_path or 'last_run_v12/last_error_preview.txt'}",
        )

        self._set_status("Lỗi xem trước Excel, đã ghi log.", "error")





# Ghi đè nhẹ, không xoá chức năng cũ.
# =========================

# V23 STRICT - DỮ LIỆU THEO ẢNH + TỔNG THEO FILE MẪU

# =========================



def _v23_cell_to_text(v):

    if v is None:

        return ""

    if isinstance(v, float):

        if v.is_integer():

            return str(int(v))

        s = ("%s" % v).rstrip("0").rstrip(".")

        return s.replace(".", ",")

    return str(v).strip()





def _v23_norm_compare(v):

    s = _v23_cell_to_text(v).strip()

    s = s.replace(" ", "")

    s = s.replace(".", ",")

    # +1,5 trong ảnh và 1,5 trong Excel coi là cùng dữ liệu số dương

    if s.startswith("+"):

        s = s[1:]

    return s.lower()





def _v23_source_should_be_text(src_name):

    n = norm(src_name)

    return any(x in n for x in [

        "ngay", "date", "gio", "time", "bat dau", "ket thuc",

        "ten", "coc", "pile", "loai", "type", "vi tri", "location",

        "ghi chu", "note", "remark"

    ])





def _v23_convert_by_source(src_name, value):

    """

    Giữ nguyên dữ liệu dễ sai format; chỉ chuyển số cho cột cần tính toán.

    """

    s = str(value or "").strip()

    if s == "":

        return ""

    n = norm(src_name)

    # các cột này phải giữ đúng như ảnh

    if any(x in n for x in ["ngay", "date", "gio", "time", "bat dau", "ket thuc", "ten", "pile", "loai", "type", "vi tri", "ghi chu", "note", "remark"]):

        if any(x in n for x in ["ngay", "date"]):

            return normalize_vietnam_date(s)

        return s

    if any(x in n for x in ["khoi luong", "kl", "weight", "luc ep", "load", "chieu dai", "length", "do dai", "so luong", "quantity", "d1", "d2", "d3", "d4", "d5", "d6"]):

        return normalize_numeric_like_text(s)

    return convert_excel_value(s)





def _v23_validate_written_cells(ws, target_rows, rows, source_cols, mapping, no_col):

    """

    Log kiểm tra: ô nào ghi trực tiếp thì so lại với dữ liệu preview.

    Không chặn save vì Excel có thể định dạng số khác ảnh, nhưng log rõ để kiểm tra.

    """

    report = []

    for r_i, dst_row in enumerate(target_rows):

        data_row = rows[r_i]

        for src_idx, excel_col in enumerate(mapping):

            if excel_col is None or excel_col == no_col:

                continue

            src = source_cols[src_idx] if src_idx < len(source_cols) else ""

            img_val = data_row[src_idx] if src_idx < len(data_row) else ""

            cell = ws.cell(dst_row, excel_col)

            if _v229_is_formula(cell.value):

                status = "FORMULA_GIU_THEO_EXCEL_MAU"

            else:

                ex_val = cell.value

                status = "OK" if _v23_norm_compare(img_val) == _v23_norm_compare(ex_val) else "CHECK_FORMAT_OR_VALUE"

            report.append({

                "row": dst_row,

                "cell": f"{get_column_letter(excel_col)}{dst_row}",

                "source": src,

                "image_value": img_val,

                "excel_value": cell.value,

                "status": status,

            })

    return report





def _v23_apply_rows_to_workbook(self, wb):

    """

    V23 chốt:

    - Dữ liệu nguồn giữ theo ảnh ở preview.

    - Khi xuất Excel: chỉ ghi vào dòng mới insert trước TỔNG.

    - Dòng TỔNG và phần sau TỔNG đẩy xuống.

    - Cột có công thức ở dòng dữ liệu giữ công thức mẫu.

    - Dòng TỔNG chỉ SUM các cột mẫu đang SUM/cần tổng; không SUM STT.

    - Có log kiểm tra từng ô ghi trực tiếp.

    """

    if not self.sheet_var.get():

        raise ValueError("Bạn chưa chọn sheet.")

    table = self.table_editor.get_current_table()

    if not table:

        raise ValueError("Chưa có dữ liệu từ ảnh.")



    fixed_tables = postprocess_to_hop_coc_d1_d2([table])

    table = fixed_tables[0] if fixed_tables else table



    ws = wb[self.sheet_var.get()]

    header_row = find_header_row_smart(ws)

    excel_headers = get_headers_smart(ws, header_row)

    self.header_row = header_row

    self.excel_headers = excel_headers



    total_row = find_total_row(ws, header_row)

    if not total_row:

        raise ValueError("Không tìm thấy dòng TỔNG/TOTAL trong Excel.")



    no_col = find_no_column_smart(ws, excel_headers, header_row, total_row)

    if not no_col:

        raise ValueError("Không tìm thấy cột STT/No trong Excel.")



    source_cols = list(table.get("columns", []))

    rows = _v229_filter_data_rows(table.get("rows", []))

    if not rows:

        raise ValueError("Không có dòng dữ liệu để nhập.")



    raw_mapping = None

    try:

        raw_mapping = self.mapping_editor.get_mapping()

    except Exception:

        raw_mapping = None

    if not raw_mapping or all(x is None for x in raw_mapping):

        raw_mapping = auto_mapping_to_excel_columns(source_cols, excel_headers)

        try:

            auto_idx = auto_map_columns(source_cols, excel_headers)

            auto_idx = ensure_no_column_in_mapping(source_cols, auto_idx, excel_headers)

            self.mapping_editor.set_mapping(source_cols, excel_headers, auto_idx)

        except Exception:

            pass



    mapping = _v229_normalize_mapping_to_excel_columns(source_cols, raw_mapping, excel_headers)

    if not mapping or all(x is None for x in mapping):

        raise ValueError("Chưa có mapping cột hoặc mapping đang bỏ qua toàn bộ cột.")



    # Check cột quan trọng không bị bỏ qua hết

    mapped_names = {norm(source_cols[i]) for i, c in enumerate(mapping) if c and i < len(source_cols)}

    critical_any = ["d1", "d2", "loai coc", "ten tim coc", "ten coc"]

    # Không chặn cứng mọi form, nhưng log cảnh báo nếu thiếu

    warnings = []

    for key in critical_any:

        if not any(key in m or m in key for m in mapped_names):

            warnings.append(f"Có thể chưa map cột: {key}")



    first_data_row, style_row, last_no = _v229_find_stt_before_total(ws, no_col, header_row, total_row)

    sum_columns = _v229_capture_sum_columns(ws, total_row, first_data_row, total_row - 1, no_col)
    table_max_col = _v229_table_max_col(no_col, mapping, sum_columns, excel_headers)



    insert_at = total_row

    row_count = len(rows)



    _v229_merged_ranges_shift_for_insert(ws, insert_at, row_count)

    total_row_after = total_row + row_count

    target_rows = list(range(insert_at, insert_at + row_count))



    for i, data_row in enumerate(rows):

        dst_row = target_rows[i]

        _v229_safe_copy_row(ws, style_row, dst_row, table_max_col)



        self._safe_set_cell_value(ws, dst_row, no_col, last_no + i + 1)



        for src_idx, excel_col in enumerate(mapping):

            if excel_col is None or excel_col == no_col:

                continue

            try:

                if _v229_is_formula(ws.cell(dst_row, excel_col).value):

                    continue

            except Exception:

                pass

            src_name = source_cols[src_idx] if src_idx < len(source_cols) else ""

            val = data_row[src_idx] if src_idx < len(data_row) else ""

            self._safe_set_cell_value(ws, dst_row, excel_col, _v23_convert_by_source(src_name, val))



    sum_first_row = first_data_row

    sum_last_row = target_rows[-1]

    for c in sorted(set(sum_columns)):

        if c == no_col:

            continue

        try:

            letter = get_column_letter(c)

            ws.cell(total_row_after, c).value = f"=SUM({letter}{sum_first_row}:{letter}{sum_last_row})"

        except Exception:

            pass



    force_workbook_recalculate(wb)

    try:

        learned_profile = _v231_capture_formula_profile_from_sheet(

            ws,

            header_row,

            total_row_after,

            no_col,

            source_name=self.excel_path,

        )

        _v231_store_formula_profile(learned_profile)

    except Exception:

        pass



    validation = _v23_validate_written_cells(ws, target_rows, rows, source_cols, mapping, no_col)



    out = last_run_dir()

    out.mkdir(exist_ok=True)

    logic = {

        "rule": "V23_CHUAN_DU_LIEU_TONG_CHINH_XAC",

        "sheet": ws.title,

        "header_row": header_row,

        "total_row_before": total_row,

        "insert_at": insert_at,

        "rows_added": row_count,

        "data_rows": target_rows,

        "total_row_after": total_row_after,

        "last_stt_before": last_no,

        "new_stt_start": last_no + 1,

        "new_stt_end": last_no + row_count,

        "sum_first_row": sum_first_row,

        "sum_last_row": sum_last_row,

        "sum_columns": sorted(set(sum_columns)),

        "warnings": warnings,

        "mapping": [

            {

                "source": source_cols[i] if i < len(source_cols) else "",

                "excel_col": c,

                "excel_letter": get_column_letter(c) if c else None,

            }

            for i, c in enumerate(mapping)

        ],

        "written_cell_validation": validation,

    }

    (out / "v23_chuan_du_lieu_tong_logic.json").write_text(

        json.dumps(logic, ensure_ascii=False, indent=2),

        encoding="utf-8"

    )



    return {

        "sheet": ws.title,

        "header_row": header_row,

        "last_stt_before": last_no,

        "start_fill_row": insert_at,

        "next_stt_start": last_no + 1,

        "rows_added": row_count,

        "total_row_after": total_row_after,

        "sum_first_row": sum_first_row,

        "sum_last_row": sum_last_row,

        "sum_columns_count": len(set(sum_columns)),

        "logic_file": str(out / "v23_chuan_du_lieu_tong_logic.json"),

    }





def _v23_run_gemini(self):

    if getattr(self, "_is_reading_table", False):
        try:
            self._set_status("Đang đọc bảng, vui lòng chờ tác vụ hiện tại hoàn tất.", "warn")
        except Exception:
            pass
        return

    image_paths = list(getattr(self, "image_paths", None) or ([] if not self.image_path else [self.image_path]))

    if not image_paths:

        messagebox.showwarning("Thiếu ảnh", "Bạn chưa chọn ảnh.")

        return

    api_key = self.api_key_var.get().strip()

    if not api_key:

        messagebox.showwarning("Thiếu khóa API", "Bạn chưa nhập khóa API.")

        return

    self.save_key()
    started_at = datetime.now()
    model_name = self.model_var.get().strip()
    self._is_reading_table = True

    try:

        self._set_status(f"Đang đọc bảng... ({len(image_paths)} ảnh)", "warn")

    except Exception:

        pass

    def _ui_alive():
        try:
            return bool(self.root.winfo_exists())
        except Exception:
            return False

    def _status(text, tone="warn"):
        try:
            self.root.after(0, lambda: self._set_status(text, tone) if _ui_alive() else None)
        except Exception:
            pass

    def _finish_success(tables, raw):
        if not _ui_alive():
            return
        try:
            out = last_run_dir()
            out.mkdir(exist_ok=True)
            (out / "ai_raw_response.txt").write_text(raw, encoding="utf-8")
            (out / "ai_tables.json").write_text(json.dumps(tables, ensure_ascii=False, indent=2), encoding="utf-8")

            self.tables = tables
            self.table_editor.set_tables(tables)
            self._refresh_daily_summary_panel(tables)
            self.current_doc_kind = "bang_khoi_luong"

            if self.excel_headers and tables:
                self.build_mapping()

            total_rows = sum(len(t.get("rows", [])) for t in tables if isinstance(t, dict))
            self._set_status(
                f"Đọc xong: {len(image_paths)} ảnh, {len(tables)} bảng sau gộp, {total_rows} dòng. Kiểm tra preview trước khi xuất.",
                "success",
            )
            ended_at = datetime.now()
            elapsed_text = _format_elapsed((ended_at - started_at).total_seconds())
            _write_ocr_timing_log(
                "ocr_timing_bang.json",
                "read_table",
                started_at,
                ended_at,
                len(image_paths),
                "success",
                f"Đọc bảng xong: {len(tables)} bảng, {total_rows} dòng.",
            )
            _show_ocr_done_notification(
                self,
                "Đọc bảng xong",
                f"Đã đọc xong {len(image_paths)} ảnh.\n"
                f"Kết quả: {len(tables)} bảng, {total_rows} dòng.\n"
                f"Thời gian: {elapsed_text}.",
            )
        finally:
            self._is_reading_table = False

    def _finish_error(error_text):
        if not _ui_alive():
            return
        try:
            out = last_run_dir()
            out.mkdir(exist_ok=True)
            (out / "last_error.txt").write_text(error_text, encoding="utf-8")
            ended_at = datetime.now()
            elapsed_text = _format_elapsed((ended_at - started_at).total_seconds())
            _write_ocr_timing_log(
                "ocr_timing_bang.json",
                "read_table",
                started_at,
                ended_at,
                len(image_paths),
                "error",
                "Lỗi đọc bảng. Xem last_run_v12/last_error.txt",
            )
            _show_ocr_done_notification(
                self,
                "Lỗi đọc bảng",
                "Có lỗi khi đọc bảng.\n"
                f"Thời gian: {elapsed_text}.\n"
                "Xem last_run_v12\\last_error.txt",
            )
            self._set_status("Lỗi đọc ảnh.", "error")
        finally:
            self._is_reading_table = False

    def _worker():
        try:
            all_tables = []
            raw_parts = []

            for idx, image_path in enumerate(image_paths, start=1):
                _status(f"Đang đọc ảnh {idx}/{len(image_paths)}...", "warn")
                tables_one, raw_one = call_gemini(image_path, api_key, model_name)

                for table in tables_one or []:
                    ensure_table_image_metadata(table, idx - 1, image_path)

                tables_one = postprocess_to_hop_coc_d1_d2(tables_one)

                for table in tables_one or []:
                    ensure_table_image_metadata(table, idx - 1, image_path)
                all_tables.extend(tables_one or [])
                raw_parts.append(f"=== IMAGE {idx}/{len(image_paths)}: {Path(image_path).name} ===\n{raw_one}")

            tables = merge_ocr_tables_for_continuous_read(all_tables)
            for table in tables:
                if isinstance(table, dict):
                    table["_source_image_count"] = len(image_paths)

            raw = "\n\n".join(raw_parts)
            try:
                self.root.after(0, lambda: _finish_success(tables, raw))
            except Exception:
                pass
        except Exception:
            error_text = traceback.format_exc()
            try:
                self.root.after(0, lambda: _finish_error(error_text))
            except Exception:
                pass

    threading.Thread(target=_worker, daemon=True).start()





# Override cuối cùng cho bản V23
# =========================

# V23.1 FINAL SUM FIX - SUM HẾT DỮ LIỆU TRƯỚC DÒNG TỔNG

# =========================



def _v231_parse_sum_first_row(formula, col_letter):

    """Lấy dòng bắt đầu từ công thức SUM mẫu ở dòng TỔNG, ví dụ =SUM(K9:K28) -> 9."""

    s = str(formula or "").replace(" ", "")

    if not s.startswith("="):

        return None

    import re

    col = str(col_letter).upper()

    m = re.search(r"SUM\(\$?" + re.escape(col) + r"\$?(\d+):\$?" + re.escape(col) + r"\$?\d+\)", s, flags=re.I)

    if m:

        try:

            return int(m.group(1))

        except Exception:

            return None

    m = re.search(r"\$?" + re.escape(col) + r"\$?(\d+):\$?" + re.escape(col) + r"\$?\d+", s, flags=re.I)

    if m:

        try:

            return int(m.group(1))

        except Exception:

            return None

    return None





def _v231_normalize_header_signature(headers):

    sig = []

    for col_idx, name in headers or []:

        n = norm(name)

        if n:

            sig.append({"col": int(col_idx), "name": str(name or ""), "norm": n})

    return sig



def _v231_profile_similarity(profile, headers, sheet_title=""):

    """

    Chấm điểm mức độ giống nhau giữa workbook hiện tại và profile công thức cũ.
    """

    prof_headers = profile.get("headers") or []

    current_headers = _v231_normalize_header_signature(headers)

    if not prof_headers or not current_headers:

        return 0

    prof_by_norm = {item.get("norm"): item for item in prof_headers if item.get("norm")}

    current_by_norm = {item.get("norm"): item for item in current_headers if item.get("norm")}

    shared_norms = set(prof_by_norm) & set(current_by_norm)

    if not shared_norms:

        return 0

    pos_score = 0

    for n in shared_norms:

        p_col = int(prof_by_norm[n].get("col") or 0)

        c_col = int(current_by_norm[n].get("col") or 0)

        if p_col and c_col:

            if p_col == c_col:

                pos_score += 5

            elif abs(p_col - c_col) <= 2:

                pos_score += 3

            else:

                pos_score += 1

    overlap_score = len(shared_norms) * 10

    order_bonus = 0

    prof_order = [item.get("norm") for item in prof_headers[:12] if item.get("norm")]

    curr_order = [item.get("norm") for item in current_headers[:12] if item.get("norm")]

    for i, n in enumerate(prof_order[:len(curr_order)]):

        if i < len(curr_order) and curr_order[i] == n:

            order_bonus += 4

    title_bonus = 0

    prof_title = norm(profile.get("sheet_title") or "")

    curr_title = norm(sheet_title)

    if prof_title and curr_title and (prof_title in curr_title or curr_title in prof_title):

        title_bonus += 10

    return overlap_score + pos_score + order_bonus + title_bonus



def _v231_find_best_formula_profile(ws, header_row, total_row, excel_headers):

    profiles = load_formula_profiles()

    if not profiles:

        return None

    sheet_title = getattr(ws, "title", "") or ""

    best = None

    best_score = 0

    for profile in profiles:

        try:

            score = _v231_profile_similarity(profile, excel_headers, sheet_title)

        except Exception:

            score = 0

        if score > best_score:

            best_score = score

            best = profile

    if best_score >= 18:

        return best

    return None



def _v231_capture_formula_profile_from_sheet(ws, header_row, total_row, no_col, source_name=""):

    headers = get_headers_smart(ws, header_row)

    header_by_col = {c: name for c, name in headers}

    first_data_row = _v231_detect_first_data_row_strict(ws, header_row, total_row, no_col)

    total_formula_cols = capture_formula_columns(ws, total_row)

    items = []

    for c in total_formula_cols:

        if c == no_col:

            continue

        letter = get_column_letter(c)

        formula = ws.cell(total_row, c).value

        start_row = _v231_parse_sum_first_row(formula, letter) or first_data_row

        items.append({

            "col": c,

            "letter": letter,

            "header": header_by_col.get(c, ""),

            "header_norm": norm(header_by_col.get(c, "")),

            "start_offset": max(0, int(start_row) - int(first_data_row)),

            "formula": str(formula or ""),

        })

    return {

        "id": uuid.uuid4().hex[:12].upper(),

        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),

        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),

        "source_name": str(source_name or ""),

        "sheet_title": ws.title,

        "sheet_title_norm": norm(ws.title),

        "header_row": int(header_row or 0),

        "total_row": int(total_row or 0),

        "no_col": int(no_col or 0) if no_col else 0,

        "headers": _v231_normalize_header_signature(headers),

        "sum_columns": items,

    }



def _v231_store_formula_profile(profile):

    if not profile:

        return

    profiles = load_formula_profiles()

    try:

        target_id = profile.get("id")

        if target_id:

            existing_idx = next((i for i, item in enumerate(profiles) if str(item.get("id") or "") == str(target_id)), None)

            if existing_idx is not None:

                profiles[existing_idx] = profile

            else:

                profiles.insert(0, profile)

        else:

            profiles.insert(0, profile)

        profiles = profiles[:120]

        save_formula_profiles(profiles)

    except Exception:

        pass



def _v231_collect_stt_rows(ws, col, header_row, total_row):

    memo = {}

    rows = []

    for r in range(header_row + 1, total_row):

        try:

            row_text = " ".join(str(ws.cell(r, c).value or "") for c in range(1, min(ws.max_column, 10) + 1))

            if is_total_marker_text(row_text):

                continue

        except Exception:

            pass

        n = None

        try:

            n = get_stt_value(ws, r, col, memo)

        except Exception:

            n = None

        if not isinstance(n, int):

            s = str(ws.cell(r, col).value or "").strip()

            if s.isdigit():

                n = int(s)

            else:

                try:

                    f = float(s.replace(",", "."))

                    if f.is_integer():

                        n = int(f)

                except Exception:

                    pass

        if isinstance(n, int):

            rows.append((r, n))

    return rows



def _v231_pick_stt_context(ws, header_row, total_row, excel_headers, preferred_no_col=None):

    """

    Chọn cột STT và dãy STT tin cậy nhất.

    Dùng khi file mới có nhiều dòng trắng hoặc header bị đọc lệch.
    """

    candidates = []

    if preferred_no_col:

        candidates.append(int(preferred_no_col))

    for col_idx, name in excel_headers or []:

        if is_no_header(name) and col_idx not in candidates:

            candidates.append(int(col_idx))

    for c in range(1, min(ws.max_column, 8) + 1):

        if c not in candidates:

            candidates.append(c)

    best = None

    best_score = -1

    for c in candidates:

        rows = _v231_collect_stt_rows(ws, c, header_row, total_row)

        if not rows:

            continue

        rows.sort(key=lambda x: x[0])

        numeric_count = len(rows)

        longest_run = 1

        current_run = 1

        increasing = 0

        for prev, cur in zip(rows, rows[1:]):

            prev_row, prev_val = prev

            cur_row, cur_val = cur

            if cur_val > prev_val:

                increasing += 1

            if 0 < (cur_row - prev_row) <= 8 and 0 < (cur_val - prev_val) <= 8:

                current_run += 1

                longest_run = max(longest_run, current_run)

            else:

                current_run = 1

        first_row = rows[0][0]

        last_row = rows[-1][0]

        last_no = rows[-1][1]

        score = numeric_count * 12 + longest_run * 20 + increasing * 3

        if first_row <= header_row + 3:

            score += 8

        if rows[0][1] in (0, 1):

            score += 8

        if last_no >= numeric_count:

            score += 4

        if score > best_score:

            best_score = score

            best = {

                "no_col": c,

                "first_row": first_row,

                "last_row": last_row,

                "last_no": last_no,

                "rows": rows,

                "score": score,

            }

    if best:

        return best

    return {

        "no_col": preferred_no_col or 1,

        "first_row": header_row + 1,

        "last_row": max(header_row + 1, total_row - 1),

        "last_no": 0,

        "rows": [],

        "score": 0,

    }



def _v231_detect_first_data_row_strict(ws, header_row, total_row, no_col):

    """Tìm dòng dữ liệu đầu tiên thật sự trước TỔNG, ưu tiên STT số/công thức STT."""

    memo = {}

    for r in range(header_row + 1, total_row):

        try:

            row_text = " ".join(str(ws.cell(r, c).value or "") for c in range(1, ws.max_column + 1))

            if is_total_marker_text(row_text):

                continue

        except Exception:

            pass

        n = None

        try:

            n = get_stt_value(ws, r, no_col, memo)

        except Exception:

            n = None

        if isinstance(n, int):

            return r

        s = str(ws.cell(r, no_col).value or "").strip()

        if s.isdigit():

            return r

    return header_row + 1





def _v231_capture_sum_columns_and_starts(ws, total_row, default_first_row, no_col, excel_headers=None, formula_profile=None):

    """

    Chốt cột cần SUM và dòng bắt đầu SUM.

    - Nếu dòng TỔNG có công thức SUM: lấy dòng bắt đầu theo công thức mẫu.

    - Nếu dòng TỔNG là số: SUM từ dòng dữ liệu đầu tiên.

    - Không SUM cột STT.

    """

    result = {}

    header_by_norm = {}

    header_by_col = {}

    for col_idx, name in (excel_headers or []):

        n = norm(name)

        if n and n not in header_by_norm:

            header_by_norm[n] = col_idx

        header_by_col[int(col_idx)] = str(name or "")

    def _is_sum_excluded_header(col_idx):
        n = norm(header_by_col.get(int(col_idx), ""))
        if not n:
            return False
        if n in {"1st", "2nd", "3rd", "4th", "5th", "6th", "d1", "d2", "d3", "d4", "d5", "d6"}:
            return True
        if any(tok in n for tok in ["luc ep", "force", "pressing force"]):
            return True
        excluded_tokens = {
            "stt", "so tt", "no",
            "ngay", "ngay thi cong", "gio", "thoi gian", "bat dau", "ket thuc",
            "ten", "ten may", "ten coc", "vi tri", "ghi chu", "noi dung",
            "loai", "ma", "ma coc", "so hieu", "ca",
            "hop dong", "contract", "contract no", "so hop dong", "hd",
        }
        return any(tok in n for tok in excluded_tokens)

    def _is_sum_allowed_header(col_idx):
        n = norm(header_by_col.get(int(col_idx), ""))
        if not n:
            return None
        if _is_sum_excluded_header(col_idx):
            return False
        allowed_tokens = {
            "tong to hop", "tong chieu dai", "chieu dai to hop",
            "do sau ep", "ep am", "ep duong", "ep am duong",
        }
        return any(tok in n for tok in allowed_tokens)

    def _is_contract_like_header(col_idx):
        n = norm(header_by_col.get(int(col_idx), ""))
        return any(tok in n for tok in ["hop dong", "contract", "contract no", "so hop dong", "hd"])

    def _looks_like_text_or_time_header(name):

        n = norm(name)

        if not n:

            return False

        bad_tokens = {

            "ngay", "gio", "bat dau", "ket thuc", "thoi gian", "ghi chu",

            "noi dung", "ten", "loai", "ma", "stt", "no", "ca",

        }

        return any(tok in n for tok in bad_tokens)

    def _column_has_enough_numbers(col_idx):

        numeric_hits = 0

        text_hits = 0

        for rr in range(default_first_row, total_row):

            vv = ws.cell(rr, col_idx).value

            if vv is None or str(vv).strip() == "":

                continue

            if isinstance(vv, (int, float)):

                numeric_hits += 1

                continue

            svv = str(vv).strip().replace(".", "").replace(",", ".")

            try:

                float(svv)

                numeric_hits += 1

            except Exception:

                text_hits += 1

        return numeric_hits, text_hits

    def _best_start_row_for_col(col_idx):

        for rr in range(default_first_row, total_row):

            vv = ws.cell(rr, col_idx).value

            if isinstance(vv, (int, float)):

                return rr

            svv = str(vv or "").strip().replace(".", "").replace(",", ".")

            try:

                if svv != "":

                    float(svv)

                    return rr

            except Exception:

                pass

        return default_first_row

    for c in range(1, ws.max_column + 1):

        if c == no_col:

            continue

        allowed_by_header = _is_sum_allowed_header(c)
        if allowed_by_header is False:
            continue

        if _is_contract_like_header(c):
            continue

        letter = get_column_letter(c)

        v = ws.cell(total_row, c).value

        if _v229_is_formula(v):

            numeric_hits, text_hits = _column_has_enough_numbers(c)
            if allowed_by_header is None and (numeric_hits <= 0 or text_hits > numeric_hits):
                continue

            parsed_first = _v231_parse_sum_first_row(v, letter)

            if parsed_first:

                first = min(parsed_first, default_first_row)

            else:

                first = default_first_row

            result[c] = first

            continue

        is_num_total = False

        if isinstance(v, (int, float)):

            is_num_total = True

        else:

            sv = str(v or "").strip().replace(".", "").replace(",", ".")

            try:

                if sv != "":

                    float(sv)

                    is_num_total = True

            except Exception:

                is_num_total = False

        if not is_num_total:

            continue

        if allowed_by_header is None:
            numeric_hits, text_hits = _column_has_enough_numbers(c)
            if numeric_hits <= 0 or text_hits > numeric_hits:
                continue

        # Có số phía trên thì đây là cột tổng cần SUM.

        for rr in range(default_first_row, total_row):

            vv = ws.cell(rr, c).value

            if isinstance(vv, (int, float)):

                result[c] = default_first_row

                break

            svv = str(vv or "").strip().replace(".", "").replace(",", ".")

            try:

                if svv != "":

                    float(svv)

                    result[c] = default_first_row

                    break

            except Exception:

                pass

    # Fallback: nếu dòng TỔNG trống, suy ra các cột số từ dữ liệu bên dưới.
    if not result:

        for c in range(1, ws.max_column + 1):

            if c == no_col:

                continue

            header_name = header_by_col.get(c, "")

            allowed_by_header = _is_sum_allowed_header(c)
            if allowed_by_header is False:
                continue

            if _is_contract_like_header(c):
                continue

            if _looks_like_text_or_time_header(header_name):

                continue

            numeric_hits, text_hits = _column_has_enough_numbers(c)

            if numeric_hits >= 2 and numeric_hits >= text_hits and (allowed_by_header is True or not header_name):

                result[c] = _best_start_row_for_col(c)

    # Nếu sheet mới chưa có công thức tổng rõ ràng, học từ profile file cũ.
    if formula_profile and isinstance(formula_profile, dict):

        for item in formula_profile.get("sum_columns") or []:

            try:

                header_norm = norm(item.get("header_norm") or item.get("header") or "")

                if not header_norm:

                    continue

                c = header_by_norm.get(header_norm)

                if not c or c == no_col:

                    continue

                if _is_sum_allowed_header(c) is False:
                    continue

                if _is_contract_like_header(c):
                    continue

                if c not in result:

                    offset = int(item.get("start_offset") or 0)

                    result[c] = max(default_first_row, default_first_row + offset)

            except Exception:

                pass

    return result


def _v231_is_non_total_measure_header(name):
    n = norm(name)
    if not n:
        return False
    if re.search(r"(^|\s)(1st|2nd|3rd|4th|5th|6th|d1|d2|d3|d4|d5|d6)(\s|$)", n):
        return True
    return any(tok in n for tok in ["luc ep", "force", "pressing force"])


def _v231_clear_non_total_measure_formulas(ws, total_row, excel_headers=None, no_col=None):
    cleared = []
    for col_idx, name in (excel_headers or []):
        try:
            c = int(col_idx)
        except Exception:
            continue
        if c == no_col:
            continue
        if not _v231_is_non_total_measure_header(name):
            continue
        try:
            cell = ws.cell(total_row, c)
            if _v229_is_formula(cell.value):
                cell.value = None
                cleared.append(c)
        except Exception:
            pass
    return cleared


def _v231_apply_rows_to_workbook(self, wb):

    """

    V23.1 chốt SUM:

    - Insert dữ liệu mới ngay trước dòng TỔNG.

    - Dòng TỔNG và toàn bộ dòng sau TỔNG tự đẩy xuống.

    - Dòng TỔNG SUM từ dòng đầu dữ liệu theo công thức mẫu đến ngay trước dòng TỔNG mới.

    - Không dừng SUM ở dòng mới cuối; vì trước TỔNG có thể còn dòng trắng/form, vẫn phải SUM hết.

    """

    if not self.sheet_var.get():

        raise ValueError("Bạn chưa chọn sheet.")

    table = self.table_editor.get_current_table()

    if not table:

        raise ValueError("Chưa có dữ liệu từ ảnh.")



    fixed_tables = postprocess_to_hop_coc_d1_d2([table])

    table = fixed_tables[0] if fixed_tables else table



    ws = wb[self.sheet_var.get()]

    header_row = find_header_row_smart(ws)

    excel_headers = get_headers_smart(ws, header_row)

    self.header_row = header_row

    self.excel_headers = excel_headers



    total_row = find_total_row(ws, header_row)

    if not total_row:

        # Fallback: tìm dòng trống đầu tiên sau dữ liệu

        for r in range(header_row + 1, ws.max_row + 2):

            if not any(str(ws.cell(r, c).value or "").strip() for c in range(1, 4)):

                total_row = r

                break

        if not total_row:

            total_row = ws.max_row + 1



    no_col = find_no_column_smart(ws, excel_headers, header_row, total_row)

    if not no_col:

        no_col = 1 # Fallback cột A



    source_cols = list(table.get("columns", []))

    rows = _v229_filter_data_rows(table.get("rows", []))

    if not rows:

        raise ValueError("Không có dòng dữ liệu để nhập.")



    raw_mapping = None

    try:

        raw_mapping = self.mapping_editor.get_mapping()

    except Exception:

        raw_mapping = None

    if not raw_mapping or all(x is None for x in raw_mapping):

        raw_mapping = auto_mapping_to_excel_columns(source_cols, excel_headers)

        try:

            auto_idx = auto_map_columns(source_cols, excel_headers)

            auto_idx = ensure_no_column_in_mapping(source_cols, auto_idx, excel_headers)

            self.mapping_editor.set_mapping(source_cols, excel_headers, auto_idx)

        except Exception:

            pass



    mapping = _v229_normalize_mapping_to_excel_columns(source_cols, raw_mapping, excel_headers)

    if not mapping or all(x is None for x in mapping):

        raise ValueError("Chưa có mapping cột hoặc mapping đang bỏ qua toàn bộ cột.")



    stt_context = _v231_pick_stt_context(ws, header_row, total_row, excel_headers, preferred_no_col=no_col)

    no_col = stt_context.get("no_col") or no_col

    first_data_row_old = stt_context.get("first_row") or header_row + 1

    style_row = stt_context.get("last_row") or max(header_row + 1, total_row - 1)

    last_no = stt_context.get("last_no") or 0

    strict_first_row = _v231_detect_first_data_row_strict(ws, header_row, total_row, no_col)

    first_data_row = min(first_data_row_old or strict_first_row, strict_first_row)

    formula_profile = _v231_find_best_formula_profile(ws, header_row, total_row, excel_headers)



    # Quan trọng: lấy mẫu công thức TỔNG trước khi insert để giữ đúng dòng bắt đầu SUM.

    sum_col_first_rows = _v231_capture_sum_columns_and_starts(

        ws,

        total_row,

        first_data_row,

        no_col,

        excel_headers=excel_headers,

        formula_profile=formula_profile,

    )
    table_max_col = _v229_table_max_col(no_col, mapping, sum_col_first_rows.keys(), excel_headers)



    row_count = len(rows)

    blank_rows = find_blank_rows_before_total(

        ws,

        header_row,

        total_row,

        row_count,

        mapping,

        no_col,

        excel_headers=excel_headers,

    )

    target_rows = list(blank_rows[:row_count])

    missing_count = row_count - len(target_rows)

    insert_at = None

    if missing_count > 0:

        insert_at = total_row

        _v229_merged_ranges_shift_for_insert(ws, insert_at, missing_count)

        target_rows.extend(range(insert_at, insert_at + missing_count))

        total_row_after = total_row + missing_count

    else:

        total_row_after = total_row

    if not target_rows:

        raise ValueError("Không tìm thấy dòng trống phù hợp để ghi dữ liệu.")

    effective_first_data_row = min(first_data_row, min(target_rows))
    data_last_row = max(target_rows)

    sum_col_first_rows = _v231_capture_sum_columns_and_starts(

        ws,

        total_row,

        effective_first_data_row,

        no_col,

        excel_headers=excel_headers,

        formula_profile=formula_profile,

    )



    for i, data_row in enumerate(rows):

        dst_row = target_rows[i]

        if dst_row >= total_row:

            _v229_safe_copy_row(ws, style_row, dst_row, table_max_col)

        existing_no = ws.cell(dst_row, no_col).value if no_col else None
        if dst_row < total_row and existing_no not in (None, ""):
            pass
        else:
            self._safe_set_cell_value(ws, dst_row, no_col, last_no + i + 1)



        for src_idx, excel_col in enumerate(mapping):

            if excel_col is None or excel_col == no_col:

                continue

            try:

                # Ô có công thức mẫu thì giữ công thức, không ghi đè OCR.

                if _v229_is_formula(ws.cell(dst_row, excel_col).value):

                    continue

            except Exception:

                pass

            src_name = source_cols[src_idx] if src_idx < len(source_cols) else ""

            val = data_row[src_idx] if src_idx < len(data_row) else ""

            self._safe_set_cell_value(ws, dst_row, excel_col, _v23_convert_by_source(src_name, val))



    # CHỐT: SUM hết tới dòng ngay trước dòng TỔNG mới.

    sum_last_row = data_last_row

    sum_columns = sorted(sum_col_first_rows.keys())

    for c in sum_columns:

        if c == no_col:

            continue

        try:

            letter = get_column_letter(c)

            start_row = sum_col_first_rows.get(c) or first_data_row

            ws.cell(total_row_after, c).value = f"=SUM({letter}{start_row}:{letter}{sum_last_row})"

        except Exception:

            pass


    updated_total_cols = []
    try:
        updated_total_cols = update_total_formulas(
            ws,
            total_row_after,
            first_data_row,
            data_last_row,
            excel_headers=excel_headers,
            no_col=no_col,
        )
    except Exception:
        updated_total_cols = []

    cleared_non_total_formula_cols = _v231_clear_non_total_measure_formulas(
        ws,
        total_row_after,
        excel_headers=excel_headers,
        no_col=no_col,
    )

    force_workbook_recalculate(wb)


    validation = _v23_validate_written_cells(ws, target_rows, rows, source_cols, mapping, no_col)

    out = last_run_dir()

    out.mkdir(exist_ok=True)

    logic = {

        "rule": "V23_1_SUM_HET_TRUOC_DONG_TONG",

        "sheet": ws.title,

        "header_row": header_row,

        "total_row_before": total_row,

        "insert_at": insert_at,

        "rows_added": row_count,

        "data_rows": target_rows,

        "total_row_after": total_row_after,

        "last_stt_before": last_no,

        "new_stt_start": last_no + 1,

        "new_stt_end": last_no + row_count,

        "first_data_row_detected_old": first_data_row_old,

        "first_data_row_strict": strict_first_row,

        "stt_context": {

            "no_col": no_col,

            "first_row": first_data_row_old,

            "style_row": style_row,

            "last_no": last_no,

            "score": stt_context.get("score", 0),

        },

        "sum_last_row_before_total": sum_last_row,

        "updated_total_cols": updated_total_cols,
        "cleared_non_total_formula_cols": [
            {"col": c, "letter": get_column_letter(c)}
            for c in cleared_non_total_formula_cols
        ],

        "sum_columns": [

            {"col": c, "letter": get_column_letter(c), "start_row": sum_col_first_rows.get(c), "formula": ws.cell(total_row_after, c).value}

            for c in sum_columns

        ],

        "mapping": [

            {

                "source": source_cols[i] if i < len(source_cols) else "",

                "excel_col": c,

                "excel_letter": get_column_letter(c) if c else None,

            }

            for i, c in enumerate(mapping)

        ],

        "written_cell_validation": validation,

    }

    (out / "v23_1_sum_het_truoc_dong_tong_logic.json").write_text(

        json.dumps(logic, ensure_ascii=False, indent=2),

        encoding="utf-8"

    )



    return {

        "sheet": ws.title,

        "header_row": header_row,

        "last_stt_before": last_no,

        "start_fill_row": insert_at if insert_at is not None else min(target_rows),

        "next_stt_start": last_no + 1,

        "rows_added": row_count,

        "total_row_after": total_row_after,

        "sum_first_row": effective_first_data_row,

        "sum_last_row": sum_last_row,

        "sum_columns_count": len(sum_columns),

        "logic_file": str(out / "v23_1_sum_het_truoc_dong_tong_logic.json"),

    }



# Override cuối cùng: ép tool dùng logic SUM hết trước dòng TỔNG.
def install_app_overrides(App):
    App.preview_excel = _v229_preview_excel
    App.run_gemini = _v23_run_gemini
    App._apply_rows_to_workbook = _v231_apply_rows_to_workbook
