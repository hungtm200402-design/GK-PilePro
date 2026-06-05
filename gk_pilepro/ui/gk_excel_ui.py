"""Excel page, mapping-template, and workbook UI actions for the main window."""

import copy
import json
import math
import os
import re
import shutil
import subprocess
import threading
import time
import tkinter as tk
import traceback
import uuid
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from gk_pilepro.gk_core import (
    append_audit_event,
    backup_file,
    last_run_dir,
    locked_file_for_write,
    load_mapping_templates,
    new_workflow_id,
    resource_path,
    report_runtime_error_to_admin,
    save_mapping_templates,
    validate_excel_before_write,
    write_role_error_log,
)
from gk_pilepro.gk_excel import (
    auto_map_columns,
    choose_best_sheet_profile,
    ensure_no_column_in_mapping,
    find_header_row_smart,
    find_no_column_smart,
    find_total_row,
    force_workbook_recalculate,
    get_headers_smart,
    norm,
    normalize_vietnam_date,
    read_formula_logic_for_workbook,
    save_selected_excel_files,
    select_longest_stt_chain,
    short_header_name,
)
from gk_pilepro.ui.gk_ui import (
    UI_BG,
    UI_BORDER,
    UI_ERROR,
    UI_MUTED,
    UI_PRIMARY,
    UI_SUCCESS,
    UI_SURFACE,
    UI_SURFACE_2,
    UI_TEXT,
    RoundedMappingEntry,
    ui_button,
    ui_font,
)


def _excel_file_key(self, path):

    try:

        return str(Path(path).resolve()).casefold()

    except Exception:

        return str(path or "").strip().casefold()


def _excel_file_label(self, path):

    try:

        p = Path(path)

        name = p.name or str(path)

        parent = p.parent.name

        if parent and parent not in {".", ""}:

            label = f"{name}  ({parent})"

        else:

            label = name

    except Exception:

        label = str(path or "")

    if len(label) > 34:

        label = f"{label[:16]}...{label[-14:]}"

    return label


def _excel_recent_modified_label(self, path):

    try:

        p = Path(path)

        if not p.exists():

            return "Không rõ"

        dt = datetime.fromtimestamp(p.stat().st_mtime)

        now = datetime.now()

        time_txt = dt.strftime("%I:%M %p").lstrip("0")

        if dt.date() == now.date():

            return f"Hôm nay lúc {time_txt}"

        if (now.date().toordinal() - dt.date().toordinal()) == 1:

            return f"Hôm qua lúc {time_txt}"

        if dt.year == now.year:

            return dt.strftime("%d Tháng %m")

        return dt.strftime("%d/%m/%Y")

    except Exception:

        return "Không rõ"


def _excel_recent_path_label(self, path):

    try:

        p = Path(path)

        parts = list(p.parts)

        if len(parts) <= 1:

            return str(path)

        tail = parts[-3:] if len(parts) >= 3 else parts[-2:]

        return " \u00bb ".join(tail)

    except Exception:

        return str(path or "")


def _clear_excel_recent_rows(self):

    inner = getattr(self, "excel_recent_inner", None)

    if inner is None:

        return

    try:

        for child in inner.winfo_children():

            child.destroy()

    except Exception:

        pass


def _render_excel_recent_rows(self):

    inner = getattr(self, "excel_recent_inner", None)

    if inner is None:

        return

    self._clear_excel_recent_rows()

    try:

        items = list(self.selected_excel_files or [])

        if getattr(self, "excel_recent_mode", "recent") == "pinned":

            items = []

        selected_key = getattr(self, "excel_recent_selected_key", None) or (self._excel_file_key(self.excel_path) if self.excel_path else None)

        if not items:

            empty = tk.Frame(inner, bg=UI_SURFACE, padx=16, pady=24)

            empty.pack(fill="x")

            if getattr(self, "excel_recent_mode", "recent") == "pinned":
                tk.Label(empty, text="Chưa có file nào được ghim.", bg=UI_SURFACE, fg=UI_MUTED, font=ui_font(11)).pack(anchor="center")
                tk.Label(empty, text="Chọn một file ở tab Gần đây, rồi ghim nếu cần trong bản tiếp theo.", bg=UI_SURFACE, fg=UI_MUTED, font=ui_font(10)).pack(anchor="center", pady=(4, 0))
            else:
                tk.Label(empty, text="Chưa có file Excel nào được chọn.", bg=UI_SURFACE, fg=UI_MUTED, font=ui_font(11)).pack(anchor="center")
                tk.Label(empty, text="Bấm Chọn Excel để thêm file vào danh sách gần đây.", bg=UI_SURFACE, fg=UI_MUTED, font=ui_font(10)).pack(anchor="center", pady=(4, 0))
            return



        for idx, path in enumerate(items):

            p = Path(path)

            key = self._excel_file_key(path)

            is_active = key == selected_key

            row_bg = "#eaf2ff" if is_active else "#ffffff"

            row_border = "#c7dfff" if is_active else "#e7edf6"

            row = tk.Frame(inner, bg=row_bg, highlightthickness=1, highlightbackground=row_border, padx=10, pady=8)

            row.pack(fill="x", pady=(0, 8))



            icon_box = tk.Frame(row, bg="#e8f5e9", width=28, height=28)

            icon_box.pack(side="left", padx=(0, 10))

            icon_box.pack_propagate(False)

            tk.Label(icon_box, text="X", bg="#1f8b4c", fg="#ffffff", font=ui_font(11, bold=True), width=2, height=1).pack(fill="both", expand=True)



            mid = tk.Frame(row, bg=row_bg)

            mid.pack(side="left", fill="both", expand=True)

            tk.Label(mid, text=p.stem, bg=row_bg, fg=UI_TEXT, font=ui_font(11, bold=True), anchor="w").pack(anchor="w")

            tk.Label(mid, text=self._excel_recent_path_label(path), bg=row_bg, fg=UI_MUTED, font=ui_font(10), anchor="w").pack(anchor="w", pady=(2, 0))



            right = tk.Frame(row, bg=row_bg)

            right.pack(side="right", padx=(10, 0))

            tk.Label(right, text=self._excel_recent_modified_label(path), bg=row_bg, fg=UI_MUTED, font=ui_font(10), anchor="e").pack(anchor="e")



            def open_row(_e=None, item_path=str(path)):

                self._load_excel_file(item_path)

                self.show_home_page()

                return "break"



            for widget in (row, icon_box, mid, right):

                widget.bind("<Double-Button-1>", open_row)

            for child in mid.winfo_children() + right.winfo_children():

                child.bind("<Double-Button-1>", open_row)

        self._bind_recent_mousewheel_recursive()

    except Exception:

        pass


def _sync_excel_recent_sidebar(self):

    inner = getattr(self, "excel_recent_inner", None)

    if inner is None:

        return

    self._render_excel_recent_rows()


def _set_excel_recent_visible(self, visible):

    panel = getattr(self, "excel_recent_panel", None)

    filters = getattr(self, "filters_card", None)

    if panel is None:

        return

    try:

        if visible:

            if filters is not None:

                panel.pack(fill="x", pady=(0, 12), before=filters)

            else:

                panel.pack(fill="x", pady=(0, 12))

        else:

            panel.pack_forget()

        self.excel_recent_visible = bool(visible)

    except Exception:

        pass


def show_excel_page(self, event=None):

    return self.show_page("excel")


def show_mapping_page(self, event=None):

    return self.show_page("mapping")


def _mapping_widget_contains(self, widget):
    targets = {
        getattr(self, "mapping_templates_canvas", None),
        getattr(self, "mapping_templates_inner", None),
    }
    while widget is not None:
        if widget in targets:
            return True
        widget = getattr(widget, "master", None)
    return False


def _on_mapping_mousewheel(self, event):
    canvas = getattr(self, "mapping_templates_canvas", None)
    if canvas is None:
        return
    if not self._mapping_widget_contains(getattr(event, "widget", None)):
        return
    try:
        delta = getattr(event, "delta", 0) or 0
        if delta:
            canvas.yview_scroll(int(-1 * (delta / 120)), "units")
            return
        num = getattr(event, "num", None)
        if num == 4:
            canvas.yview_scroll(-1, "units")
        elif num == 5:
            canvas.yview_scroll(1, "units")
    except Exception:
        pass


def _bind_mapping_mousewheel_recursive(self, widget=None):
    widget = widget or getattr(self, "mapping_templates_inner", None)
    if widget is None:
        return
    try:
        widget.bind("<MouseWheel>", self._on_mapping_mousewheel, add="+")
        widget.bind("<Button-4>", self._on_mapping_mousewheel, add="+")
        widget.bind("<Button-5>", self._on_mapping_mousewheel, add="+")
    except Exception:
        pass
    try:
        for child in widget.winfo_children():
            self._bind_mapping_mousewheel_recursive(child)
    except Exception:
        pass


def _mapping_template_kind_label(self):

    kind = str(getattr(self, "current_doc_kind", "") or "").strip().lower()

    if "phieu" in kind or "coc" in kind:

        return "Phiếu cọc"

    table = None

    try:

        table = self.table_editor.get_current_table()

    except Exception:

        table = None

    table_name = str((table or {}).get("name") or "").lower()

    if "phiếu" in table_name or "phieu" in table_name or "cọc" in table_name or "coc" in table_name:

        return "Phiếu cọc"

    return "Khối lượng"


def ask_mapping_template_name(self, default_name):

    result = {"value": None}

    win = tk.Toplevel(self.root)

    win.title("Lưu mẫu mapping")

    win.configure(bg=UI_SURFACE)

    win.resizable(False, False)

    try:

        win.transient(self.root)

    except Exception:

        pass

    win.grab_set()

    body = tk.Frame(win, bg=UI_SURFACE, padx=28, pady=24)

    body.pack(fill="both", expand=True)

    tk.Label(body, text="Lưu mẫu mapping", bg=UI_SURFACE, fg=UI_TEXT, font=ui_font(12, bold=True)).pack(anchor="w")

    tk.Label(body, text="Đặt tên để dễ nhận biết khi dùng lại trong mục Mẫu mapping.", bg=UI_SURFACE, fg=UI_MUTED, font=ui_font(11)).pack(anchor="w", pady=(4, 14))

    tk.Label(body, text="Tên mẫu mapping", bg=UI_SURFACE, fg=UI_TEXT, font=ui_font(11, bold=True)).pack(anchor="w", pady=(0, 6))

    name_var = tk.StringVar(value=default_name)

    entry_wrap = tk.Frame(body, bg="#f8fbff", highlightthickness=1, highlightbackground="#bcd2ee")

    entry_wrap.pack(fill="x")

    name_entry = tk.Entry(

        entry_wrap,

        textvariable=name_var,

        relief="flat",

        bd=0,

        bg="#f8fbff",

        fg=UI_TEXT,

        insertbackground=UI_TEXT,

        font=ui_font(11),

    )

    name_entry.pack(fill="x", padx=10, pady=8)

    error_var = tk.StringVar(value="")

    tk.Label(body, textvariable=error_var, bg=UI_SURFACE, fg="#b91c1c", font=ui_font(10)).pack(anchor="w", pady=(6, 0))

    bottom = tk.Frame(win, bg="#f8fafc", padx=18, pady=14, highlightthickness=1, highlightbackground="#e2e8f0")

    bottom.pack(fill="x")

    actions = tk.Frame(bottom, bg="#f8fafc")

    actions.pack(anchor="e")

    def submit(_event=None):

        value = name_var.get().strip()

        if not value:

            error_var.set("Vui lòng nhập tên mẫu mapping.")

            return "break"

        result["value"] = value

        win.destroy()

        return "break"

    def cancel(_event=None):

        win.destroy()

        return "break"

    ui_button(actions, "Hủy", cancel, width=9, variant="soft").pack(side="right", padx=(8, 0))

    ui_button(actions, "Lưu mẫu", submit, width=11, variant="primary").pack(side="right")

    win.bind("<Return>", submit)

    win.bind("<Escape>", cancel)

    self._center_dialog_on_screen(win)

    try:

        win.lift()

        win.focus_force()

        name_entry.focus_force()

        name_entry.select_range(0, "end")

    except Exception:

        pass

    self.root.wait_window(win)

    return result["value"]


def save_current_mapping_template(self):

    editor = getattr(self, "mapping_editor", None)

    if editor is None:

        return

    mapping = editor.get_mapping()

    source_cols = list(getattr(editor, "table_cols", []) or [])

    excel_headers = list(getattr(editor, "excel_headers", []) or [])

    if not source_cols or not excel_headers or not mapping:

        messagebox.showwarning("Chưa có mapping", "Bạn cần auto map hoặc chọn mapping cột trước khi lưu mẫu.")

        return

    header_by_col = {}

    for col_idx, name in excel_headers:

        try:

            header_by_col[int(col_idx)] = str(name or "")

        except Exception:

            pass

    pairs = []

    for idx, source_name in enumerate(source_cols):

        target_col = mapping[idx] if idx < len(mapping) else None

        if target_col is None:

            continue

        try:

            target_col = int(target_col)

        except Exception:

            continue

        pairs.append(

            {

                "source": str(source_name or ""),

                "target_col": target_col,

                "target_letter": get_column_letter(target_col),

                "target": header_by_col.get(target_col, ""),

            }

        )

    if not pairs:

        messagebox.showwarning("Chưa có mapping", "Chưa có cặp cột nào được chọn để lưu mẫu.")

        return

    kind_label = self._mapping_template_kind_label()

    default_name = f"{kind_label} - {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    name = self.ask_mapping_template_name(default_name)

    if not name:

        return

    name = name.strip()

    if not name:

        return

    templates = list(getattr(self, "mapping_templates", []) or load_mapping_templates())

    existing_idx = next((i for i, item in enumerate(templates) if str(item.get("name") or "").strip().lower() == name.lower()), None)

    if existing_idx is not None:

        if not messagebox.askyesno("Trùng tên mẫu", "Tên mẫu này đã tồn tại. Bạn có muốn ghi đè không?"):

            return

        template_id = templates[existing_idx].get("id") or uuid.uuid4().hex[:12].upper()

        created_at = templates[existing_idx].get("created_at") or datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    else:

        template_id = uuid.uuid4().hex[:12].upper()

        created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    template = {

        "id": template_id,

        "name": name,

        "kind": kind_label,

        "created_at": created_at,

        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),

        "source_columns": source_cols,

        "excel_headers": [{"col": int(col), "name": str(label or "")} for col, label in excel_headers],

        "mapping": pairs,

    }

    if existing_idx is not None:

        templates[existing_idx] = template

    else:

        templates.insert(0, template)

    self.mapping_templates = templates[:200]

    save_mapping_templates(self.mapping_templates)

    self._render_mapping_templates()

    try:

        self._set_status(f"Đã lưu mẫu mapping: {name}", "success")

    except Exception:

        pass

    messagebox.showinfo("Đã lưu", "Mẫu mapping đã được lưu vào mục Mẫu mapping.")


def apply_mapping_template(self, template):

    editor = getattr(self, "mapping_editor", None)

    if editor is None:

        return

    try:

        table = self.table_editor.get_current_table()

    except Exception:

        table = None

    source_cols = list((table or {}).get("columns") or getattr(editor, "table_cols", []) or [])

    excel_headers = list(getattr(self, "excel_headers", []) or getattr(editor, "excel_headers", []) or [])

    if not source_cols or not excel_headers:

        messagebox.showwarning("Thiếu dữ liệu", "Cần có bảng OCR và Excel trước khi áp dụng mẫu mapping.")

        return

    target_idx_by_col = {}

    for idx, (col_idx, _name) in enumerate(excel_headers):

        try:

            target_idx_by_col[int(col_idx)] = idx

        except Exception:

            pass

    pair_by_source = {}

    for pair in template.get("mapping") or []:

        try:

            pair_by_source[norm(pair.get("source"))] = int(pair.get("target_col"))

        except Exception:

            continue

    auto_idx = []

    for source_name in source_cols:

        target_col = pair_by_source.get(norm(source_name))

        auto_idx.append(target_idx_by_col.get(target_col))

    editor.set_mapping(source_cols, excel_headers, auto_idx)

    self.show_home_page()

    try:

        self._set_status(f"Đã áp dụng mẫu mapping: {template.get('name') or ''}", "success")

    except Exception:

        pass


def delete_mapping_template(self, template_id):

    if not messagebox.askyesno("Xóa mẫu", "Bạn có chắc muốn xóa mẫu mapping này không?"):

        return

    self.mapping_templates = [

        item for item in (getattr(self, "mapping_templates", []) or []) if item.get("id") != template_id

    ]

    save_mapping_templates(self.mapping_templates)

    self._render_mapping_templates()


def _render_mapping_templates(self):

    host = getattr(self, "mapping_templates_inner", None)

    if host is None:

        return

    for child in host.winfo_children():

        child.destroy()

    self.mapping_templates = load_mapping_templates()

    if not self.mapping_templates:

        empty = tk.Frame(host, bg=UI_SURFACE, padx=20, pady=30)

        empty.pack(fill="x")

        tk.Label(empty, text="Chưa có mẫu mapping nào.", bg=UI_SURFACE, fg=UI_TEXT, font=ui_font(12, bold=True)).pack(anchor="center")

        tk.Label(empty, text="Sau khi xác nhận mapping cột, bấm Lưu mẫu để lưu lại tại đây.", bg=UI_SURFACE, fg=UI_MUTED, font=ui_font(11)).pack(anchor="center", pady=(6, 0))

        return

    for template in self.mapping_templates:

        item = tk.Frame(host, bg="#fbfdff", padx=12, pady=10, highlightthickness=1, highlightbackground=UI_BORDER)

        item.pack(fill="x", pady=(0, 10))

        top = tk.Frame(item, bg="#fbfdff")

        top.pack(fill="x")

        title_box = tk.Frame(top, bg="#fbfdff")

        title_box.pack(side="left", fill="x", expand=True)

        tk.Label(

            title_box,

            text=str(template.get("name") or "Mẫu mapping"),

            bg="#fbfdff",

            fg=UI_TEXT,

            font=ui_font(11, bold=True),

        ).pack(anchor="w")

        meta = f"{template.get('kind') or 'Mapping'} • {template.get('updated_at') or template.get('created_at') or ''}"

        tk.Label(title_box, text=meta, bg="#fbfdff", fg=UI_MUTED, font=ui_font(10)).pack(anchor="w", pady=(2, 0))

        ui_button(top, "Áp dụng", lambda tpl=template: self.apply_mapping_template(tpl), width=9, variant="primary").pack(side="right", padx=(8, 0))

        ui_button(top, "Xóa", lambda tid=template.get("id"): self.delete_mapping_template(tid), width=7, variant="warn").pack(side="right")

        rows = tk.Frame(item, bg="#fbfdff")

        rows.pack(fill="x", pady=(8, 0))

        for pair in (template.get("mapping") or [])[:14]:

            target = f"{pair.get('target_letter') or ''}: {pair.get('target') or ''}".strip()

            line = f"{pair.get('source') or ''}  →  {target}"

            tk.Label(rows, text=line, bg="#fbfdff", fg=UI_TEXT, font=ui_font(10), anchor="w").pack(fill="x", pady=1)

        total = len(template.get("mapping") or [])

        if total > 14:

            tk.Label(rows, text=f"... còn {total - 14} cặp mapping", bg="#fbfdff", fg=UI_MUTED, font=ui_font(10)).pack(anchor="w", pady=(2, 0))

    self._bind_mapping_mousewheel_recursive()


def _set_excel_recent_mode(self, mode):

    if mode not in {"recent", "pinned"}:

        mode = "recent"

    self.excel_recent_mode = mode

    try:

        if hasattr(self, "excel_tab_recent"):

            self.excel_tab_recent.config(fg=UI_TEXT if mode == "recent" else UI_MUTED, font=ui_font(11, bold=(mode == "recent")))

        if hasattr(self, "excel_tab_pinned"):

            self.excel_tab_pinned.config(fg=UI_TEXT if mode == "pinned" else UI_MUTED, font=ui_font(11, bold=(mode == "pinned")))

    except Exception:

        pass

    self._render_excel_recent_rows()


def _remember_selected_excel_file(self, path):

    if not path:

        return

    key = self._excel_file_key(path)

    new_list = [p for p in self.selected_excel_files if self._excel_file_key(p) != key]

    new_list.insert(0, str(Path(path).resolve()))

    self.selected_excel_files = new_list

    self.excel_recent_selected_key = key

    save_selected_excel_files(self.selected_excel_files)

    self._sync_excel_recent_sidebar()


def _select_loaded_workbook_sheet(self, sheets, profiles, prefer_previous=False):
    previous_sheet = self.sheet_var.get() if hasattr(self, "sheet_var") else ""
    active_sheet = ""
    try:
        active_sheet = self.workbook.active.title
    except Exception:
        active_sheet = ""

    selected = ""
    if prefer_previous and previous_sheet in sheets:
        selected = previous_sheet
    elif active_sheet in sheets:
        selected = active_sheet
    else:
        best = choose_best_sheet_profile(profiles)
        if best and best.get("sheet") in sheets:
            selected = best["sheet"]
        elif sheets:
            selected = sheets[0]

    if selected:
        self.sheet_var.set(selected)
        try:
            self.sheet_combo.current(sheets.index(selected))
        except Exception:
            try:
                self.sheet_combo.set(selected)
            except Exception:
                pass
    return selected


def _visible_workbook_sheet_names(workbook):
    names = []
    try:
        for ws in workbook.worksheets:
            if getattr(ws, "sheet_state", "visible") == "visible":
                names.append(ws.title)
    except Exception:
        names = []
    return names or list(getattr(workbook, "sheetnames", []) or [])


def _load_excel_file(self, path):

    previous_path = getattr(self, "excel_path", "")
    resolved_path = str(Path(path).resolve())
    prefer_previous = bool(previous_path and self._excel_file_key(previous_path) == self._excel_file_key(resolved_path))

    self.excel_path = resolved_path

    self.workbook = load_workbook(self.excel_path, data_only=False)

    sheets = _visible_workbook_sheet_names(self.workbook)

    try:
        self.sheet_combo.set_values(sheets)
    except Exception:
        self.sheet_combo["values"] = sheets



    profiles = self._profile_workbook(self.excel_path)

    selected_sheet = self._select_loaded_workbook_sheet(sheets, profiles, prefer_previous=prefer_previous)

    out = last_run_dir()

    out.mkdir(exist_ok=True)

    (out / "current_workbook_profiles.json").write_text(json.dumps(profiles, ensure_ascii=False, indent=2), encoding="utf-8")



    self._display_profiles(profiles, "Đã đọc toàn bộ file Excel vừa chọn")

    self._remember_selected_excel_file(self.excel_path)

    self.excel_recent_selected_key = self._excel_file_key(self.excel_path)

    self._set_status(f"Đã đọc Excel: {len(sheets)} sheet. Đang chọn sheet {selected_sheet or 'trống'}.", "success")

    if not self.current_workflow_id:

        self.current_workflow_id = new_workflow_id()


def read_current_excel_formulas(self):

    if not self.excel_path:

        messagebox.showwarning("Thiếu Excel", "Bạn chưa chọn file Excel.")

        return

    try:

        logic = read_formula_logic_for_workbook(self.excel_path)

        out = last_run_dir()

        out.mkdir(exist_ok=True)

        (out / "excel_formula_logic.json").write_text(json.dumps(logic, ensure_ascii=False, indent=2), encoding="utf-8")



        self.excel_info.delete("1.0", "end")

        lines = []

        lines.append("ĐÃ ĐỌC CÔNG THỨC & CÁCH LÀM EXCEL\n")

        lines.append("=" * 70 + "\n")

        lines.append(f"File: {self.excel_path}\n\n")

        for sh in logic.get("sheets", []):

            lines.append(f"Sheet: {sh.get('sheet')}\n")

            if sh.get("error"):

                lines.append(f"Lỗi: {sh.get('error')}\n\n")

                continue

            lines.append(f"- Header row: {sh.get('header_row')}\n")

            lines.append(f"- Total row: {sh.get('total_row')}\n")

            lines.append(f"- STT col: {sh.get('stt_col')}\n")

            lines.append(f"- Số ô công thức: {sh.get('formula_count')}\n")

            if sh.get("formula_columns"):

                lines.append("- Các cột có công thức:\n")

                for fc in sh.get("formula_columns", [])[:30]:

                    lines.append(f"  + {fc['col']} ({fc.get('header','')}): {fc['count']} ô, mẫu {fc['sample_cell']} = {fc['sample_formula']}\n")

            if sh.get("total_formulas"):

                lines.append("- Công thức dòng TỔNG:\n")

                for tf in sh.get("total_formulas", [])[:30]:

                    lines.append(f"  + {tf['cell']} ({tf.get('header','')}): {tf['formula']}\n")

            if sh.get("rules_text"):

                lines.append("- Cách làm tóm tắt:\n")

                for rule in sh.get("rules_text", [])[:30]:

                    lines.append(f"  + {rule}\n")

            lines.append("\n")

        lines.append("Log đầy đủ: last_run_v12\\excel_formula_logic.json\n")

        self.excel_info.insert("1.0", "".join(lines))

        self._set_status("Đã đọc công thức và cách làm Excel.", "success")

    except Exception:

        out = last_run_dir()

        out.mkdir(exist_ok=True)

        (out / "last_error_formula_logic.txt").write_text(traceback.format_exc(), encoding="utf-8")

        messagebox.showerror("Lỗi đọc công thức", "Có lỗi. Xem last_run_v12/last_error_formula_logic.txt")

        self._set_status("Lỗi đọc công thức Excel.", "error")


def scan_excel_folder(self):

    folder = filedialog.askdirectory(title="Chọn thư mục chứa các file Excel")

    if not folder:

        return

    self.excel_folder = folder

    paths = []

    for ext in ("*.xlsx", "*.xlsm"):

        paths.extend(Path(folder).glob(ext))

    paths = [p for p in paths if not p.name.startswith("~$")]



    if not paths:

        messagebox.showwarning("Không có Excel", "Thư mục này không có file .xlsx/.xlsm.")

        return



    all_profiles = []

    errors = []

    for p in paths:

        try:

            all_profiles.extend(self._profile_workbook(p))

        except Exception as e:

            errors.append({"file": str(p), "error": repr(e)})



    out = last_run_dir()

    out.mkdir(exist_ok=True)

    (out / "all_excel_profiles.json").write_text(json.dumps(all_profiles, ensure_ascii=False, indent=2), encoding="utf-8")

    if errors:

        (out / "all_excel_scan_errors.json").write_text(json.dumps(errors, ensure_ascii=False, indent=2), encoding="utf-8")



    self._display_profiles(all_profiles, f"Đã quét {len(paths)} file Excel trong thư mục")

    self._set_status(f"Đã quét {len(paths)} file Excel. Log: last_run_v12/all_excel_profiles.json", "success")


def choose_excel(self):

    p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx;*.xlsm"), ("All", "*.*")])

    if not p:

        return

    try:

        self._load_excel_file(p)

    except Exception:

        out = last_run_dir()

        out.mkdir(exist_ok=True)

        (out / "last_error_open_excel.txt").write_text(traceback.format_exc(), encoding="utf-8")

        messagebox.showerror("Lỗi mở Excel", "Có lỗi khi đọc Excel. Xem last_run_v12/last_error_open_excel.txt")

        self._set_status("Lỗi mở Excel.", "error")


def open_selected_excel_from_sidebar(self, event=None):

    try:

        selected_key = getattr(self, "excel_recent_selected_key", None)

        if selected_key:

            for path in self.selected_excel_files:

                if self._excel_file_key(path) == selected_key:

                    self._load_excel_file(path)

                    self.show_excel_page()

                    self._sync_excel_recent_sidebar()

                    return

        if self.excel_path:

            self._load_excel_file(self.excel_path)

            self.show_excel_page()

    except Exception:

        pass


def refresh_excel_header_info(self):

    self.excel_info.delete("1.0", "end")

    if not self.workbook or not self.sheet_var.get():

        self.excel_info.insert("1.0", "Chưa chọn Excel hoặc sheet.\n")

        return



    try:

        ws = self.workbook[self.sheet_var.get()]



        self.header_row = find_header_row_smart(ws)

        self.excel_headers = get_headers_smart(ws, self.header_row)



        total_row = find_total_row(ws, self.header_row)

        no_col = find_no_column_smart(ws, self.excel_headers, self.header_row, total_row) if total_row else None



        txt = []

        txt.append(f"File: {self.excel_path}\n")

        txt.append(f"Sheet: {ws.title}\n")

        txt.append(f"Header: dòng {self.header_row}")

        if total_row:

            txt.append(f" | TỔNG: dòng {total_row}")

        if no_col:

            txt.append(f" | STT: cột {get_column_letter(no_col)}")

        txt.append("\n\n")



        txt.append("Cột phát hiện:\n")

        for col_idx, name in self.excel_headers:

            txt.append(f"- {get_column_letter(col_idx)}: {short_header_name(name)}\n")



        if total_row and no_col:

            try:

                best = select_longest_stt_chain(ws, no_col, self.header_row, total_row)

                if best:

                    txt.append(

                        f"\nChuỗi STT chọn: {best[0][1]} → {best[-1][1]} "

                        f"(dòng {best[0][0]} → {best[-1][0]}) | STT cuối chuỗi: {best[-1][1]}\n"

                    )

                else:

                    txt.append("\nChưa tìm thấy chuỗi STT chuẩn.\n")

            except Exception:

                txt.append("\nKhông đọc được chuỗi STT.\n")



        self.excel_info.insert("1.0", "".join(txt))



    except Exception:

        out = last_run_dir()

        out.mkdir(exist_ok=True)

        (out / "last_error_excel_info.txt").write_text(traceback.format_exc(), encoding="utf-8")

        self.excel_info.insert("1.0", "Lỗi đọc thông tin Excel. Xem last_run_v12/last_error_excel_info.txt\n")

        self._set_status("Lỗi đọc thông tin Excel.", "error")


def build_mapping(self):

    table = self.table_editor.get_current_table()

    if not table:
        table = {"columns": [], "rows": []}

    source_cols = list(table.get("columns") or [])
    if not source_cols:
        try:
            source_cols = [str(col) for col in list(self.table_editor.tree["columns"] or []) if str(col).strip()]
        except Exception:
            source_cols = []

    if not self.excel_headers:

        messagebox.showwarning("Thiếu Excel", "Bạn chưa chọn Excel hoặc sheet.")

        return

    if not source_cols:
        self.mapping_editor.set_mapping([], self.excel_headers, [])
        messagebox.showwarning("Thiếu bảng OCR", "Chưa có cột bảng OCR để mapping. Hãy bấm Đọc bảng hoặc Đọc phiếu cọc trước.")
        return


    auto = auto_map_columns(source_cols, self.excel_headers)

    auto = ensure_no_column_in_mapping(source_cols, auto, self.excel_headers)

    self.mapping_editor.set_mapping(source_cols, self.excel_headers, auto)

    self._set_status("Đã auto map từ bảng trong ảnh sang cột của file Excel.", "success")


def preview_excel(self):

    if not self.excel_path:

        messagebox.showwarning("Thiếu Excel", "Bạn chưa chọn file Excel.")

        return

    try:

        wb = load_workbook(self.excel_path)

        info = self._apply_rows_to_workbook(wb)



        out = last_run_dir()

        out.mkdir(exist_ok=True)

        preview_path = out / "preview_sau_khi_ghep.xlsx"

        force_workbook_recalculate(wb)

        wb.save(preview_path)



        try:

            os.startfile(str(preview_path))

        except Exception:

            pass



        self._set_status(f"Đã tạo file xem trước: {preview_path}", "success")

        messagebox.showinfo(

            "Đã tạo xem trước",

            f"Đã tạo file preview để kiểm tra trước khi lưu thật.\n\n"

            f"Sheet: {info['sheet']}\n"

            f"Bắt đầu chèn từ dòng: {info['start_fill_row']}\n"

            f"Số dòng thêm: {info['rows_added']}\n\n"

            f"File preview:\n{preview_path}\n\n"

            f"Lưu ý: nếu STT bắt đầu không đúng, bạn đang chọn file đã bị ghi thử trước đó. Hãy dùng lại file Excel gốc sạch.\n\n"

            f"Nếu đúng thì quay lại tool bấm: Điền tiếp vào Excel."

        )

    except Exception:

        out = last_run_dir()

        out.mkdir(exist_ok=True)

        (out / "last_error_preview.txt").write_text(traceback.format_exc(), encoding="utf-8")

        messagebox.showerror("Lỗi xem trước", "Có lỗi. Xem last_run_v12/last_error_preview.txt")

        self._set_status("Lỗi xem trước Excel.", "error")


def fill_excel(self):

    if not self.excel_path:

        messagebox.showwarning("Thiếu Excel", "Bạn chưa chọn file Excel.")

        return


    backup_path = None

    try:
        backup_path = backup_file(self.excel_path, "excel")

        # Luôn nạp lại file gốc để preview không làm nhân đôi dữ liệu trong bộ nhớ

        wb = load_workbook(self.excel_path)
        validate_excel_before_write(wb)

        info = self._apply_rows_to_workbook(wb)



        out_path = self.excel_path
        
        force_workbook_recalculate(wb)
        
        try:
            with locked_file_for_write(out_path, timeout=10):
                wb.save(out_path)
        except PermissionError:
            if backup_path:
                try:
                    shutil.copy2(backup_path, out_path)
                except Exception as rollback_exc:
                    write_role_error_log("excel_rollback_permission_error", rollback_exc, {"backup": str(backup_path), "target": out_path})
            append_audit_event("export_excel", status="error", file_path=out_path, message="Không thể lưu Excel vì file đang mở hoặc bị khóa.", extra={"backup": str(backup_path or "")})
            messagebox.showerror("Lỗi ghi file", f"Không thể lưu trực tiếp vào file Excel.\nVui lòng đóng file Excel '{Path(out_path).name}' trước khi điền dữ liệu.")
            return
        except TimeoutError as e:
            append_audit_event("export_excel", status="error", file_path=out_path, message=str(e), extra={"backup": str(backup_path or "")})
            messagebox.showerror("File đang được ghi", "File Excel đang được máy khác ghi.\nVui lòng chờ rồi thử lại để tránh sai dữ liệu.")
            return
        except Exception as e:
            if backup_path:
                try:
                    shutil.copy2(backup_path, out_path)
                except Exception as rollback_exc:
                    write_role_error_log("excel_rollback_save_error", rollback_exc, {"backup": str(backup_path), "target": out_path})
            write_role_error_log("excel_save_error", e, {"excel_path": out_path, "backup": str(backup_path or "")})
            append_audit_event("export_excel", status="error", file_path=out_path, message=f"Lỗi lưu Excel: {e}", extra={"backup": str(backup_path or "")})
            messagebox.showerror("Lỗi", f"Có lỗi khi lưu file: {str(e)}")
            return



        out = last_run_dir()

        out.mkdir(exist_ok=True)

        (out / "last_fill_info.json").write_text(json.dumps(info, ensure_ascii=False, indent=2), encoding="utf-8")
        current_table = self.table_editor.get_current_table() or {}
        filled_rows = current_table.get("rows") or []
        numeric_total = 0.0
        numeric_cells = 0
        for row in filled_rows:
            values = row.values() if isinstance(row, dict) else row if isinstance(row, (list, tuple)) else []
            for value in values:
                try:
                    text = str(value).strip().replace(".", "").replace(",", ".")
                    if re.fullmatch(r"-?\d+(\.\d+)?", text):
                        numeric_total += float(text)
                        numeric_cells += 1
                except Exception:
                    pass
        audit_extra = {
            "sheet": info.get("sheet"),
            "rows_added": info.get("rows_added"),
            "start_fill_row": info.get("start_fill_row"),
            "backup_path": str(backup_path or ""),
            "source_excel_path": self.excel_path,
            "source_image_path": self.image_path,
            "numeric_cells": numeric_cells,
            "numeric_total": numeric_total,
        }
        audit_record = append_audit_event(
            "export_excel",
            file_path=out_path,
            message=f"Đã xuất {info.get('rows_added', 0)} dòng vào Excel.",
            extra=audit_extra,
        )
        reconcile_report = {
            "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "file": out_path,
            "sheet": info.get("sheet"),
            "rows_added": info.get("rows_added"),
            "start_fill_row": info.get("start_fill_row"),
            "backup_path": str(backup_path or ""),
            "numeric_cells": numeric_cells,
            "numeric_total": numeric_total,
            "audit_hash": (audit_record or {}).get("hash", ""),
        }
        (out / "last_reconcile_report.json").write_text(json.dumps(reconcile_report, ensure_ascii=False, indent=2), encoding="utf-8")
        current_table_title = str(current_table.get("title") or "").strip()
        current_table_title_l = current_table_title.lower()
        export_ocr_type = str(self.current_doc_kind or "").strip().lower()
        if export_ocr_type not in {"bang_khoi_luong", "phieu_coc"}:
            if "phiếu cọc" in current_table_title_l or "phieu coc" in current_table_title_l:
                export_ocr_type = "phieu_coc"
            elif "khối lượng" in current_table_title_l or "khoi luong" in current_table_title_l:
                export_ocr_type = "bang_khoi_luong"
            else:
                export_ocr_type = "bang_khoi_luong"

        self._record_history(

            "export_excel",

            file_path=out_path,

            sheet=info.get("sheet"),

            rows=info.get("rows_added"),

            message=f"Đã xuất {info.get('rows_added', 0)} dòng",

            extra={

                "ocr_type": export_ocr_type,

                "start_fill_row": info.get("start_fill_row"),

                "source_excel_path": self.excel_path,

                "source_excel_name": Path(self.excel_path).name if self.excel_path else "",

                "output_path": out_path,

                "output_name": Path(out_path).name,

                "source_image_path": self.image_path,

                "source_table_title": current_table_title,

                "table_columns": len(current_table.get("columns", [])),

                "filled_table_title": current_table_title,

                "filled_table_columns": current_table.get("columns") or [],

                "filled_rows_data": current_table.get("rows") or [],
                "backup_path": str(backup_path or ""),
                "audit_hash": (audit_record or {}).get("hash", ""),
                "reconcile_report_path": str(out / "last_reconcile_report.json"),
                "numeric_cells": numeric_cells,
                "numeric_total": numeric_total,

            },

            workflow_id=self.current_workflow_id,

            workflow_label=self.current_workflow_label,

        )



        self._set_status(f"Đã điền {info['rows_added']} dòng vào sheet {info['sheet']}, bắt đầu từ dòng {info['start_fill_row']}.", "success")

        messagebox.showinfo(

            "Xong",

            f"Đã điền {info['rows_added']} dòng vào Excel.\n"

            f"Bắt đầu từ dòng {info['start_fill_row']}\n"

            f"File: {out_path}"

        )

        try:
            os.startfile(str(out_path))
        except Exception:
            pass

    except Exception as exc:
        if backup_path:
            try:
                shutil.copy2(backup_path, self.excel_path)
            except Exception as rollback_exc:
                write_role_error_log("excel_rollback_outer_error", rollback_exc, {"backup": str(backup_path), "target": self.excel_path})
        append_audit_event("export_excel", status="error", file_path=self.excel_path, message="Lỗi khi xuất Excel, đã rollback nếu có backup.", extra={"backup": str(backup_path or "")})

        out = last_run_dir()

        out.mkdir(exist_ok=True)

        error_path, _log_path = report_runtime_error_to_admin(
            "fill_excel",
            exc,
            {
                "excel_path": self.excel_path,
                "backup": str(backup_path or ""),
                "workflow_id": self.current_workflow_id,
                "workflow_label": self.current_workflow_label,
            },
            error_file_name="last_error_fill.txt",
            notify_server=True,
        )

        detail = str(exc).strip() or exc.__class__.__name__
        messagebox.showerror(
            "Lỗi điền Excel",
            f"Xuất Excel bị lỗi:\n{detail}\n\nĐã rollback nếu có backup, ghi log và gửi cho Admin nếu server đang bật.\nFile lỗi: {error_path or 'last_run_v12/last_error_fill.txt'}",
        )

        self._set_status("Lỗi điền Excel, đã ghi log.", "error")

        self._record_history(

            "export_error",

            status="error",

            file_path=self.excel_path,

            message="Lỗi khi xuất Excel",

            workflow_id=self.current_workflow_id,

            workflow_label=self.current_workflow_label,

        )


def install_excel_ui(app_cls):
    app_cls._excel_file_key = _excel_file_key
    app_cls._excel_file_label = _excel_file_label
    app_cls._excel_recent_modified_label = _excel_recent_modified_label
    app_cls._excel_recent_path_label = _excel_recent_path_label
    app_cls._clear_excel_recent_rows = _clear_excel_recent_rows
    app_cls._render_excel_recent_rows = _render_excel_recent_rows
    app_cls._sync_excel_recent_sidebar = _sync_excel_recent_sidebar
    app_cls._set_excel_recent_visible = _set_excel_recent_visible
    app_cls.show_excel_page = show_excel_page
    app_cls.show_mapping_page = show_mapping_page
    app_cls._mapping_widget_contains = _mapping_widget_contains
    app_cls._on_mapping_mousewheel = _on_mapping_mousewheel
    app_cls._bind_mapping_mousewheel_recursive = _bind_mapping_mousewheel_recursive
    app_cls._mapping_template_kind_label = _mapping_template_kind_label
    app_cls.ask_mapping_template_name = ask_mapping_template_name
    app_cls.save_current_mapping_template = save_current_mapping_template
    app_cls.apply_mapping_template = apply_mapping_template
    app_cls.delete_mapping_template = delete_mapping_template
    app_cls._render_mapping_templates = _render_mapping_templates
    app_cls._set_excel_recent_mode = _set_excel_recent_mode
    app_cls._remember_selected_excel_file = _remember_selected_excel_file
    app_cls._visible_workbook_sheet_names = _visible_workbook_sheet_names
    app_cls._select_loaded_workbook_sheet = _select_loaded_workbook_sheet
    app_cls._load_excel_file = _load_excel_file
    app_cls.read_current_excel_formulas = read_current_excel_formulas
    app_cls.scan_excel_folder = scan_excel_folder
    app_cls.choose_excel = choose_excel
    app_cls.open_selected_excel_from_sidebar = open_selected_excel_from_sidebar
    app_cls.refresh_excel_header_info = refresh_excel_header_info
    app_cls.build_mapping = build_mapping
    app_cls.preview_excel = preview_excel
    app_cls.fill_excel = fill_excel
