# -*- coding: utf-8 -*-

import copy
import difflib
import glob
import json
import math
import os
import re
import shutil
import subprocess
import tempfile
import time
import traceback
import unicodedata
from datetime import datetime
from pathlib import Path

from PIL import Image
from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple

from gk_pilepro.gk_core import (
    DEFAULT_MODEL,
    FALLBACK_MODELS,
    backup_file,
    load_formula_profiles,
    save_formula_profiles,
    selected_excel_files_path,
)


TEMPLATE_PRESETS = {"Bảng bất kỳ - tự nhận cột": {}}
CANONICAL_TEMPLATE_COLUMNS = {}


# Giữ dạng không dấu vì hàm norm() đã bỏ dấu tiếng Việt.

TOTAL_MARKERS = {

    "tong", "total", "cong", "sum",

}



SYNONYM_GROUPS = [

    {"stt", "no", "so thu tu", "tt"},

    {"ngay", "date", "ngay thi cong", "ngay ep", "ngay xuat", "ngay thang", "ngay, thang", "ngay xuat date"},

    {"may ep", "machine", "may ep machine"},

    {"ten coc", "pile name", "pile no", "ten coc pile name", "ten coc pile no", "pile"},

    {"loai coc", "type of pile", "pile type", "loai coc type of pile"},

    {"vi tri", "location", "vi tri location"},

    {"d1", "đ1", "1st", "to hop coc 1", "to hop 1", "cot to hop 1"},

    {"d2", "đ2", "2nd", "to hop coc 2", "to hop 2", "cot to hop 2"},

    {"d3", "đ3", "3rd", "to hop coc 3", "to hop 3", "cot to hop 3"},

    {"d4", "đ4", "4th", "to hop coc 4", "to hop 4", "cot to hop 4"},

    {"d5", "đ5", "5th", "to hop coc 5", "to hop 5", "cot to hop 5"},

    {"d6", "đ6", "6th", "to hop coc 6", "to hop 6", "cot to hop 6"},

    {"chieu dai to hop", "tong to hop", "total", "total m", "length of pile", "chieu dai coc", "chieu dai to hop m", "tong so m", "tong so met", "tong so met coc", "tong so m coc", "tong so m coc d500", "total of pile detail length", "total of pile detail length m", "total of pile detail length (m)", "pile detail length", "pile detail length m", "pile detail length (m)"},

    {"length of pile under ground", "under ground length", "ground length", "length under ground"},

    {"length of pile pressing positive", "pressing positive length", "pressing length positive"},

    {"chieu sau ep thuc te", "pressing depth", "thuc te", "reality", "actual depth"},

    {"tai trong dung ep", "luc ep khi dung", "pressing load", "jacking stopping load", "load", "tan", "luc ep", "pressing force", "pressing force ton"},

    {"bat dau", "start", "thoi gian bat dau", "time start"},

    {"ket thuc", "finish", "end", "thoi gian ket thuc", "time finish"},

    {"dau coc thiet ke", "design pile head", "design"},

    {"mat dat tu nhien", "natural ground", "mat dat tu nhien m"},

    {"dau coc thuc te", "actual pile head", "actual"},

    {"ghi chu", "note", "remark", "remarks", "ghi chu remark"},

    {"hop dong", "contract", "contract no", "so hop dong", "hd"},

]



def load_selected_excel_files():

    try:

        data = json.loads(selected_excel_files_path().read_text(encoding="utf-8"))

        if isinstance(data, list):

            out = []

            seen = set()

            for item in data:

                path = str(item or "").strip()

                if not path:

                    continue

                key = str(Path(path).resolve()).casefold()

                if key in seen:

                    continue

                seen.add(key)

                out.append(str(Path(path).resolve()))

            return out

    except Exception:

        pass

    return []



def save_selected_excel_files(paths):

    try:

        unique = []

        seen = set()

        for path in paths or []:

            p = str(path or "").strip()

            if not p:

                continue

            key = str(Path(p).resolve()).casefold()

            if key in seen:

                continue

            seen.add(key)

            unique.append(str(Path(p).resolve()))

        backup_file(selected_excel_files_path(), "config")

        selected_excel_files_path().write_text(

            json.dumps(unique, ensure_ascii=False, indent=2),

            encoding="utf-8",

        )

    except Exception:

        pass



def clean_text(v):

    s = str(v or "").strip()

    s = s.replace("−","-").replace("–","-").replace("—","-")

    if re.fullmatch(r"-?\d+\.\d+", s):

        s = s.replace(".", ",")

    return s



def norm(s):

    s = str(s or "").strip().lower()

    vietnamese_map = str.maketrans(

        "àáạảãâầấậẩẫăằắặẳẵèéẹẻẽêềếệểễìíịỉĩòóọỏõôồốộổỗơờớợởỡùúụủũưừứựửữỳýỵỷỹđ",

        "aaaaaaaaaaaaaaaaaeeeeeeeeeeeiiiiiooooooooooooooooouuuuuuuuuuuyyyyyd"

    )

    s = s.translate(vietnamese_map)

    s = re.sub(r"[\n\r\t/()\-_:;,\.]+", " ", s)

    s = re.sub(r"\s+", " ", s).strip()

    return s



def extract_json(text):

    s = str(text or "").strip()

    s = re.sub(r"^```json\s*", "", s, flags=re.I)

    s = re.sub(r"^```\s*", "", s)

    s = re.sub(r"\s*```$", "", s)

    try:

        return json.loads(s)

    except Exception:

        pass

    m = re.search(r"\{[\s\S]*\}", s)

    if m:

        return json.loads(m.group(0))

    m = re.search(r"\[[\s\S]*\]", s)

    if m:

        return {"tables":[{"title":"","columns":[],"rows":json.loads(m.group(0))}]}

    raise ValueError("Không parse được JSON từ AI.")



def build_prompt():

    return """

Bạn là công cụ đọc bảng từ ảnh để nhập Excel. Nhiệm vụ là đọc THẬT CHUẨN dữ liệu trong bảng.



NGUYÊN TẮC BẮT BUỘC:

1. Chỉ đọc dữ liệu nằm TRONG BẢNG chính.

2. Không đọc chữ ngoài bảng, chữ ký, tên người ký, tiêu đề dưới cuối trang.

3. Không tự bịa dữ liệu.

4. Ô nào trống thì trả về "".

5. Ô nào không chắc thì trả về "" hoặc giữ đúng phần đọc được, không đoán.

6. Giữ đúng thứ tự cột từ trái sang phải.

7. Giữ đúng thứ tự dòng từ trên xuống dưới.

8. Không gộp nhiều ô thành một ô.

9. Không tách một ô thành nhiều ô nếu ảnh chỉ có một ô.

10. Không bỏ dòng dữ liệu có STT.

11. Không lấy các dòng trống phía dưới bảng.

12. Không lấy dòng chữ ký, đại diện chủ đầu tư, đại diện đơn vị thi công.



CÁCH ĐỌC BẢNG:

- Đọc theo đường kẻ bảng, không đọc theo chữ rời rạc.

- Xác định header trước, sau đó đọc từng dòng dữ liệu.

- Nếu header nhiều tầng thì ghép tên cột cho rõ nghĩa.

- Nếu một header cha có nhiều cột con thì phải tách từng cột con riêng.

- Nếu nhiều vùng bảng có cùng header và STT nối tiếp nhau, coi đó là một bảng kéo dài: trả về 1 object trong `tables` và nối toàn bộ rows theo thứ tự STT.

- Nếu cùng một bảng bị ngắt trang, ngắt ảnh, hoặc chia thành nhiều khung nhưng STT vẫn nối tiếp, phải gộp hết rows vào một bảng duy nhất.

- Chỉ trả nhiều object trong `tables` khi đó là các bảng khác loại, khác header, hoặc không phải dữ liệu nối tiếp nhau.

- Không bỏ các dòng sau phần ngắt bảng; phải đọc hết đến dòng dữ liệu cuối cùng.



QUY TẮC TỔ HỢP CỌC:

- Nếu ảnh có cột "Tổ hợp cọc" và dưới đó có nhiều cột con:

  + cột con thứ 1 từ trái sang phải = D1

  + cột con thứ 2 = D2

  + cột con thứ 3 = D3

  + cột con thứ 4 = D4

  + cột con thứ 5 = D5

  + cột con thứ 6 = D6

- Nếu ảnh không ghi sẵn D1/D2/D3... nhưng ô "Tổ hợp cọc" chứa nhiều giá trị kiểu 11 + 13 + 14,

  thì tự tách theo thứ tự trái sang phải thành D1, D2, D3... để điền vào từng cột riêng.

- Ví dụ trong ảnh có "Tổ hợp cọc" gồm 2 ô: 6 | 10

  thì trả về D1 = 6, D2 = 10.

- Không được gộp thành "6 10".

- Nếu ảnh đã ghi sẵn D1/D2 hoặc 1st/2nd thì giữ theo tên đó.



QUY TẮC SỐ LIỆU:

- Giữ dấu phẩy thập phân nếu ảnh dùng dấu phẩy: 14,5; 16,20; -1,20.

- Giữ dấu cộng/trừ: +1,5; -1,20.

- Giữ giờ theo ảnh: 15h00, 15h30, 16h00.

- Giữ ngày theo ảnh: 11/05 hoặc 11/05/2026.

- Không đổi 90 thành 9.0.

- Không đổi D300 thành D30 hoặc 0300.

- Không đổi tên cọc/tim cọc, ví dụ 22, 21, C262, RBH1-C46.

- Với khối lượng, lực ép, chiều dài, số lượng:

  + chỉ trả phần số

  + bỏ đơn vị như t, kg, m, mm nếu có

  + giữ nguyên dấu thập phân đang xuất hiện trong ảnh

  + không tự làm tròn hoặc tự đổi dấu phẩy thành dấu chấm



QUY TẮC DỪNG DÒNG:

- Chỉ lấy các dòng có dữ liệu thật.

- Nếu STT từ 1 đến 12 có dữ liệu, còn 13 trở xuống trống thì chỉ trả 12 dòng.

- Không trả các dòng trống 13,14,15...

- Không lấy nét gạch chéo/ký tên làm dữ liệu.



CỘT THƯỜNG GẶP CẦN GIỮ ĐÚNG:

- STT

- Ngày tháng

- Giờ ép hoặc Thời gian bắt đầu/kết thúc

- Tên tim cọc hoặc Tên cọc

- Loại cọc

- D1, D2, D3, D4, D5, D6

- Chiều dài cọc

- Chiều dài ép

- Ép âm dương

- Lực ép

- Ghi chú



BẮT BUỘC SOÁT LỖI THEO Ô:

- Đọc từng ô theo đường kẻ, không suy từ dòng bên trên nếu ô đó trống.

- Dữ liệu nào viết trong ảnh thì giữ y nguyên ở JSON: ngày, giờ, tên cọc, loại cọc, dấu +/-, dấu phẩy, ghi chú.

- Nếu một ô khó đọc, để trống đúng ô đó; không sửa cả dòng, không tự đoán.

- Sau khi đọc xong, tự kiểm tra lại từng cột: STT, ngày, giờ, tên cọc, loại cọc, D1, D2, chiều dài, ép âm dương, lực ép, ghi chú.

- Không tự tính tổng từ ảnh; tổng sẽ do Excel tính lại theo file mẫu.



KIỂM TRA TRƯỚC KHI TRẢ JSON:

- Số ô mỗi dòng phải bằng số cột.

- Nếu thiếu ô thì điền "" để đủ cột.

- Không được lệch cột.

- Dữ liệu ở cột nào phải đúng cột đó.

- Tên cọc không được nhảy sang Loại cọc.

- D1/D2 không được nhảy sang Chiều dài.

- Ghi chú không được nhảy sang Lực ép.





QUY TẮC TỔ HỢP CỌC CỰC KỲ QUAN TRỌNG:

- Nếu header là "Tổ hợp cọc" và trong mỗi dòng có 2 số nằm dưới vùng đó, ví dụ 6 | 10:

  + số thứ nhất phải là D1

  + số thứ hai phải là D2

- Không được trả một cột chung tên "Tổ hợp cọc".

- Không được làm lệch cột sau nó.

- Ví dụ đúng:

  Loại cọc=D300, D1=6, D2=10, Chiều dài cọc=16, Chiều dài ép=14,5, Ép âm dương=+1,5, Lực ép=90, Ghi chú=cắt cọc.



Output JSON thuần, không markdown, không giải thích:

{

  "tables": [

    {

      "title": "tên bảng nếu đọc được, không có thì để rỗng",

      "columns": ["cột 1", "cột 2", "cột 3"],

      "rows": [

        ["ô11", "ô12", "ô13"],

        ["ô21", "ô22", "ô23"]

      ]
    },

    {

      "title": "bảng khác loại nếu ảnh có nhiều bảng khác header",

      "columns": ["cột 1", "cột 2"],

      "rows": [

        ["ô11", "ô12"]

      ]

    }

  ]

}

"""





def build_prompt_phieu_coc(excel_columns=None):

    """

    Prompt đọc phiếu cọc.

    - Nếu có excel_columns: chỉ trả 1 bảng duy nhất với cột ĐÚNG Y HỆT Excel.

    - Nếu không có: trả cả "Thông tin phiếu" lẫn "Danh sách cọc".

    """

    if excel_columns:

        # Lọc bỏ cột STT/No vì tool tự nối STT

        skip_keys = {"stt", "no", "so thu tu", "tt"}

        cols_to_fill = [c for c in excel_columns if norm(c) not in skip_keys]

        col_list = json.dumps(cols_to_fill, ensure_ascii=False)

        n_cols = len(cols_to_fill)

        example_row = json.dumps(["" for _ in cols_to_fill], ensure_ascii=False)

        # Tạo hướng dẫn đặc biệt cho cột dạng "Độ dài X"

        import re as _re

        do_dai_cols = []

        for c in cols_to_fill:

            m = _re.search(r"(?:do\s*dai|chieu\s*dai|length|dai)\s*(\d+[\.,]?\d*)", norm(c))

            if m:

                do_dai_cols.append((c, m.group(1).replace(",", ".")))



        do_dai_hint = ""

        if do_dai_cols:

            do_dai_lines = "\n".join(

                f"  - Cột \"{c}\": số lượng cọc có chiều dài = {dl}m trong phiếu"

                for c, dl in do_dai_cols

            )

            do_dai_hint = f"""

QUY TẮC ĐẶC BIỆT – CỘT ĐỘ DÀI:

Excel có các cột theo từng chiều dài cọc:

{do_dai_lines}



Cách điền:
- GỘP TOÀN BỘ CÁC LOẠI ĐỘ DÀI CỦA 1 PHIẾU VÀO 1 DÒNG DUY NHẤT.
- Xem trong phiếu có các loại chiều dài nào, hãy điền số lượng tương ứng vào các cột "Độ dài X" trên cùng 1 dòng JSON đó.
- Ví dụ: 1 phiếu ghi "6m - 7 cây", "9m - 7 cây", "10m - 10 cây" -> Bạn chỉ tạo 1 dòng JSON, trong đó cột "Độ dài 6" = 7, cột "Độ dài 9" = 7, cột "Độ dài 10" = 10.
- Nếu phiếu không có cọc độ dài đó thì để "".
- TUYỆT ĐỐI KHÔNG tạo nhiều dòng JSON cho cùng 1 phiếu.

"""

        col_guide = "\n".join(
            f"  - Cột {i+1}: \"{c}\" → tìm thông tin tương ứng trong phiếu, để \"\" nếu không có"
            for i, c in enumerate(cols_to_fill)
        )

        return f"""

Bạn là công cụ đọc PHIẾU CỌC từ ảnh để điền vào file Excel.

FILE EXCEL CÓ {n_cols} CỘT SAU – bạn PHẢI trả đúng {n_cols} cột này, ĐÚNG THỨ TỰ, ĐÚNG TÊN:

{col_list}

HƯỚNG DẪN TỪNG CỘT:

{col_guide}

{do_dai_hint}

NGUYÊN TẮC:

1. Đọc phiếu cọc trong ảnh (phiếu xuất, nhập, giao cọc).

2. GỘP TOÀN BỘ thông tin của 1 phiếu cọc (dù phiếu có liệt kê nhiều loại độ dài khác nhau) thành 1 ROW DUY NHẤT trong JSON. Khác với trước đây, KHÔNG tách mỗi dòng trong phiếu thành một row riêng biệt. Mọi độ dài đều được điền vào các cột tương ứng trên cùng 1 row.

3. Mỗi row phải có đúng {n_cols} ô tương ứng với {n_cols} cột ở trên theo đúng thứ tự.

4. Ô không có thông tin → "".

5. Không bịa dữ liệu.

6. Giữ nguyên: ngày tháng, số lượng, chiều dài, loại cọc, mã cọc.

7. Không đọc dòng Tổng/Cộng làm dòng dữ liệu.

8. Không lấy chữ ký.



QUY TẮC SỐ:

- Giữ dấu phẩy thập phân: 14,5 giữ là 14,5.

- Giữ ngày như ảnh: 11/05/2026.

- Nếu AI trả ngày kiểu 2026-05-26 hoặc 2026/05/26 thì đổi về dd/mm/yyyy.

- D300 giữ là D300, PHC500 giữ là PHC500.

- Nếu phiếu có ô hoặc cột "Tổng số mét" thì phải điền đúng giá trị đó vào cột Excel tương ứng "Tổng số m cọc D500" / "Tổng số m cọc" / "Tổng số m", không được để trống.

- Với cột khối lượng/số lượng/chiều dài/lực ép:

  - chỉ lấy số

  - bỏ đơn vị nếu có

  - giữ nguyên dấu thập phân đang có trong ảnh

  - không tự làm tròn và không tự đổi dấu phẩy thành dấu chấm



KIỂM TRA: mỗi row phải có đúng {n_cols} ô.



Output JSON thuần, không markdown:

{{

  "tables": [

    {{

      "title": "Dữ liệu phiếu cọc",

      "columns": {col_list},

      "rows": [

        {example_row}

      ]

    }}

  ]

}}

"""

    else:

        # Không có Excel: trả cả 2 bảng

        return """

Bạn là công cụ đọc PHIẾU CỌC từ ảnh để nhập vào Excel.



NGUYÊN TẮC:

1. Đọc toàn bộ thông tin: phần form chung và bảng danh sách cọc.

2. Không bịa dữ liệu.

3. Giữ ngày tháng theo kiểu Việt Nam, ví dụ 11/05/2026.

4. Nếu ngày đang ở dạng 2026-05-26 hoặc 2026/05/26 thì đổi về dd/mm/yyyy.

5. Không lấy dòng tổng/cộng làm dữ liệu.

6. Không lấy chữ ký.



Cột thường gặp: STT, Ngày xuất, Số HĐ, Loại cọc, Chiều dài (m), Số lượng, Khối lượng, Ghi chú.



Output JSON thuần:

{

  "tables": [

    {

      "title": "Thông tin phiếu",

      "columns": ["Trường", "Giá trị"],

      "rows": [["Số phiếu","..."],["Ngày lập","..."]]

    },

    {

      "title": "Danh sách cọc",

      "columns": ["STT","Ngày xuất","Loại cọc","Chiều dài (m)","Số lượng","Ghi chú"],

      "rows": [["1","11/05/2026","D300","6","10",""]]

    }

  ]

}

"""





def call_gemini_phieu_coc(image_path, api_key, model_name, excel_columns=None):

    """

    Gọi AI đọc phiếu cọc từ ảnh.

    Nếu truyền excel_columns, AI sẽ trả về dữ liệu khớp với cột Excel.

    """

    from google import genai

    from google.genai import types

    import time



    client = genai.Client(api_key=api_key)

    image = Image.open(image_path)



    preferred = (model_name or DEFAULT_MODEL).strip()

    models_to_try = []

    for m in [preferred] + FALLBACK_MODELS:

        if m and m not in models_to_try:

            models_to_try.append(m)



    last_error = None

    tried = []

    prompt = build_prompt_phieu_coc(excel_columns)



    for model in models_to_try:

        tried.append(model)

        for attempt in range(3):

            try:

                response = client.models.generate_content(

                    model=model,

                    contents=[prompt, image],

                    config=types.GenerateContentConfig(

                        temperature=0,

                        response_mime_type="application/json",

                    ),

                )



                raw = response.text or ""

                data = extract_json(raw)

                if "tables" not in data:

                    raise ValueError("AI không trả về khóa tables.")



                tables = []

                for t in data.get("tables", []):

                    title = clean_text(t.get("title", ""))

                    columns = [clean_text(c) for c in t.get("columns", [])]

                    rows = t.get("rows", [])



                    norm_rows = []

                    if rows and isinstance(rows[0], dict):

                        if not columns:

                            columns = list(rows[0].keys())

                        for r in rows:

                            norm_rows.append([clean_text(r.get(c, "")) for c in columns])

                    else:

                        for r in rows:

                            if isinstance(r, list):

                                rr = [clean_text(x) for x in r]

                                if columns:

                                    rr = rr[:len(columns)] + [""] * max(0, len(columns) - len(rr))

                                norm_rows.append(rr)



                    if columns and norm_rows:

                        tables.append({"title": title, "columns": columns, "rows": norm_rows})



                raw_with_meta = f"MODEL_USED={model}\nTRIED_MODELS={tried}\nLOAI=PHIEU_COC\n\n{raw}"

                return tables, raw_with_meta



            except Exception as e:

                last_error = e

                msg = str(e)

                if "503" in msg or "500" in msg or "INTERNAL" in msg or "UNAVAILABLE" in msg or "high demand" in msg:

                    time.sleep(3 + attempt * 2)

                    continue

                if any(x in msg for x in [

                    "429", "RESOURCE_EXHAUSTED", "Quota",

                    "INVALID_ARGUMENT", "API key not valid",

                    "not found", "404"

                ]):

                    break

                raise



    raise RuntimeError(

        "Không gọi được AI (phiếu cọc) sau khi thử các model: "

        + ", ".join(tried)

        + "\nLỗi cuối: "

        + repr(last_error)

    )





def call_gemini(image_path, api_key, model_name):

    """

    Gọi AI đọc bảng từ ảnh.

    Có retry khi 503 quá tải và fallback model.

    """

    from google import genai

    from google.genai import types

    import time



    client = genai.Client(api_key=api_key)

    image = Image.open(image_path)



    preferred = (model_name or DEFAULT_MODEL).strip()

    models_to_try = []

    for m in [preferred] + FALLBACK_MODELS:

        if m and m not in models_to_try:

            models_to_try.append(m)



    last_error = None

    tried = []



    for model in models_to_try:

        tried.append(model)

        for attempt in range(3):

            try:

                response = client.models.generate_content(

                    model=model,

                    contents=[build_prompt(), image],

                    config=types.GenerateContentConfig(

                        temperature=0,

                        response_mime_type="application/json",

                    ),

                )



                raw = response.text or ""

                data = extract_json(raw)

                if "tables" not in data:

                    raise ValueError("AI không trả về khóa tables.")



                tables = []

                for t in data.get("tables", []):

                    title = clean_text(t.get("title", ""))

                    columns = [clean_text(c) for c in t.get("columns", [])]

                    rows = t.get("rows", [])



                    norm_rows = []

                    if rows and isinstance(rows[0], dict):

                        if not columns:

                            columns = list(rows[0].keys())

                        for r in rows:

                            norm_rows.append([clean_text(r.get(c, "")) for c in columns])

                    else:

                        for r in rows:

                            if isinstance(r, list):

                                rr = [clean_text(x) for x in r]

                                if columns:

                                    rr = rr[:len(columns)] + [""] * max(0, len(columns) - len(rr))

                                norm_rows.append(rr)



                    if columns and norm_rows:

                        tables.append({"title": title, "columns": columns, "rows": norm_rows})



                raw_with_meta = f"MODEL_USED={model}\nTRIED_MODELS={tried}\n\n{raw}"

                return tables, raw_with_meta



            except Exception as e:

                last_error = e

                msg = str(e)



                # Dịch vụ AI quá tải tạm thời

                if "503" in msg or "500" in msg or "INTERNAL" in msg or "UNAVAILABLE" in msg or "high demand" in msg:

                    time.sleep(3 + attempt * 2)

                    continue



                # Quota/rate/model/key: chuyển model hoặc báo lỗi cuối

                if any(x in msg for x in [

                    "429", "RESOURCE_EXHAUSTED", "Quota",

                    "INVALID_ARGUMENT", "API key not valid",

                    "not found", "404"

                ]):

                    break



                raise



    raise RuntimeError(

        "Không gọi được AI sau khi thử các model: "

        + ", ".join(tried)

        + "\nLỗi cuối: "

        + repr(last_error)

    )







def copy_style_row(ws, src_row, dst_row, max_col):

    """

    Copy format/style từ dòng mẫu sang dòng mới.

    Không copy dữ liệu cũ.

    """

    for c in range(1, max_col + 1):

        src = ws.cell(src_row, c)

        dst = ws.cell(dst_row, c)



        if dst.__class__.__name__ == "MergedCell":

            continue



        try:

            if src.has_style:

                dst._style = copy.copy(src._style)



            dst.font = copy.copy(src.font)

            dst.fill = copy.copy(src.fill)

            dst.border = copy.copy(src.border)

            dst.alignment = copy.copy(src.alignment)

            dst.number_format = src.number_format

            dst.protection = copy.copy(src.protection)

        except Exception:

            pass

    try:

        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height

        ws.row_dimensions[dst_row].hidden = False

        ws.row_dimensions[dst_row].collapsed = False

    except Exception:

        pass





def copy_row_dimension(ws, src_row, dst_row):

    """

    Copy chiều cao dòng và bỏ ẩn dòng.

    Sửa lỗi dòng mới bị dính/sát nhau do row height nhỏ hoặc hidden.

    """

    try:

        src_dim = ws.row_dimensions[src_row]

        dst_dim = ws.row_dimensions[dst_row]

        dst_dim.height = src_dim.height

        dst_dim.hidden = False

        dst_dim.outlineLevel = src_dim.outlineLevel

        dst_dim.collapsed = False

    except Exception:

        pass





def find_no_col_from_headers(excel_headers):

    for col_idx, name in excel_headers:

        if is_no_header(name):

            return col_idx

    return None



def ensure_no_column_in_mapping(source_cols, mapping, excel_headers):

    """

    Cột No/STT trong Excel không lấy từ ảnh nữa.

    Tool tự nối STT từ số cuối. Vì vậy nguồn STT trong ảnh sẽ bỏ qua.

    """

    no_col = find_no_col_from_headers(excel_headers)

    if not no_col:

        return mapping

    out = list(mapping)

    for i, src in enumerate(source_cols):

        if is_no_header(src):

            out[i] = None

    return out





def find_best_source_for_target(target_name, source_cols):

    tn = norm(target_name)

    best_i = None

    best_score = -1

    for i, src in enumerate(source_cols):

        sn = norm(src)

        score = 0

        if sn == tn:

            score = 100

        elif sn in tn or tn in sn:

            score = 90

        elif group_of(src) == group_of(target_name) and len(group_of(src)) > 1:

            score = 95

        else:

            import difflib

            score = difflib.SequenceMatcher(None, sn, tn).ratio() * 70

        if score > best_score:

            best_score = score

            best_i = i

    return best_i if best_score >= 40 else None



def normalize_table_for_template(table, template_name):

    """

    Nếu chọn preset có khung cố định, ép bảng AI về đúng thứ tự cột chuẩn của mẫu đó.

    """

    canonical = CANONICAL_TEMPLATE_COLUMNS.get(template_name)

    if not canonical or not table:

        return table



    src_cols = table.get("columns", [])

    rows = table.get("rows", [])

    idx_map = [find_best_source_for_target(t, src_cols) for t in canonical]



    norm_rows = []

    for row in rows:

        out = []

        for idx in idx_map:

            if idx is None or idx >= len(row):

                out.append("")

            else:

                out.append(row[idx])

        norm_rows.append(out)



    return {

        "title": table.get("title", ""),

        "columns": canonical,

        "rows": norm_rows

    }



def preset_map_columns(source_cols, excel_headers, preset_name):

    preset = TEMPLATE_PRESETS.get(preset_name, {})

    if not preset:

        return auto_map_columns(source_cols, excel_headers)



    excel_norm_by_idx = [(i, col_idx, norm(name), name) for i, (col_idx, name) in enumerate(excel_headers)]

    result = []



    for src in source_cols:

        srcn = norm(src)

        targets = None

        for key, target_list in preset.items():

            kn = norm(key)

            if srcn == kn or srcn in kn or kn in srcn:

                targets = target_list

                break



        if targets == []:

            result.append(None)

            continue



        best = None

        best_score = -1

        if targets:

            target_norms = [norm(t) for t in targets]

            for i, col_idx, exn, name in excel_norm_by_idx:

                score = 0

                for tn in target_norms:

                    if exn == tn:

                        score = max(score, 100)

                    elif tn in exn or exn in tn:

                        score = max(score, 85)

                    else:

                        score = max(score, difflib.SequenceMatcher(None, tn, exn).ratio() * 70)

                if score > best_score:

                    best_score = score

                    best = i

            result.append(best if best_score >= 45 else None)

        else:

            result.append(None)



    return result



def is_total_marker_text(v):

    s = norm(v)

    if not s:

        return False

    return any(m in s for m in TOTAL_MARKERS)



def find_total_row(ws, header_row):

    """

    Tìm dòng TỔNG thật, tránh nhầm header như 'Tổng tổ hợp'.

    Nhận:

    - Dòng có chữ TỔNG/TOTAL đứng riêng ở vài ô đầu.

    - Dòng tổng kết dạng "KL CỌC NHẬP VỀ", "CỘNG", "TỔNG CỘNG" ngoài header.

    - Dòng tổng kết kiểu nhãn ở giữa dòng như "Ép cọc đại trà" kèm các ô SUM/numeric phía sau.

    Không nhận các header cột như 'Tổng tổ hợp', 'Tổng số m cọc...'

    """

    start_row = max(1, header_row + 1)

    def _is_num_or_formula(v):
        if is_formula_value(v):
            return True
        s = str(v or "").strip()
        if not s:
            return False
        try:
            float(s.replace(",", "."))
            return True
        except Exception:
            return False

    def _looks_like_midrow_total(vals):
        non_empty = [v for v in vals if str(v or "").strip()]
        if len(non_empty) < 3:
            return False
        if len(non_empty) > max(8, ws.max_column // 4):
            return False

        first_non_empty_idx = None
        for i, v in enumerate(vals[:6]):
            if str(v or "").strip():
                first_non_empty_idx = i
                break

        # Dòng kiểu tổng kết thường bỏ trống các cột định danh đầu bảng.
        if first_non_empty_idx is None or first_non_empty_idx < 4:
            return False

        summary_cells = sum(1 for v in vals if _is_num_or_formula(v))
        if summary_cells < 2:
            return False

        label_text = " ".join(
            str(v).strip()
            for v in vals[:12]
            if str(v or "").strip() and not _is_num_or_formula(v)
        )
        label_norm = norm(label_text)
        if not label_norm:
            return False

        if any(k in label_norm for k in ["ep coc dai tra", "dai tra", "tong hop", "khoi luong", "nhap ve", "cong", "tong", "sum"]):
            return True

        return summary_cells >= 3 and first_non_empty_idx >= 5

    for r in range(start_row, ws.max_row + 1):

        vals = [str(ws.cell(r, c).value or "").strip() for c in range(1, ws.max_column + 1)]

        non_empty = [v for v in vals if v]

        if not non_empty:

            continue



        first_cells = " ".join(vals[:5])

        first_norm = norm(first_cells)

        row_norm = norm(" ".join(vals))



        # TỔNG thật thường nằm ở vài cột đầu và là chữ riêng

        if re.search(r"(^|\s)(tong|total)(\s|$)", first_norm):

            # loại các header dạng "tong to hop", "tong so m coc" nằm trong header

            if any(x in first_norm for x in ["tong to hop", "tong so m", "tong so met"]):

                continue

            return r



        # Nhận dạng dòng tổng kết đặc thù: "KL CỌC NHẬP VỀ", "CỘNG"

        kl_markers = ["kl coc", "kl cọc", "khoi luong coc", "cong tong", "tong cong"]

        if any(k in row_norm for k in kl_markers):

            return r


        if _looks_like_midrow_total(vals):

            return r



        # fallback: dòng ít ô, có tổng/total rõ ràng

        if len(non_empty) <= max(4, ws.max_column // 5):

            if re.search(r"(^|\s)(tong|total)(\s|$)", row_norm):

                if not any(x in row_norm for x in ["tong to hop", "tong so m", "tong so met"]):

                    return r



    return None



def find_insert_row(ws, header_row, header_cols):

    total_row = find_total_row(ws, header_row)

    if total_row:

        return total_row

    # Fallback: tìm dòng trống đầu tiên sau header

    for r in range(header_row + 1, ws.max_row + 2):

        vals = [str(ws.cell(r, c).value or "").strip() for c in range(1, min(ws.max_column, 10) + 1)]

        if not any(vals):

            return r

    return ws.max_row + 1



def last_number_above(ws, col_idx, before_row, header_row):

    """

    Lấy STT cuối cùng trong vùng trắng phía trên dòng chèn, không lấy vùng xám/chữ ký.

    """

    last = 0

    if not col_idx:

        return last

    for r in range(header_row + 1, before_row):

        if row_has_grey_background(ws, r):

            continue

        v = ws.cell(r, col_idx).value

        try:

            if str(v).strip().isdigit():

                last = int(str(v).strip())

        except Exception:

            pass

    return last



def _is_grey_fill(cell):

    try:

        if not cell.fill or not cell.fill.fill_type:

            return False

        

        # Lấy mã màu RGB

        fg = str(cell.fill.fgColor.rgb or "").upper()

        # openpyxl đôi khi trả về 00000000 hoặc 00FFFFFF cho ô không màu/trắng

        if not fg or fg in {"00000000", "FFFFFFFF", "FF000000", "00FFFFFF"}:

            return False



        # Các mã màu xám phổ biến

        if fg in {"FF808080", "FFC0C0C0", "FFBFBFBF", "FF999999", "FFEAEAEA", "FFF2F2F2"}:

            return True

            

        # Nếu có màu nền (không phải trắng/trong suốt) thì kiểm tra xem có phải vàng/xanh không

        # Ưu tiên coi là vùng non-data (xám) nếu không phải các màu highlight dữ liệu quen thuộc

        if fg not in {"FF00FF00", "FFFF00", "FFFFFF00", "FFCCFFCC"}:

            return True

    except Exception:

        pass

    return False



def _is_grey_fill(cell):
    try:
        if not cell.fill or not cell.fill.fill_type:
            return False

        # File này dùng theme fill cho cả vùng dữ liệu, nên chỉ coi là xám
        # khi openpyxl trả về màu rõ ràng hoặc indexed grey.
        color = cell.fill.fgColor
        ctype = getattr(color, "type", None)

        if ctype == "rgb":
            fg = str(color.rgb or "").upper()
            if not fg or fg in {"00000000", "FFFFFFFF", "FF000000", "00FFFFFF"}:
                return False
            return fg in {"FF808080", "FFC0C0C0", "FFBFBFBF", "FF999999", "FFEAEAEA", "FFF2F2F2"}

        if ctype == "indexed":
            idx = getattr(color, "indexed", None)
            return idx in {22, 23, 24, 25}

        return False
    except Exception:
        return False


def row_has_grey_background(ws, r, max_col=None):

    max_col = max_col or ws.max_column

    grey_count = 0

    checked = 0

    for c in range(1, max_col + 1):

        cell = ws.cell(r, c)

        if cell.__class__.__name__ == "MergedCell":

            continue

        checked += 1

        if _is_grey_fill(cell):

            grey_count += 1

    return checked > 0 and grey_count >= max(2, checked // 3)



def row_is_mostly_blank(ws, r, key_cols):

    vals = [str(ws.cell(r, c).value or "").strip() for c in key_cols]

    return not any(vals)



def find_last_data_row_before(ws, before_row, header_row, key_cols):

    """

    Tìm dòng dữ liệu thật ngay trước dòng TỔNG để copy style.

    Bỏ qua dòng trống/merged/tổng phụ/vùng xám chữ ký.

    """

    for r in range(before_row - 1, header_row, -1):

        vals = [str(ws.cell(r, c).value or "").strip() for c in key_cols]

        text = " ".join(vals)

        if any(vals) and not is_total_marker_text(text) and not row_has_grey_background(ws, r):

            return r

    return max(header_row + 1, before_row - 1)



def find_insert_row_in_white_area(ws, header_row, key_cols):

    """

    Tìm vị trí chèn đúng vùng bảng trắng:

    - Ưu tiên dòng TỔNG/TOTAL.

    - Nếu phía trên TỔNG có dòng trống trong vùng trắng thì điền vào dòng trống đó.

    - Không điền xuống vùng xám/chữ ký.

    """

    total_row = find_total_row(ws, header_row)

    if not total_row:

        # Tìm dòng trống đầu tiên sau header

        for r in range(header_row + 1, ws.max_row + 2):

            if row_is_mostly_blank(ws, r, key_cols):

                return r, None, False

        return ws.max_row + 1, None, False



    # Tìm dòng trống gần nhất phía trên dòng TỔNG trong vùng trắng

    for r in range(total_row - 1, header_row, -1):

        if row_has_grey_background(ws, r):

            continue

        if row_is_mostly_blank(ws, r, key_cols):

            return r, total_row, False



    # Không có dòng trống thì chèn ngay trên TỔNG

    return total_row, total_row, True



def is_no_header(name):

    n = norm(name)

    # Chỉ nhận No/STT thật, tránh nhầm Note

    if n in {"no", "stt", "so thu tu", "tt", "stt no", "no stt", "stt no", "no.", "số tt", "số thứ tự"}:

        return True

    if n.startswith("stt ") or n.endswith(" stt") or "stt" == n:

        return True

    if n.startswith("no ") or n.endswith(" no") or "no" == n:

        return True

    return False



def is_row_total_header(name):

    n = norm(name)

    return any(x in n for x in ["chieu dai to hop", "total m", "total", "length of pile"])



def is_summary_sum_header(name):

    n = norm(name)

    return any(x in n for x in [

        "chieu dai to hop", "total m", "total",

        "chieu sau ep thuc te", "pressing depth",

        "chieu dai coc", "length of pile"

    ])



def is_actual_pressing_depth_header(name):

    n = norm(name)

    return any(x in n for x in [

        "chieu sau ep thuc te",
        "pressing depth",
        "actual depth",
        "chieu sau thuc te",
        "ep thuc te",
    ])


def is_segment_header(name):

    n = norm(name)

    return n in {"d1", "d2", "d3", "d4", "d5", "d6", "1st", "2nd", "3rd", "4th", "5th", "6th", "đ1", "đ2", "đ3", "đ4", "đ5", "đ6"}



def find_first_data_row(ws, header_row, no_col, total_row):

    """

    Tìm dòng dữ liệu đầu tiên để tính tổng.

    Ưu tiên cột STT/No có số, bỏ qua header/blank.

    """

    if no_col:

        for r in range(header_row + 1, total_row):

            v = ws.cell(r, no_col).value

            try:

                if str(v).strip().isdigit():

                    return r

            except Exception:

                pass

    return header_row + 1



def excel_col_letter(col_idx):

    return get_column_letter(col_idx)





def is_formula_value(v):

    return isinstance(v, str) and v.startswith("=")



def translate_formula_to_row(formula, from_cell, to_cell):

    try:

        return Translator(formula, origin=from_cell).translate_formula(to_cell)

    except Exception:

        return formula



def capture_formula_columns(ws, row_idx):

    """

    Lấy tất cả cột có công thức ở một dòng.

    Dùng để biết cột nào cần tự sum/tự công thức, không hard-code.

    """

    cols = []

    if not row_idx:

        return cols

    for c in range(1, ws.max_column + 1):

        v = ws.cell(row_idx, c).value

        if is_formula_value(v):

            cols.append(c)

    return cols



def capture_total_sum_columns(ws, total_row, first_data_row=None, last_data_row=None):

    """

    Dòng TỔNG cần SUM cột nào:

    - Nếu ô dòng TỔNG có công thức: lấy.

    - Nếu ô dòng TỔNG có số và cột đó có số ở vùng dữ liệu: cũng lấy.

    """

    cols = []

    def _is_contract_like_header(name):
        n = norm(name)
        if not n:
            return False
        return any(tok in n for tok in ["hop dong", "contract", "contract no", "so hop dong", "hd"])

    header_by_col = {}
    try:
        header_row = find_header_row_smart(ws)
        header_by_col = {int(c): str(v or "") for c, v in get_headers_smart(ws, header_row)}
    except Exception:
        header_by_col = {}

    if not total_row:

        return cols

    for c in range(1, ws.max_column + 1):
        if _is_contract_like_header(header_by_col.get(c, "")):
            continue
        v = ws.cell(total_row, c).value

        if is_formula_value(v):

            cols.append(c)

            continue

        sv = str(v or "").strip().replace(".", "").replace(",", ".")

        try:

            if sv != "":

                float(sv)

                found_num = False

                if first_data_row and last_data_row and last_data_row >= first_data_row:

                    for r in range(first_data_row, last_data_row + 1):

                        vv = str(ws.cell(r, c).value or "").strip().replace(".", "").replace(",", ".")

                        try:

                            float(vv)

                            found_num = True

                            break

                        except Exception:

                            pass

                else:

                    found_num = True

                if found_num:

                    cols.append(c)

        except Exception:

            pass

    return sorted(set(cols))



def apply_row_formulas_from_template(ws, template_row, dst_row):

    """

    Dòng mới cần công thức gì thì lấy theo dòng mẫu phía trên.

    Ví dụ:

    A16 = A15+1  -> A17 = A16+1

    L16 = SUM(F16:K16) -> L17 = SUM(F17:K17)

    """

    for c in range(1, ws.max_column + 1):

        src = ws.cell(template_row, c)

        dst = ws.cell(dst_row, c)

        if dst.__class__.__name__ == "MergedCell":

            continue

        if is_formula_value(src.value):

            dst.value = translate_formula_to_row(

                src.value,

                f"{excel_col_letter(c)}{template_row}",

                f"{excel_col_letter(c)}{dst_row}"

            )



def set_total_formulas_by_template(ws, total_row, formula_cols, first_data_row, last_data_row):

    """

    Dòng TỔNG cần sum cột nào thì dựa vào chính file mẫu:

    cột nào ở dòng TỔNG đang có công thức thì đặt SUM lại cho cột đó.

    Không hard-code chỉ 1-2 cột.

    """

    if not total_row or last_data_row < first_data_row:

        return

    for c in formula_cols:

        cell = ws.cell(total_row, c)

        if cell.__class__.__name__ == "MergedCell":

            continue

        col = excel_col_letter(c)

        cell.value = f"=SUM({col}{first_data_row}:{col}{last_data_row})"



def _update_sum_formula_range(formula, col_letter, first_data_row, last_data_row):

    """

    Nếu công thức là SUM thì cập nhật lại range theo dòng dữ liệu thực tế.
    Không đụng các công thức khác.
    """

    if not is_formula_value(formula):

        return None

    s = str(formula or "").replace(" ", "")
    col = str(col_letter).upper()

    m = re.match(
        r"^=SUM\((\$?" + re.escape(col) + r"\$?\d+):(\$?" + re.escape(col) + r"\$?\d+)\)$",
        s,
        flags=re.I,
    )
    if m:
        return f"=SUM({col}{first_data_row}:{col}{last_data_row})"

    m = re.match(
        r"^=\+?SUM\((.*)\)$",
        s,
        flags=re.I,
    )
    if m:
        return f"=SUM({col}{first_data_row}:{col}{last_data_row})"

    return None



def update_total_formulas(ws, total_row, first_data_row, last_data_row, excel_headers=None, no_col=None):

    """

    Cập nhật công thức dòng TỔNG theo vùng dữ liệu thực tế.

    - Nếu ô TỔNG đã có SUM: giữ công thức, đổi range.
    - Nếu ô TỔNG đang trống nhưng cột có số liệu: tạo SUM mới.
    - Không tính STT.
    """

    if not total_row or not first_data_row or not last_data_row or last_data_row < first_data_row:

        return []

    updated = []
    header_by_col = {int(c): str(name or "") for c, name in (excel_headers or [])}

    def _header_is_numeric_candidate(name):

        n = norm(name)
        if not n:
            return False
        if any(tok in n for tok in ["hop dong", "contract", "contract no", "so hop dong", "hd"]):
            return False
        if any(tok in n for tok in ["stt", "no", "ngay", "gio", "bat dau", "ket thuc", "ghi chu", "ten", "loai", "ca"]):
            return False
        return any(tok in n for tok in ["chieu dai", "chieu sau", "tai trong", "khoi luong", "do sau", "khoi luong ep", "ep thuc te", "tong hop", "do dai", "m)", "(m)", "(t)"]) or is_actual_pressing_depth_header(n)

    def _col_has_numeric_data(col_idx):

        hits = 0
        for r in range(first_data_row, last_data_row + 1):
            v = ws.cell(r, col_idx).value
            if v in (None, ""):
                continue
            if is_formula_value(v):
                hits += 1
                continue
            if isinstance(v, (int, float)):
                hits += 1
                continue
            s = str(v).strip().replace(".", "").replace(",", ".")
            try:
                float(s)
                hits += 1
            except Exception:
                pass
        return hits

    candidate_cols = []
    for c in range(1, ws.max_column + 1):
        if c == no_col:
            continue
        cell = ws.cell(total_row, c)
        if is_formula_value(cell.value):
            candidate_cols.append(c)

    if not candidate_cols:
        for c in range(1, ws.max_column + 1):
            if c == no_col:
                continue
            if _header_is_numeric_candidate(header_by_col.get(c, "")) and _col_has_numeric_data(c) > 0:
                candidate_cols.append(c)

    if not candidate_cols:
        for c in range(1, ws.max_column + 1):
            if c == no_col:
                continue
            if _col_has_numeric_data(c) > 0:
                candidate_cols.append(c)

    for c in sorted(set(candidate_cols)):
        cell = ws.cell(total_row, c)
        existing = cell.value
        col_letter = get_column_letter(c)

        new_formula = None
        if is_formula_value(existing):
            new_formula = _update_sum_formula_range(existing, col_letter, first_data_row, last_data_row)
            if new_formula is None and str(existing).strip().upper().startswith("=SUM("):
                new_formula = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
        elif _header_is_numeric_candidate(header_by_col.get(c, "")) or _col_has_numeric_data(c) > 0:
            new_formula = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"

        if new_formula:
            cell.value = new_formula
            updated.append(c)

    return updated







def group_of(s):

    ns = norm(s)

    groups = globals().get("SYNONYM_GROUPS", [])

    for g in groups:

        gn = {norm(x) for x in g}

        if ns in gn:

            return gn

    return {ns}



def auto_map_columns(source_cols, excel_headers):

    """

    Auto map động từ cột ảnh -> cột Excel.

    Không dùng mẫu cố định.

    Mỗi cột Excel chỉ nhận 1 cột ảnh để tránh đè dữ liệu.

    """

    result = [None] * len(source_cols)

    used_excel_idx = set()



    excel_names = [h for _, h in excel_headers]

    excel_norms = [norm(h) for h in excel_names]

    candidates = []



    for s_idx, src in enumerate(source_cols):

        srcn = norm(src)

        src_group = group_of(src)



        if is_no_header(src):

            continue



        for e_idx, ex in enumerate(excel_names):

            exn = excel_norms[e_idx]

            ex_group = group_of(ex)

            score = 0.0

            if exn == srcn:

                score = 100

            elif src_group == ex_group and len(src_group) > 1:

                score = 95

            elif srcn in exn or exn in srcn:

                score = 85

            else:

                score = difflib.SequenceMatcher(None, srcn, exn).ratio() * 70



            if srcn in {"d1","d2","d3","d4","d5","d6"} and exn == srcn:

                score = 110

            if srcn in {"bat dau", "gio bat dau", "start"} and ("bat dau" in exn or exn == "start"):

                score = max(score, 105)

            if srcn in {"ket thuc", "gio ket thuc", "end"} and ("ket thuc" in exn or exn == "end"):

                score = max(score, 105)

            candidates.append((score, s_idx, e_idx))



    candidates.sort(reverse=True, key=lambda x: x[0])

    for score, s_idx, e_idx in candidates:

        if score < 45:

            continue

        if result[s_idx] is not None:

            continue

        if e_idx in used_excel_idx:

            continue

        result[s_idx] = e_idx

        used_excel_idx.add(e_idx)



    return result





def find_last_stt_in_white_area(ws, no_col, header_row, total_row):

    """

    Lấy STT cuối cùng trong vùng trắng phía trên dòng TỔNG.

    Không lấy vùng xám/chữ ký.

    """

    if not no_col or not total_row:

        return 0

    last = 0

    for r in range(header_row + 1, total_row):

        if row_has_grey_background(ws, r):

            continue

        v = ws.cell(r, no_col).value

        try:

            if str(v).strip().isdigit():

                last = int(str(v).strip())

        except Exception:

            pass

    return last



def find_first_data_row_for_sum(ws, no_col, header_row, total_row):

    """

    Dòng đầu để SUM: dòng đầu trong vùng trắng có STT số.

    """

    if no_col:

        for r in range(header_row + 1, total_row):

            if row_has_grey_background(ws, r):

                continue

            v = ws.cell(r, no_col).value

            try:

                if str(v).strip().isdigit():

                    return r

            except Exception:

                pass

    return header_row + 1











def get_stt_value(ws, row, col, memo=None, depth=0):

    """

    Đọc STT thật trong cột STT.

    Hỗ trợ:

    - Số trực tiếp: 27

    - Chữ kèm số: "1.", "01", "(1)"

    - Công thức đơn giản: =A16+1, =A16 + 1, =+A16+1

    """

    if memo is None:

        memo = {}

    key = (row, col)

    if key in memo:

        return memo[key]

    if depth > 50:

        return None



    v = ws.cell(row, col).value

    if v is None:

        return None

        

    if isinstance(v, (int, float)):

        memo[key] = int(v)

        return memo[key]



    s = str(v).strip()

    if not s:

        return None



    # Thử parse số trực tiếp

    if s.isdigit():

        memo[key] = int(s)

        return memo[key]



    # Dạng số có dấu chấm/ngoặc: "1.", "(1)", "1/"

    m_num = re.search(r"^\(?(\d+)\)?[\./]?$", s)

    if m_num:

        memo[key] = int(m_num.group(1))

        return memo[key]



    # Dạng số float nhưng hiển thị nguyên

    try:

        f_val = float(s.replace(",", "."))

        if f_val.is_integer():

            memo[key] = int(f_val)

            return memo[key]

    except Exception:

        pass



    # Dạng công thức: =A16+1 hoặc =+A16+1

    if s.startswith("="):

        f = s.replace(" ", "")

        if f.startswith("=+"):

            f = "=" + f[2:]

        

        # Regex cho =A16+1

        m = re.fullmatch(r"=([A-Z]+)(\d+)\+(\d+)", f, flags=re.I)

        if m:

            ref_col_letters = m.group(1).upper()

            ref_row = int(m.group(2))

            plus = int(m.group(3))

            try:

                ref_col = coordinate_to_tuple(ref_col_letters + "1")[1]

            except Exception:

                ref_col = col

            base = get_stt_value(ws, ref_row, ref_col, memo, depth + 1)

            if base is not None:

                memo[key] = base + plus

                return memo[key]



    return None





def find_all_stt_chains(ws, no_col, header_row, total_row):

    """

    Tìm tất cả chuỗi STT liên tục trước dòng TỔNG.

    Đọc được cả STT là công thức =A16+1.

    Hỗ trợ gộp các chuỗi bị đứt đoạn nhẹ (thiếu 1-2 số).

    """

    if not no_col or not total_row:

        return []



    seq = []

    memo = {}

    for r in range(header_row + 1, total_row):

        try:

            if row_has_grey_background(ws, r):

                continue

        except Exception:

            pass



        n = get_stt_value(ws, r, no_col, memo)

        if isinstance(n, int):

            seq.append((r, n))



    if not seq:

        return []



    # Gom nhóm các số liên tục hoặc gần liên tục

    chains = []

    cur = [seq[0]]

    for item in seq[1:]:

        # Nếu là số tiếp theo (+1) hoặc nhảy nhẹ (+2) nhưng dòng cũng gần nhau

        val_gap = item[1] - cur[-1][1]

        row_gap = item[0] - cur[-1][0]

        

        if val_gap == 1 or (0 < val_gap <= 3 and 0 < row_gap <= 3):

            cur.append(item)

        else:

            if cur:

                chains.append(cur)

            cur = [item]

    if cur:

        chains.append(cur)

        

    return chains


def score_stt_column_candidate(ws, col, header_row, total_row):

    """

    Chấm điểm một cột xem có giống cột STT/No hay không.

    Hàm này chịu được trường hợp người dùng đã xóa/cắt vài dòng,
    khiến chuỗi số không còn liền mạch tuyệt đối.
    """

    if not col or not total_row:

        return 0, []



    memo = {}

    rows = []

    for r in range(header_row + 1, total_row):

        try:

            if row_has_grey_background(ws, r):

                continue

        except Exception:

            pass

        try:

            row_text = " ".join(str(ws.cell(r, c).value or "") for c in range(1, ws.max_column + 1))

            if is_total_marker_text(row_text):

                continue

        except Exception:

            pass

        n = None

        try:

            n = get_stt_value(ws, r, col, memo)

        except Exception:

            n = None

        if not isinstance(n, int):

            s = str(ws.cell(r, col).value or "").strip()

            if s.isdigit():

                n = int(s)

            else:

                try:

                    f = float(s.replace(",", "."))

                    if f.is_integer():

                        n = int(f)

                except Exception:

                    pass

        if isinstance(n, int):

            rows.append((r, n))



    if not rows:

        return 0, []



    rows.sort(key=lambda x: x[0])

    longest_run = 1

    current_run = 1

    increasing_steps = 0

    small_gap_steps = 0

    for prev, cur in zip(rows, rows[1:]):

        prev_row, prev_val = prev

        cur_row, cur_val = cur

        row_gap = cur_row - prev_row

        val_gap = cur_val - prev_val

        if val_gap > 0:

            increasing_steps += 1

        if 0 < row_gap <= 8 and 0 < val_gap <= 8:

            small_gap_steps += 1

            current_run += 1

            longest_run = max(longest_run, current_run)

        else:

            current_run = 1



    first_row = rows[0][0]

    last_val = rows[-1][1]

    numeric_count = len(rows)

    score = numeric_count * 12 + longest_run * 20 + increasing_steps * 3 + small_gap_steps * 2

    if first_row <= header_row + 3:

        score += 8

    if rows[0][1] in (0, 1):

        score += 6

    if last_val >= numeric_count:

        score += 4



    return score, rows



def merge_contiguous_stt_chains(chains):

    """
    Ghép các chuỗi STT tách rời nhưng vẫn nối tiếp nhau theo số.

    Ví dụ: 1->37 rồi sau đó 38->40 sẽ được ghép thành một chuỗi.
    """

    if not chains:
        return []

    merged = [list(chains[0])]

    for chain in chains[1:]:
        if not chain:
            continue

        try:
            curr_first_row, curr_first_val = chain[0]
        except Exception:
            merged.append(list(chain))
            continue

        # Không chỉ nối với chuỗi cuối cùng:
        # nếu có một dòng rác / dòng tổng / block phụ chen giữa,
        # vẫn phải nhận ra chuỗi STT tiếp nối (vd. 37 -> 38).
        merge_idx = None

        for idx in range(len(merged) - 1, -1, -1):
            prev = merged[idx]
            try:
                prev_last_row, prev_last_val = prev[-1]
            except Exception:
                continue

            if (
                isinstance(prev_last_val, int)
                and isinstance(curr_first_val, int)
                and curr_first_val == prev_last_val + 1
                and curr_first_row >= prev_last_row
            ):
                merge_idx = idx
                break

        if merge_idx is not None:
            merged[merge_idx].extend(chain)
            continue

        merged.append(list(chain))

    return [list(ch) for ch in merged]



def select_longest_stt_chain(ws, no_col, header_row, total_row):

    """

    Chọn chuỗi STT chuẩn:

    - Ưu tiên chuỗi dài nhất.

    - Nếu bằng nhau, ưu tiên chuỗi bắt đầu nhỏ hơn.

    - Nếu vẫn bằng, ưu tiên STT cuối lớn hơn.

    """

    chains = merge_contiguous_stt_chains(find_all_stt_chains(ws, no_col, header_row, total_row))

    if not chains:

        return None

    return sorted(chains, key=lambda ch: (len(ch), -ch[0][1], ch[-1][1]), reverse=True)[0]



def find_stt_sequence_region(ws, no_col, header_row, total_row):

    """

    Lấy chuỗi STT liên tục dài nhất trước dòng TỔNG.

    Không lấy dòng rác 19,20,21 dưới bảng.

    """

    best = select_longest_stt_chain(ws, no_col, header_row, total_row)

    if not best:

        return None, None, 0

    return best[0][0], best[-1][0], best[-1][1]



def get_row_values_nonempty(ws, row, cols):

    return [str(ws.cell(row, c).value or "").strip() for c in cols]







def find_no_column_smart(ws, excel_headers, header_row, total_row):

    """

    Tìm cột STT/No chắc hơn:

    1. Theo tên header đã đọc.

    2. Quét vùng header 3 dòng đầu xem có chữ STT/No.

    3. Nếu header đọc sai, tự chọn cột có chuỗi số liên tục dài nhất trước dòng TỔNG.

    """

    # 1. Theo header

    for col_idx, name in excel_headers:

        if is_no_header(name):

            return col_idx



    # 2. Quét trực tiếp vài dòng header

    max_r = min(ws.max_row, header_row + 3)

    for c in range(1, min(ws.max_column, 8) + 1):

        texts = []

        for r in range(max(1, header_row - 1), max_r + 1):

            texts.append(str(ws.cell(r, c).value or ""))

        joined = norm(" ".join(texts))

        if is_no_header(joined):

            return c

        # cell A đôi khi là "STT" ở dòng trên và "No" ở dòng dưới

        if "stt" in joined and ("no" in joined or c == 1):

            return c



    # 3. Fallback: cột có chuỗi số liên tục dài nhất trước TỔNG, ưu tiên cột bên trái

    best_col = None

    best_len = 0

    best_last = 0

    best_score = 0

    scan_max_col = min(ws.max_column, 8)

    for c in range(1, scan_max_col + 1):

        chains = find_all_stt_chains(ws, c, header_row, total_row)

        if not chains:

            score, rows = score_stt_column_candidate(ws, c, header_row, total_row)

            if score > best_score:

                best_col = c

                best_len = len(rows)

                best_last = rows[-1][1] if rows else 0

                best_score = score

            continue

        best_chain = sorted(chains, key=lambda ch: (len(ch), ch[-1][1]), reverse=True)[0]

        ln = len(best_chain)

        last = best_chain[-1][1]

        score, rows = score_stt_column_candidate(ws, c, header_row, total_row)

        if (
            score > best_score
            or (score == best_score and ln > best_len)
            or (score == best_score and ln == best_len and best_col is not None and c < best_col)
        ):

            best_col = c

            best_len = ln

            best_last = last

            best_score = score



    if best_col and (best_len >= 3 or best_score >= 40):

        return best_col



    return None







def convert_excel_value(value):

    """

    Chuyển dữ liệu OCR dạng số thành số thật để công thức SUM tính được.

    Giữ nguyên ngày, giờ, mã cọc, loại cọc, vị trí.

    """

    if value is None:

        return ""

    s = str(value).strip()

    if s == "":

        return ""



    # giữ nguyên ngày, giờ, mã có chữ

    lower = s.lower()

    if "/" in s or "h" in lower:

        return s

    if any(ch.isalpha() for ch in s):

        return s



    # bỏ khoảng trắng, đổi dấu phẩy thập phân sang dấu chấm

    t = s.replace(" ", "").replace(",", ".")

    # bỏ dấu + dư

    if t.startswith("+"):

        t = t[1:]



    try:

        if t.count(".") <= 1:

            num = float(t)

            if num.is_integer():

                return int(num)

            return num

    except Exception:

        pass



    return s



def normalize_vietnam_date(value):

    """

    Chuẩn hóa ngày về dd/mm/yyyy để phiếu cọc ghi đúng kiểu Việt Nam.

    Chỉ xử lý các dạng ngày phổ biến; nếu không nhận ra thì trả nguyên chuỗi.
    """

    if value is None:

        return ""

    if isinstance(value, datetime):

        return value.strftime("%d/%m/%Y")

    try:

        from datetime import date

        if isinstance(value, date):

            return value.strftime("%d/%m/%Y")

    except Exception:

        pass

    s = str(value).strip()

    if not s:

        return ""

    s = s.split(" ")[0].strip()

    m = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$", s)

    if m:

        y, mo, d = m.groups()

        return f"{int(d):02d}/{int(mo):02d}/{y}"

    m = re.match(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{4})$", s)

    if m:

        d, mo, y = m.groups()

        return f"{int(d):02d}/{int(mo):02d}/{y}"

    return s



def normalize_numeric_like_text(value):

    """

    Lấy số đầu tiên trong một chuỗi OCR có thể dính đơn vị/khoảng trắng.

    Giữ nguyên nếu không tìm thấy số hợp lệ.

    """

    if value is None:

        return ""

    s = str(value).strip()

    if s == "":

        return ""

    if not re.search(r"\d", s):

        return s



    cleaned = s.replace("\u00a0", " ").replace(" ", "")

    matches = re.findall(r"[-+]?\d[\d.,]*", cleaned)

    if not matches:

        return convert_excel_value(s)



    token = max(matches, key=lambda x: sum(ch.isdigit() for ch in x))

    token = token.strip(".,")



    if not token:

        return convert_excel_value(s)



    if "," in token and "." in token:

        last_comma = token.rfind(",")

        last_dot = token.rfind(".")

        decimal_sep = "," if last_comma > last_dot else "."

        thousand_sep = "." if decimal_sep == "," else ","

        token = token.replace(thousand_sep, "")

        token = token.replace(decimal_sep, ".")

    elif token.count(",") == 1 and token.count(".") == 0:

        token = token.replace(",", ".")

    elif token.count(".") == 1 and token.count(",") == 0:

        left, right = token.split(".")

        if len(right) == 3 and left.lstrip("+-").isdigit() and len(left.lstrip("+-")) > 1:

            token = left + right

    elif token.count(",") > 1 and "." not in token:

        token = token.replace(",", "")

    elif token.count(".") > 1 and "," not in token:

        token = token.replace(".", "")



    try:

        f = float(token)

        if f.is_integer():

            return int(f)

        return f

    except Exception:

        return token.replace(".", ",")



def _static_jacking_to_float(value):

    if value is None:

        return None

    if isinstance(value, (int, float)):

        return float(value)

    s = str(value).strip()

    if not s:

        return None

    m = re.search(r"[-+]?\d[\d.,]*", s.replace("\u00a0", " "))
    if not m:

        return None

    token = m.group(0)

    if "," in token and "." in token:

        last_comma = token.rfind(",")

        last_dot = token.rfind(".")

        decimal_sep = "," if last_comma > last_dot else "."

        thousand_sep = "." if decimal_sep == "," else ","

        token = token.replace(thousand_sep, "")

        token = token.replace(decimal_sep, ".")

    elif token.count(",") == 1 and token.count(".") == 0:

        token = token.replace(",", ".")

    elif token.count(",") > 1 and "." not in token:

        token = token.replace(",", "")

    elif token.count(".") > 1 and "," not in token:

        token = token.replace(".", "")

    try:

        return float(token)

    except Exception:

        return None


def _summary_date_sort_key(date_text):

    s = str(date_text or "").strip()

    if not s:

        return (1, "")

    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):

        try:

            return (0, datetime.strptime(s, fmt))

        except Exception:

            pass

    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)

    if m:

        d, mo, y = m.groups()

        try:

            return (0, datetime(int(y), int(mo), int(d)))

        except Exception:

            pass

    m = re.match(r"^(\d{1,2})/(\d{1,2})$", s)

    if m:

        d, mo = m.groups()

        try:

            return (0, datetime(datetime.now().year, int(mo), int(d)))

        except Exception:

            pass

    return (1, s)


def build_static_jacking_daily_summary_lines(tables):

    summary_lines = []

    if not tables:

        return summary_lines

    def _is_match(name, aliases):

        n = norm(name)

        return any(alias in n for alias in aliases)

    def _find_col(cols, aliases):

        for idx, col in enumerate(cols):

            if _is_match(col, aliases):

                return idx

        return None

    def _score_date_cell(value):

        s = str(value or "").strip()

        if not s:

            return 0

        if normalize_vietnam_date(s) != s:

            return 4

        if re.search(r"\b\d{1,2}[/-]\d{1,2}([/-]\d{2,4})?\b", s):

            return 3

        if re.search(r"\b20\d{2}\b", s) and re.search(r"\b\d{1,2}\b", s):

            return 2

        return 0

    def _score_depth_cell(value):

        s = str(value or "").strip()

        if not s:

            return 0

        if re.search(r"\d", s) and ("m" in norm(s) or "." in s or "," in s):

            return 3

        if _static_jacking_to_float(s) is not None:

            return 2

        return 0

    def _score_pile_cell(value):

        s = str(value or "").strip()

        if not s:

            return 0

        if re.search(r"\d", s) and not re.search(r"\b\d{1,2}[/-]\d{1,2}\b", s):

            return 2

        if len(s) >= 2 and any(ch.isalpha() for ch in s):

            return 1

        return 0

    def _normalize_summary_date(value):

        s = str(value or "").strip()

        if not s:

            return ""

        normalized = normalize_vietnam_date(s)

        if re.search(r"\b\d{1,2}[/-]\d{1,2}([/-]\d{2,4})?\b", str(normalized)):

            return normalized

        return ""

    def _find_by_data_score(rows, cols_len, scorer, avoid=None):

        avoid = set(avoid or [])
        best_idx = None
        best_score = 0
        for idx in range(cols_len):
            if idx in avoid:
                continue
            score = 0
            for row in rows:
                if idx < len(row):
                    score += scorer(row[idx])
            if score > best_score:
                best_idx = idx
                best_score = score
        return best_idx, best_score

    def _header_score(name, aliases, bonus_words=()):

        n = norm(name)
        score = 0
        for alias in aliases:
            if alias in n:
                score += 10 + min(6, len(alias) // 4)
        for word in bonus_words:
            if word in n:
                score += 4
        return score

    def _find_smart_col(cols, rows, aliases, scorer, avoid=None, bonus_words=(), min_score=5):

        avoid = set(avoid or [])
        best_idx = None
        best_score = 0
        for idx, col in enumerate(cols):
            if idx in avoid:
                continue
            score = _header_score(col, aliases, bonus_words=bonus_words)
            for row in rows[:20]:
                if idx < len(row):
                    score += scorer(row[idx])
            if score > best_score:
                best_idx = idx
                best_score = score
        if best_idx is not None and best_score >= min_score:
            return best_idx
        return None

    def _infer_date_from_table(candidate, rows):

        probes = [
            candidate.get("title"),
            candidate.get("name"),
            candidate.get("sheet"),
            candidate.get("caption"),
        ]
        for row in rows[:8]:
            probes.extend(row[: min(len(row), 8)])
        for value in probes:
            date_value = _normalize_summary_date(value)
            if date_value:
                return date_value
        return ""

    grouped = {}
    matched_tables = 0
    source_image_count = 0

    for candidate in tables:

        if not isinstance(candidate, dict):

            continue

        cols = [str(c or "").strip() for c in (candidate.get("columns") or [])]
        rows = [list(r) if isinstance(r, (list, tuple)) else [r] for r in (candidate.get("rows") or [])]
        if not cols or not rows:

            continue

        date_aliases = ("ngay ep", "ngay thi cong", "ngay thi cong ep", "ngay", "date", "jacking date")
        pile_aliases = ("ten tim coc", "ten coc", "ma coc", "so hieu coc", "pile name", "pile no", "ten tim", "tim coc", "tim", "pile")
        depth_aliases = (
            "chieu sau ep thuc te",
            "chieu sau thuc te",
            "chieu sau ep tt",
            "sau ep thuc te",
            "do sau thuc te",
            "do sau ep thuc te",
            "chieu sau ep",
            "do sau ep",
            "chieu sau",
            "pressing depth actual",
            "actual pressing depth",
            "actual depth",
            "ep thuc te",
            "depth",
        )

        date_idx = _find_smart_col(cols, rows, date_aliases, _score_date_cell, bonus_words=("ngay", "date"))
        pile_idx = _find_smart_col(cols, rows, pile_aliases, _score_pile_cell, avoid={date_idx} - {None}, bonus_words=("tim", "coc", "pile"))
        depth_idx = _find_smart_col(
            cols,
            rows,
            depth_aliases,
            _score_depth_cell,
            avoid={date_idx, pile_idx} - {None},
            bonus_words=("thuc te", "chieu sau", "do sau", "depth", "ep"),
        )

        if date_idx is None:
            date_idx, _ = _find_by_data_score(rows, len(cols), _score_date_cell)
        if pile_idx is None:
            pile_idx, _ = _find_by_data_score(rows, len(cols), _score_pile_cell, avoid={date_idx} if date_idx is not None else set())
        if depth_idx is None:
            depth_idx, _ = _find_by_data_score(rows, len(cols), _score_depth_cell, avoid={date_idx, pile_idx} - {None})

        if date_idx is None or pile_idx is None or depth_idx is None:

            continue

        matched_tables += 1
        try:

            source_image_count = max(source_image_count, int(candidate.get("_source_image_count") or 0))

        except Exception:

            pass

        for row in rows:

            if pile_idx >= len(row) or date_idx >= len(row) or depth_idx >= len(row):

                continue

            pile_value = str(row[pile_idx] or "").strip()

            if not pile_value:

                continue

            date_value = _normalize_summary_date(row[date_idx]) if date_idx < len(row) else ""

            if not date_value:

                date_value = _infer_date_from_table(candidate, rows) or "Không rõ ngày"

            depth_num = _static_jacking_to_float(row[depth_idx])

            bucket = grouped.setdefault(date_value, {"tim": 0, "depth": 0.0})

            bucket["tim"] += 1

            if depth_num is not None:

                bucket["depth"] += depth_num

    if not matched_tables:

        return [
            "Chưa đủ cột rõ ràng để tổng hợp chính xác.",
            "Cần tối thiểu: cột ngày ép, cột tên tim/cọc, và cột chiều sâu ép thực tế.",
        ]

    if not grouped:

        return ["Chưa có dòng hợp lệ để tổng hợp."]

    if source_image_count > 0:

        summary_lines.append(f"Tổng hợp theo ngày từ {source_image_count} ảnh đã đọc")

    else:

        summary_lines.append("Tổng hợp theo ngày từ toàn bộ dữ liệu đã đọc")

    for date_value, bucket in sorted(grouped.items(), key=lambda item: _summary_date_sort_key(item[0])):

        total_depth = bucket["depth"]

        if abs(total_depth - round(total_depth)) < 1e-9:

            depth_text = str(int(round(total_depth)))

        else:

            depth_text = str(round(total_depth, 3)).rstrip("0").rstrip(".")

        summary_lines.append(f"{date_value}: {bucket['tim']} tim - {depth_text} m")

    return summary_lines



def force_workbook_recalculate(wb):

    """

    Bắt Excel/WPS tính lại công thức khi mở file.

    """

    try:

        wb.calculation.fullCalcOnLoad = True

        wb.calculation.forceFullCalc = True

        wb.calculation.calcMode = "auto"

    except Exception:

        pass

    try:

        wb.calculation_properties.fullCalcOnLoad = True

        wb.calculation_properties.forceFullCalc = True

        wb.calculation_properties.calcMode = "auto"

    except Exception:

        pass







def find_header_row_smart(ws):

    """

    Tìm header thật cho file Excel:

    - Bỏ qua dòng title merge (nhiều cột cùng 1 giá trị như 'BẢNG THỐNG KÊ...').

    - Ưu tiên dòng có nhiều cột riêng biệt: STT, Ngày, Tên cọc, Loại cọc...

    - Không chọn dòng dữ liệu số.

    """

    best_row = 1

    best_score = -999

    max_r = min(ws.max_row, 40)

    max_c = min(ws.max_column, 30)



    for r in range(1, max_r + 1):

        # Lấy giá trị từng ô riêng biệt trong dòng này

        cell_vals = []

        for c in range(1, max_c + 1):

            v = str(ws.cell(r, c).value or "").strip()

            cell_vals.append(v)



        non_empty = [v for v in cell_vals if v]

        if not non_empty:

            continue



        # --- PHẠT NẶNG dòng title merge ---

        # Nếu tất cả ô không rỗng đều có cùng 1 giá trị → dòng merge tiêu đề

        unique_vals = set(non_empty)

        if len(unique_vals) == 1 and len(non_empty) >= 3:

            continue  # bỏ hoàn toàn dòng này



        # Nếu > 70% ô có cùng giá trị → khả năng cao là merge title

        most_common_count = max(non_empty.count(v) for v in unique_vals)

        merge_ratio = most_common_count / len(non_empty) if non_empty else 0

        if merge_ratio > 0.7 and len(non_empty) >= 3:

            continue  # bỏ dòng merge title



        # Tính điểm dựa trên các keyword cột thực

        text = " ".join(cell_vals)

        # Cũng lấy thêm vài dòng tiếp theo để hỗ trợ header đa tầng

        extra = " ".join(

            str(ws.cell(rr, c).value or "")

            for rr in range(r + 1, min(r + 3, ws.max_row) + 1)

            for c in range(1, max_c + 1)

        )

        n = norm(text + " " + extra)

        score = 0



        # Thưởng cho các keyword cột thực

        col_keywords = [

            "stt", "no", "ngay", "date", "ten coc", "pile", "loai coc",

            "vi tri", "d1", "d2", "d3", "d4", "d5", "d6", "tong", "total",

            "so hop dong", "so phieu", "xe van chuyen", "chieu dai", "so luong",

            "khoi luong", "ghi chu", "note", "bat dau", "ket thuc", "luc ep",

            "hop dong", "chung loai", "mui", "ky hieu"

        ]

        for kw in col_keywords:

            if kw in n:

                score += 10



        # Thưởng thêm cho số lượng cột riêng biệt (header đa cột riêng biệt)

        score += min(len(unique_vals), 15) * 3



        # Phạt dòng dữ liệu số nhiều

        nums = len(re.findall(r"\b\d+[,.]?\d*\b", norm(text)))

        if nums > 10:

            score -= 30



        # Phạt dòng có text dài (tiêu đề bảng, không phải header cột)

        if non_empty and max(len(v) for v in non_empty) > 60:

            score -= 20



        if score > best_score:

            best_score = score

            best_row = r



    return best_row







def get_cell_value_with_merge(ws, row, col):

    """

    Lấy giá trị ô, nếu ô nằm trong vùng merge thì lấy ô góc trên-trái.

    """

    v = ws.cell(row, col).value

    if v not in (None, ""):

        return v

    try:

        for rng in ws.merged_cells.ranges:

            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:

                return ws.cell(rng.min_row, rng.min_col).value

    except Exception:

        pass

    return v



def detect_header_rows_from_real_cells(ws, header_row, max_depth=4):

    """

    Chỉ lấy header từ chính file Excel, không fallback sang mẫu cố định.

    Header nhiều tầng thì lấy vài dòng liên tiếp có chữ/ô merge.

    """

    rows = [header_row]

    for r in range(header_row + 1, min(ws.max_row, header_row + max_depth) + 1):

        vals = [str(get_cell_value_with_merge(ws, r, c) or "").strip() for c in range(1, ws.max_column + 1)]

        non_empty = [v for v in vals if v]

        if not non_empty:

            continue



        numeric_like = 0

        for v in non_empty:

            vv = v.replace(",", ".").replace("/", "").replace("-", "").replace(":", "").replace("h", "").strip()

            if vv.replace(".", "").isdigit():

                numeric_like += 1



        has_text = any(re.search(r"[A-Za-zÀ-ỹ]", v) for v in non_empty)

        has_header_token = any(norm(v) in {"d1","d2","d3","d4","d5","d6","1st","2nd","3rd","4th","5th","6th"} for v in non_empty)



        if (has_text and numeric_like < max(3, len(non_empty) // 2)) or has_header_token:

            rows.append(r)

        else:

            break



    return sorted(set(rows))





def get_headers_smart(ws, header_row):

    """

    Đọc cột THẬT trong Excel, xử lý header đa tầng merge:

      - Dòng header: merged parent (VD: "Chủng loại PHC")

      - Dòng dưới:   merged sub-label (VD: "D500 - Độ dài")

      - Dòng dưới nữa: số phân biệt (VD: 5, 6, 7, 8, ...)

    → Kết quả: "Độ dài 5", "Độ dài 6"...

    Các cột thường (STT, Ngày xuất...) giữ nguyên tên.

    """

    used = find_used_range(ws)

    if used:

        min_col = used["min_col"]

        max_col = used["max_col"]

    else:

        min_col = 1

        max_col = ws.max_column



    header_rows = detect_header_rows_from_real_cells(ws, header_row)

    last_header_row = max(header_rows)



    # --- Tìm dòng số phân biệt (leaf row): ngay sau last_header_row ---

    # Đây là dòng chứa số nguyên riêng biệt: 5, 6, 7, 8, 9, 10, 11...

    numeric_leaf_row = None

    for r in range(last_header_row + 1, min(ws.max_row, last_header_row + 5) + 1):

        vals = [ws.cell(r, c).value for c in range(min_col, max_col + 1)]

        non_empty = [v for v in vals if v is not None and str(v).strip() != ""]

        if not non_empty:

            continue

        # Đa số là số nguyên/float

        num_count = sum(

            1 for v in non_empty

            if isinstance(v, (int, float)) or

               re.match(r"^\d+(\.\d+)?$", str(v).strip().replace(",", "."))

        )

        

        # Nhận diện đây có phải là dòng dữ liệu thực hay không?

        # Nếu dòng này chứa giá trị ở các cột "định danh" như STT, Ngày, Tên (chữ) thì nó là dòng data, không phải label

        is_data_row = False

        text_count = 0

        for i, v in enumerate(non_empty):

            sv = str(v).strip()

            # Dữ liệu STT, hoặc chữ

            if re.search(r"[a-zA-ZÀ-ỹ]", sv) and len(sv) > 2:

                text_count += 1

        

        # Nếu có chữ hoặc nếu là dòng ngay sát dữ liệu có cấu trúc đầy đủ, bỏ qua

        if text_count >= 2:

            is_data_row = True

            

        # Kiểm tra thêm: cột đầu tiên (thường là STT/Ngày) có giá trị liên tục không?

        # Nếu có, đích thị là data.

        first_vals = [str(vals[c] or "").strip() for c in range(min(3, len(vals)))]

        if any(first_vals):

            is_data_row = True



        if not is_data_row and num_count >= max(2, len(non_empty) * 0.6):

            numeric_leaf_row = r

            break



    # --- Kiểm tra xem cột có nằm trong vùng header merge không ---

    def is_col_in_merged_header(col_idx):

        for rng in ws.merged_cells.ranges:

            if rng.min_row <= last_header_row and rng.max_row >= header_rows[0]:

                if rng.min_col <= col_idx <= rng.max_col and rng.max_col > rng.min_col:

                    return True

        return False



    # --- Xây dựng tên cột ---

    headers = []

    for c in range(min_col, max_col + 1):

        # Đọc tất cả header rows

        parts = []

        for r in header_rows:

            v = str(get_cell_value_with_merge(ws, r, c) or "").strip()

            if v and v not in parts:

                parts.append(v)



        name_from_headers = " / ".join(parts).strip() if parts else ""



        # Nếu có numeric_leaf_row: chỉ dùng nếu cột này nằm trong merged group

        if numeric_leaf_row and is_col_in_merged_header(c):

            leaf_val = ws.cell(numeric_leaf_row, c).value

            if leaf_val is not None and str(leaf_val).strip() != "":

                leaf_str = str(leaf_val).strip()

                # Chỉ dùng leaf nếu là số (chiều dài cọc)

                if re.match(r"^\d+(\.\d+)?$", leaf_str.replace(",", ".")):

                    # Lấy tên label gần nhất (phần cuối) để làm prefix

                    if parts:

                        label = parts[-1]

                        for p in reversed(parts):

                            pn = norm(p)

                            if any(k in pn for k in ["dai", "do dai", "length", "chieu"]):

                                label = p

                                break

                        # Chỉ giữ phần sau dấu "-" hoặc "/" nếu có

                        for sep in [" - ", "/ ", "/"]:

                            if sep in label:

                                label = label.split(sep)[-1].strip()

                    else:

                        label = "Độ dài"

                    name = f"{label} {leaf_str}"

                    headers.append((c, name))

                    continue



        # Cột thường: dùng tên ghép từ header rows

        name = name_from_headers or f"Cột {get_column_letter(c)}"

        headers.append((c, name))



    # --- Giải quyết tên trùng (nếu còn) ---

    seen: dict = {}

    result = []

    for c, name in headers:

        if name not in seen:

            seen[name] = 0

        else:

            seen[name] += 1

            name = f"{name} ({seen[name]})"

        result.append((c, name))



    return result











def choose_best_sheet_profile(profiles):

    """

    Chọn sheet có khả năng là bảng nhập liệu tốt nhất:

    ưu tiên có dòng TỔNG, có cột STT, có chuỗi STT.

    """

    best = None

    best_score = -1

    for p in profiles:

        if p.get("error"):

            continue

        score = 0

        if p.get("total_row"):

            score += 30

        if p.get("stt_col"):

            score += 30

        sc = p.get("selected_chain")

        if sc:

            score += min(40, sc.get("length", 0))

        if len(p.get("headers", [])) >= 6:

            score += 10

        if score > best_score:

            best = p

            best_score = score

    return best







def is_excel_formula(v):

    return isinstance(v, str) and v.startswith("=")



def cell_addr(row, col):

    return f"{get_column_letter(col)}{row}"



def normalize_formula_to_pattern(formula, origin_cell):

    """

    Dùng để gom nhóm công thức giống nhau theo cột.

    Không cần quá phức tạp; chủ yếu giúp hiển thị công thức mẫu.

    """

    try:

        # dịch công thức về hàng 1 cùng cột để so pattern tương đối

        col_letters = re.match(r"([A-Z]+)", origin_cell).group(1)

        return Translator(formula, origin=origin_cell).translate_formula(f"{col_letters}1")

    except Exception:

        return formula



def formula_references(formula):

    """

    Lấy nhanh các tham chiếu ô/vùng trong công thức để mô tả cách làm.

    """

    if not formula:

        return []

    refs = re.findall(r"(?:'[^']+'!)?\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?", str(formula))

    # loại trùng, giữ thứ tự

    out = []

    for r in refs:

        if r not in out:

            out.append(r)

    return out



def read_formula_logic_for_sheet(ws, header_row=None, total_row=None, no_col=None):

    """

    Đọc công thức và cách làm trong sheet:

    - công thức từng dòng dữ liệu

    - công thức dòng TỔNG

    - nhóm công thức theo cột

    - merged cells

    """

    header_row = header_row or find_header_row_smart(ws)

    total_row = total_row or find_total_row(ws, header_row)

    headers = get_headers_smart(ws, header_row)

    header_by_col = {c: name for c, name in headers}

    no_col = no_col or (find_no_column_smart(ws, headers, header_row, total_row) if total_row else None)



    formulas = []

    formula_cols = {}

    max_rows_scan = ws.max_row



    for r in range(1, max_rows_scan + 1):

        for c in range(1, ws.max_column + 1):

            v = ws.cell(r, c).value

            if is_excel_formula(v):

                addr = cell_addr(r, c)

                item = {

                    "cell": addr,

                    "row": r,

                    "col": c,

                    "col_letter": get_column_letter(c),

                    "header": header_by_col.get(c, ""),

                    "formula": v,

                    "references": formula_references(v),

                    "pattern": normalize_formula_to_pattern(v, addr),

                    "role": "normal"

                }

                if total_row and r == total_row:

                    item["role"] = "total_row"

                elif total_row and header_row < r < total_row:

                    item["role"] = "data_row"

                formulas.append(item)

                formula_cols.setdefault(c, []).append(item)



    # Tìm công thức mẫu cho từng cột

    formula_columns = []

    for c, items in sorted(formula_cols.items()):

        patterns = {}

        for it in items:

            patterns[it["pattern"]] = patterns.get(it["pattern"], 0) + 1

        common_pattern = sorted(patterns.items(), key=lambda x: x[1], reverse=True)[0][0] if patterns else ""

        sample = items[0]

        formula_columns.append({

            "col": get_column_letter(c),

            "col_index": c,

            "header": header_by_col.get(c, ""),

            "count": len(items),

            "sample_cell": sample["cell"],

            "sample_formula": sample["formula"],

            "common_pattern": common_pattern,

            "roles": sorted(set(it["role"] for it in items)),

        })



    total_formulas = [it for it in formulas if it["role"] == "total_row"]

    data_formulas = [it for it in formulas if it["role"] == "data_row"]



    merged_ranges = [str(rng) for rng in ws.merged_cells.ranges]



    # Mô tả ngắn "cách làm"

    rules = []

    for it in total_formulas:

        rules.append(f"TỔNG {it['col_letter']} ({it.get('header','')}): {it['formula']}")

    # lấy tối đa 20 công thức dòng dữ liệu mẫu

    seen_cols = set()

    for it in data_formulas:

        if it["col"] in seen_cols:

            continue

        seen_cols.add(it["col"])

        rules.append(f"Dòng dữ liệu cột {it['col_letter']} ({it.get('header','')}): mẫu {it['sample_cell'] if 'sample_cell' in it else it['cell']} = {it['formula']}")

        if len(seen_cols) >= 20:

            break



    return {

        "sheet": ws.title,

        "header_row": header_row,

        "total_row": total_row,

        "stt_col": get_column_letter(no_col) if no_col else None,

        "formula_count": len(formulas),

        "formula_columns": formula_columns,

        "total_formulas": total_formulas,

        "data_formulas_sample": data_formulas[:80],

        "merged_ranges": merged_ranges[:200],

        "rules_text": rules[:80],

    }



def read_formula_logic_for_workbook(excel_path):

    wb = load_workbook(excel_path, data_only=False)

    result = {

        "file": str(excel_path),

        "sheets": []

    }

    for ws in wb.worksheets:

        try:

            header_row = find_header_row_smart(ws)

            headers = get_headers_smart(ws, header_row)

            total_row = find_total_row(ws, header_row)

            no_col = find_no_column_smart(ws, headers, header_row, total_row) if total_row else None

            logic = read_formula_logic_for_sheet(ws, header_row, total_row, no_col)

            result["sheets"].append(logic)

        except Exception as e:

            result["sheets"].append({"sheet": ws.title, "error": repr(e)})

    return result







def auto_mapping_to_excel_columns(source_cols, excel_headers):

    """

    Trả mapping dạng source_idx -> excel column number.

    Dùng khi người dùng quên bấm Auto map cột.

    """

    auto_idx = auto_map_columns(source_cols, excel_headers)

    auto_idx = ensure_no_column_in_mapping(source_cols, auto_idx, excel_headers)

    out = []

    for idx in auto_idx:

        if idx is None:

            out.append(None)

        else:

            try:

                out.append(excel_headers[idx][0])

            except Exception:

                out.append(None)

    return out







def cell_text(v):

    return str(v or "").strip()



def row_non_empty_count(ws, r):

    return sum(1 for c in range(1, ws.max_column + 1) if cell_text(ws.cell(r, c).value))



def find_used_range(ws):

    min_r, min_c = None, None

    max_r, max_c = 0, 0

    for r in range(1, ws.max_row + 1):

        for c in range(1, ws.max_column + 1):

            if cell_text(ws.cell(r, c).value):

                min_r = r if min_r is None else min(min_r, r)

                min_c = c if min_c is None else min(min_c, c)

                max_r = max(max_r, r)

                max_c = max(max_c, c)

    if min_r is None:

        return None

    return {"min_row": min_r, "min_col": min_c, "max_row": max_r, "max_col": max_c}



def infer_sheet_type(text):

    n = norm(text)

    if any(x in n for x in ["ngay xuat", "so hop dong", "so phieu", "xe van chuyen", "tong so m coc", "mui d300", "chung loai phc"]):

        return "Bảng xuất/nhập cọc - hợp đồng/phiếu/xe vận chuyển"

    if any(x in n for x in ["tong hop coc thuc te", "tai trong dung ep", "cao do", "chieu sau ep thuc te", "luc ep"]):

        return "Bảng báo cáo ép cọc / nhật ký ép cọc"

    if any(x in n for x in ["summary construction", "pile press", "pile combination", "pressing load"]):

        return "Bảng tổng hợp khối lượng ép cọc song ngữ"

    if any(x in n for x in ["nghiem thu", "bien ban", "xac nhan"]):

        return "Biên bản/nghiệm thu"

    return "Bảng Excel khác / chưa phân loại chắc"



def build_multiline_headers(ws, header_rows, min_col, max_col):

    headers = []

    for c in range(min_col, max_col + 1):

        parts = []

        for r in header_rows:

            v = cell_text(ws.cell(r, c).value)

            if v and v not in parts:

                parts.append(v)

        name = " / ".join(parts).strip()

        if not name:

            name = f"Cột {get_column_letter(c)}"

        headers.append({"col": get_column_letter(c), "index": c, "name": name})

    return headers



def detect_header_rows_general(ws, used):

    """

    Dò các dòng header của từng sheet, không cố định form.

    Lấy vùng có nhiều chữ/ô gộp trước data.

    """

    if not used:

        return []

    min_r, max_scan = used["min_row"], min(used["max_row"], used["min_row"] + 25)

    best_r = min_r

    best_score = -999



    for r in range(min_r, max_scan + 1):

        vals = [cell_text(ws.cell(r, c).value) for c in range(used["min_col"], used["max_col"] + 1)]

        joined = " ".join(vals)

        n = norm(joined)

        non_empty = sum(1 for v in vals if v)

        alpha = sum(1 for v in vals if re.search(r"[A-Za-zÀ-ỹ]", v))

        nums = sum(1 for v in vals if re.fullmatch(r"[-+]?\d+[,.]?\d*", v.replace(" ", "")))



        score = non_empty * 2 + alpha * 3 - nums * 2

        for kw in ["stt", "ngay", "ten", "loai", "vi tri", "d1", "1st", "tong", "ghi chu",

                   "so hop dong", "so phieu", "xe van chuyen", "chung loai", "tai trong", "cao do"]:

            if kw in n:

                score += 15



        if score > best_score:

            best_score = score

            best_r = r



    # Header có thể nhiều tầng: lấy best_r và 1-2 dòng tiếp theo nếu có chữ, không lấy dòng dữ liệu quá số

    header_rows = [best_r]

    for rr in range(best_r + 1, min(best_r + 4, used["max_row"] + 1)):

        vals = [cell_text(ws.cell(rr, c).value) for c in range(used["min_col"], used["max_col"] + 1)]

        joined = " ".join(vals)

        has_alpha = any(re.search(r"[A-Za-zÀ-ỹ]", v) for v in vals if v)

        numeric_heavy = sum(1 for v in vals if re.fullmatch(r"[-+]?\d+[,.]?\d*", v.replace(" ", ""))) >= max(3, len(vals)//3)

        if has_alpha and not numeric_heavy:

            header_rows.append(rr)

        elif any(norm(v) in {"d1","d2","d3","d4","d5","d6","1st","2nd","3rd","4th","5th"} for v in vals):

            header_rows.append(rr)

    return sorted(set(header_rows))



def detect_data_rows_general(ws, used, header_rows):

    if not used or not header_rows:

        return []

    start = max(header_rows) + 1

    rows = []

    for r in range(start, used["max_row"] + 1):

        vals = [cell_text(ws.cell(r, c).value) for c in range(used["min_col"], used["max_col"] + 1)]

        joined = " ".join(vals)

        if not joined.strip():

            continue

        if is_total_marker_text(joined):

            continue

        # dòng dữ liệu thường có số/ngày/mã hoặc nhiều ô

        if sum(1 for v in vals if v) >= 2:

            rows.append(r)

    return rows



def analyze_sheet_content(ws):

    used = find_used_range(ws)

    if not used:

        return {

            "sheet": ws.title,

            "empty": True,

            "summary": "Sheet trống"

        }



    all_text = []

    for r in range(used["min_row"], used["max_row"] + 1):

        vals = [cell_text(ws.cell(r, c).value) for c in range(used["min_col"], used["max_col"] + 1)]

        if any(vals):

            all_text.append(" ".join(vals))

    joined_all = "\n".join(all_text)



    header_rows = detect_header_rows_general(ws, used)

    headers = build_multiline_headers(ws, header_rows, used["min_col"], used["max_col"]) if header_rows else []



    data_rows = detect_data_rows_general(ws, used, header_rows)

    total_rows = []

    for r in range(used["min_row"], used["max_row"] + 1):

        row_text = " ".join(cell_text(ws.cell(r, c).value) for c in range(used["min_col"], used["max_col"] + 1))

        if is_total_marker_text(row_text):

            total_rows.append(r)



    formula_cells = []

    for r in range(used["min_row"], used["max_row"] + 1):

        for c in range(used["min_col"], used["max_col"] + 1):

            v = ws.cell(r, c).value

            if is_formula_value(v):

                formula_cells.append({

                    "cell": f"{get_column_letter(c)}{r}",

                    "formula": v,

                    "col": get_column_letter(c),

                })



    sample_rows = []

    for r in data_rows[:8]:

        row_data = {}

        for h in headers:

            v = cell_text(ws.cell(r, h["index"]).value)

            if v:

                row_data[h["name"]] = v

        if row_data:

            sample_rows.append({"row": r, "values": row_data})



    merged = [str(x) for x in list(ws.merged_cells.ranges)[:80]]



    return {

        "sheet": ws.title,

        "empty": False,

        "sheet_type": infer_sheet_type(joined_all),

        "used_range": {

            "from": f"{get_column_letter(used['min_col'])}{used['min_row']}",

            "to": f"{get_column_letter(used['max_col'])}{used['max_row']}",

            "min_row": used["min_row"],

            "max_row": used["max_row"],

            "min_col": used["min_col"],

            "max_col": used["max_col"],

        },

        "header_rows": header_rows,

        "headers": headers,

        "data_row_count": len(data_rows),

        "data_rows_first_last": [data_rows[0], data_rows[-1]] if data_rows else None,

        "total_rows": total_rows,

        "formula_count": len(formula_cells),

        "formula_samples": formula_cells[:50],

        "merged_ranges_count": len(ws.merged_cells.ranges),

        "merged_ranges_sample": merged,

        "sample_rows": sample_rows,

        "summary": f"{infer_sheet_type(joined_all)} | {len(headers)} cột | {len(data_rows)} dòng dữ liệu | {len(formula_cells)} ô công thức"

    }



def analyze_workbook_sheets(excel_path):

    wb = load_workbook(excel_path, data_only=False)

    return {

        "file": str(excel_path),

        "sheets": [analyze_sheet_content(ws) for ws in wb.worksheets]

    }







def short_header_name(name, max_len=45):

    """

    Rút gọn tên cột hiển thị:

    - Bỏ tiêu đề bảng dài kiểu BẢNG DIỄN GIẢI...

    - Giữ phần header thật của cột

    - Không ảnh hưởng dữ liệu/mapping gốc

    """

    s = str(name or "").replace("\n", " ").replace("\r", " ")

    parts = [re.sub(r"\s+", " ", p).strip() for p in s.split("/") if str(p).strip()]



    cleaned = []

    for p in parts:

        np = norm(p)

        # bỏ title/slogan dài, không phải tên cột

        if "bang dien giai" in np or "khoi luong thi cong" in np:

            continue

        if "summary construction" in np or "bang tong hop" in np:

            continue

        if len(p) > 55 and ("bang" in np or "cong" in np or "construction" in np):

            continue

        if p not in cleaned:

            cleaned.append(p)



    if not cleaned:

        cleaned = parts[-2:] if len(parts) >= 2 else parts



    # Nếu chỉ có đơn vị M/T/Tim thì giữ kèm cha phía trước

    if len(cleaned) >= 2 and cleaned[-1].lower() in {"m", "t", "tim"}:

        out = f"{cleaned[-2]} ({cleaned[-1]})"

    else:

        out = " / ".join(cleaned[-3:]) if cleaned else str(name or "")



    out = re.sub(r"\s+", " ", out).strip()

    if len(out) > max_len:

        out = out[:max_len-3].rstrip() + "..."

    return out







def find_last_data_row_before_total(ws, header_row, total_row, mapping_cols=None, no_col=None):

    """

    Fallback khi không có chuỗi STT liên tục:

    tìm dòng dữ liệu cuối cùng trước dòng TỔNG dựa vào các cột có dữ liệu/mapping.

    """

    end_row = (total_row - 1) if total_row else ws.max_row

    cols = []

    if mapping_cols:

        cols.extend([c for c in mapping_cols if c])

    if no_col:

        cols.append(no_col)

    cols = sorted(set(cols))

    if not cols:

        cols = list(range(1, ws.max_column + 1))



    last = None

    for r in range(header_row + 1, end_row + 1):

        if row_has_grey_background(ws, r):

            continue

        vals = [str(ws.cell(r, c).value or "").strip() for c in cols if c <= ws.max_column]

        if any(vals):

            # bỏ qua dòng header phụ nếu toàn chữ cột

            row_text = norm(" ".join(vals))

            if row_text and not is_total_marker_text(row_text):

                last = r

    return last



def find_last_stt_number_loose(ws, no_col, header_row, total_row=None):

    """

    Lấy số STT cuối cùng, không yêu cầu liên tục.

    Dùng cho các sheet STT bị đứt hoặc có công thức chưa tính.

    """

    if not no_col:

        return 0

    end_row = (total_row - 1) if total_row else ws.max_row

    last = 0

    memo = {}

    for r in range(header_row + 1, end_row + 1):

        if row_has_grey_background(ws, r):

            continue

        n = None

        try:

            n = get_stt_value(ws, r, no_col, memo)

        except Exception:

            n = None

        if isinstance(n, int):

            last = max(last, n)

            continue

        s = str(ws.cell(r, no_col).value or "").strip()

        if s.isdigit():

            last = max(last, int(s))

    return last









def row_has_big_merge_area(ws, row):

    """

    Tránh ghi vào vùng merge lớn kiểu chữ ký/trang trống.

    """

    try:

        for rng in ws.merged_cells.ranges:

            if rng.min_row <= row <= rng.max_row:

                row_span = rng.max_row - rng.min_row + 1

                col_span = rng.max_col - rng.min_col + 1

                if row_span >= 2 and col_span >= 3:

                    return True

    except Exception:

        pass

    return False





def row_is_empty_for_new_data(ws, row, mapping_cols, no_col):

    """

    Dòng được coi là trống nếu các cột cần ghi dữ liệu đều đang trống.

    Không xét cột STT vì có file đã kẻ sẵn hoặc có công thức STT.

    """

    try:

        row_text = " ".join(str(ws.cell(row, c).value or "") for c in range(1, ws.max_column + 1))

        if is_total_marker_text(row_text):

            return False

    except Exception:

        pass



    if row_has_big_merge_area(ws, row):

        return False



    cols = []

    if mapping_cols:

        cols.extend([c for c in mapping_cols if c])

    cols = sorted(set([c for c in cols if c != no_col]))



    if not cols:

        return False



    for c in cols:

        v = ws.cell(row, c).value

        if v not in (None, ""):

            return False



    return True





def find_blank_rows_before_total(ws, start_after_row, total_row, need_count, mapping_cols, no_col):

    """

    Tìm các dòng trống thật sự nằm trước dòng TỔNG để ghi dữ liệu.

    Ưu tiên giữ nguyên form, không insert/xóa dòng.

    """

    rows = []

    for r in range(start_after_row + 1, total_row):

        if row_is_empty_for_new_data(ws, r, mapping_cols, no_col):

            rows.append(r)

            if len(rows) >= need_count:

                break

    return rows

def postprocess_to_hop_coc_d1_d2(tables):

    """

    Fix riêng cho form mới:

    Nếu AI đọc header "Tổ hợp cọc" thành 1 cột nhưng dữ liệu thực tế là 2 cột con,

    ví dụ: D300 | 6 | 10 | 16 | 14,5 | +1,5 | 90

    thì đổi thành:

    D300 | D1=6 | D2=10 | Chiều dài cọc=16 | Chiều dài ép=14,5 | Ép âm dương=+1,5 | Lực ép=90



    Không ảnh hưởng các form cũ đã có sẵn D1/D2 hoặc 1st/2nd.

    """

    if not tables:

        return tables



    def _n(s):

        try:

            return norm(s)

        except Exception:

            return str(s or "").lower().strip()



    for t in tables:

        cols = list(t.get("columns", []))

        rows = t.get("rows", [])



        if not cols:

            continue



        norm_cols = [_n(c) for c in cols]



        # Nếu đã có D1/D2 hoặc 1st/2nd rồi thì bỏ qua, giữ cấu trúc cũ

        has_d1 = any(x in {"d1", "đ1", "1st"} for x in norm_cols)

        has_d2 = any(x in {"d2", "đ2", "2nd"} for x in norm_cols)

        if has_d1 and has_d2:

            continue



        # Tìm cột "Tổ hợp cọc"

        idx = None

        for i, nc in enumerate(norm_cols):

            if "to hop coc" in nc or nc == "to hop" or "pile combination" in nc:

                idx = i

                break



        if idx is None:

            continue



        # Nếu ngay sau Tổ hợp cọc là Chiều dài cọc thì khả năng cao AI bị thiếu header D2

        next_name = norm_cols[idx + 1] if idx + 1 < len(norm_cols) else ""

        should_split = False



        if "chieu dai coc" in next_name or "length of pile" in next_name:

            should_split = True



        # Hoặc nếu dữ liệu ở cột tổ hợp và cột kế tiếp đều là số ngắn, cũng tách

        if not should_split and idx + 1 < len(cols):

            sample_count = 0

            ok_count = 0

            for r in rows[:12]:

                if idx + 1 >= len(r):

                    continue

                a = str(r[idx]).strip()

                b = str(r[idx + 1]).strip()

                if a or b:

                    sample_count += 1

                    if a.replace(",", ".").replace(".", "").isdigit() and b.replace(",", ".").replace(".", "").isdigit():

                        ok_count += 1

            if sample_count and ok_count >= max(1, sample_count // 2):

                should_split = True



        if not should_split:

            continue



        # Đổi header: Tổ hợp cọc -> D1, chèn D2 ngay sau đó

        new_cols = cols[:]

        new_cols[idx] = "D1"

        new_cols.insert(idx + 1, "D2")



        # Không chèn giá trị vào rows.

        # Vì rows hiện tại đang có: 6,10,16,14.5...

        # Chỉ cần chèn header D2 là các giá trị tự dịch đúng cột.

        fixed_rows = []

        for r in rows:

            rr = list(r)



            # Nếu dòng thiếu ô so với header mới thì pad trống cuối dòng

            if len(rr) < len(new_cols):

                rr = rr + [""] * (len(new_cols) - len(rr))



            fixed_rows.append(rr)



        t["columns"] = new_cols

        t["rows"] = fixed_rows

        t["title"] = t.get("title") or "Bảng đã tách Tổ hợp cọc D1/D2"



    return tables




def merge_ocr_tables_for_continuous_read(tables):

    """

    Gộp các bảng OCR cùng cấu trúc thành một bảng liên tục để preview, mapping,

    ghi Excel và tổng hợp ngày đều dùng đủ dữ liệu đã đọc.

    """

    if not tables:

        return tables

    def _clean_col(name):

        try:

            return norm(name)

        except Exception:

            return str(name or "").strip().lower()

    def _signature(columns):

        return tuple(_clean_col(c) for c in (columns or []))

    def _is_stt_col(name):

        n = _clean_col(name)

        return n in {"stt", "no", "so thu tu", "tt", "no."} or n.startswith("stt ")

    groups = []

    passthrough = []

    for table in tables or []:

        if not isinstance(table, dict):

            passthrough.append(table)

            continue

        cols = list(table.get("columns") or [])

        rows = [list(r) if isinstance(r, (list, tuple)) else [r] for r in (table.get("rows") or [])]

        if not cols or not rows:

            passthrough.append(table)

            continue

        # Bảng key-value thường là phần thông tin phiếu, không phải dòng dữ liệu.

        if len(cols) <= 2 and {_clean_col(c) for c in cols} <= {"truong", "gia tri", "field", "value"}:

            passthrough.append(table)

            continue

        sig = _signature(cols)

        found = None

        for group in groups:

            if group["signature"] == sig:

                found = group

                break

        if found is None:

            found = {

                "signature": sig,

                "columns": cols,

                "rows": [],

                "titles": [],

            }

            groups.append(found)

        title = str(table.get("title") or "").strip()

        if title:

            found["titles"].append(title)

        width = len(found["columns"])

        for row in rows:

            rr = list(row)

            if len(rr) < width:

                rr += [""] * (width - len(rr))

            found["rows"].append(rr[:width])

    merged = []

    for group in groups:

        cols = list(group["columns"])

        rows = [list(r) for r in group["rows"]]

        stt_idx = next((i for i, name in enumerate(cols) if _is_stt_col(name)), None)

        if stt_idx is not None:

            next_no = 1

            for row in rows:

                if any(str(v or "").strip() for i, v in enumerate(row) if i != stt_idx):

                    row[stt_idx] = str(next_no)

                    next_no += 1

                else:

                    row[stt_idx] = ""

        title = "Bảng dữ liệu đã gộp"

        if len(group["titles"]) == 1:

            title = group["titles"][0]

        elif group["titles"]:

            title = f"Bảng dữ liệu đã gộp ({len(group['titles'])} phần)"

        merged.append({"title": title, "columns": cols, "rows": rows})

    merged.sort(key=lambda t: (len(t.get("columns") or []), len(t.get("rows") or [])), reverse=True)

    return merged + passthrough

