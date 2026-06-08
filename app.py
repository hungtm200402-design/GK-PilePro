# -*- coding: utf-8 -*-

import os

import sys

import math

import hashlib

import shutil

import subprocess

import time

import uuid

import socket

import threading

import tempfile

from datetime import datetime

import unicodedata

import copy, re, json, traceback, difflib, glob

from urllib import error as urllib_error, parse as urllib_parse, request as urllib_request

from pathlib import Path

def _configure_tcl_tk_runtime():
    if not getattr(sys, "frozen", False):
        return
    base = Path(getattr(sys, "_MEIPASS", Path(sys.executable).parent))
    for env_name, folders in (
        ("TCL_LIBRARY", ("_tcl_data", "tcl_data", "tcl8.6")),
        ("TK_LIBRARY", ("_tk_data", "tk_data", "tk8.6")),
    ):
        if os.environ.get(env_name):
            continue
        for folder in folders:
            candidate = base / folder
            if (candidate / "init.tcl").exists() or (candidate / "tk.tcl").exists():
                os.environ[env_name] = str(candidate)
                break


_configure_tcl_tk_runtime()

import tkinter as tk

from tkinter import filedialog, messagebox, ttk



from PIL import Image, ImageDraw, ImageFilter, ImageFont, ImageTk # type: ignore

from dotenv import load_dotenv # type: ignore

from openpyxl import load_workbook

from openpyxl.styles import PatternFill, Font, Border, Alignment, Protection

from openpyxl.utils import get_column_letter

from openpyxl.utils.cell import coordinate_to_tuple

from openpyxl.formula.translate import Translator

from gk_pilepro.gk_core import (
    DEFAULT_MODEL,
    FALLBACK_MODELS,
    app_dir,
    _safe_path_part,
    app_data_root,
    app_data_dir,
    app_data_path,
    role_log_path,
    write_role_error_log,
    clean_role_log_text_for_report,
    backup_file,
    validate_excel_before_write,
    last_run_dir,
    resource_dir,
    resource_path,
    current_user_role_labels,
    current_os_username,
    default_app_user_name,
    load_app_user_name_setting,
    current_app_user_name,
    lookup_assigned_machine_user_name,
    resolve_member_display_name,
    is_admin_build,
    approval_path,
    admin_approved_machines_path,
    code_versions_path,
    legacy_revoked_machines_path,
    get_machine_code,
    is_server_owner_machine,
    make_approval_code,
    load_revoked_machines,
    save_revoked_machines,
    machine_approval_version,
    invalidate_machine_approval_code,
    is_machine_approved,
    save_machine_approval,
    load_admin_approved_machines,
    save_admin_approved_machines,
    remember_admin_approved_machine,
    update_admin_machine_last_seen,
    parse_machine_datetime,
    format_machine_last_seen,
    is_machine_active_recently,
    delete_admin_approved_machine,
    import_local_approval_to_admin_list,
    sync_presence_machines_to_admin_list,
    env_path,
    user_settings_path,
    selected_excel_files_path,
    history_entries_path,
    mapping_templates_path,
    formula_profiles_path,
    load_history_entries,
    save_history_entries,
    load_mapping_templates,
    save_mapping_templates,
    load_formula_profiles,
    save_formula_profiles,
    append_history_entry,
    new_workflow_id,
    load_env_values,
    save_env,
    normalize_presence_server_url,
    _detect_local_ip_address,
    resolve_presence_server_url,
    DEFAULT_PRESENCE_SERVER_URL,
    presence_server_url_from_env,
    _presence_client_request_json,
    file_sha256,
    fetch_user_update_info,
    download_user_update,
    send_presence_heartbeat,
    check_presence_server_alive,
    fetch_presence_machines,
    fetch_presence_approved_machines,
    send_presence_approved_machine,
    send_presence_error_log,
    fetch_presence_error_logs,
    resolve_presence_error_log,
    delete_presence_approved_machine,
)

from gk_pilepro.gk_excel import (
    TEMPLATE_PRESETS,
    CANONICAL_TEMPLATE_COLUMNS,
    TOTAL_MARKERS,
    SYNONYM_GROUPS,
    load_selected_excel_files,
    save_selected_excel_files,
    clean_text,
    norm,
    extract_json,
    build_prompt,
    build_prompt_phieu_coc,
    call_gemini_phieu_coc,
    call_gemini,
    copy_style_row,
    copy_row_dimension,
    find_no_col_from_headers,
    ensure_no_column_in_mapping,
    find_best_source_for_target,
    normalize_table_for_template,
    preset_map_columns,
    is_total_marker_text,
    find_total_row,
    find_insert_row,
    last_number_above,
    _is_grey_fill,
    row_has_grey_background,
    row_is_mostly_blank,
    find_last_data_row_before,
    find_insert_row_in_white_area,
    is_no_header,
    is_row_total_header,
    is_summary_sum_header,
    is_actual_pressing_depth_header,
    is_segment_header,
    find_first_data_row,
    excel_col_letter,
    is_formula_value,
    translate_formula_to_row,
    capture_formula_columns,
    capture_total_sum_columns,
    apply_row_formulas_from_template,
    set_total_formulas_by_template,
    _update_sum_formula_range,
    update_total_formulas,
    group_of,
    auto_map_columns,
    find_last_stt_in_white_area,
    find_first_data_row_for_sum,
    get_stt_value,
    find_all_stt_chains,
    score_stt_column_candidate,
    merge_contiguous_stt_chains,
    select_longest_stt_chain,
    find_stt_sequence_region,
    get_row_values_nonempty,
    find_no_column_smart,
    convert_excel_value,
    normalize_vietnam_date,
    postprocess_to_hop_coc_d1_d2,
    merge_ocr_tables_for_continuous_read,
    normalize_numeric_like_text,
    _static_jacking_to_float,
    _summary_date_sort_key,
    build_static_jacking_daily_summary_lines,
    force_workbook_recalculate,
    find_header_row_smart,
    get_cell_value_with_merge,
    detect_header_rows_from_real_cells,
    get_headers_smart,
    choose_best_sheet_profile,
    is_excel_formula,
    cell_addr,
    normalize_formula_to_pattern,
    formula_references,
    read_formula_logic_for_sheet,
    read_formula_logic_for_workbook,
    auto_mapping_to_excel_columns,
    cell_text,
    row_non_empty_count,
    find_used_range,
    infer_sheet_type,
    build_multiline_headers,
    detect_header_rows_general,
    detect_data_rows_general,
    analyze_sheet_content,
    analyze_workbook_sheets,
    short_header_name,
    find_last_data_row_before_total,
    find_last_stt_number_loose,
    row_has_big_merge_area,
    row_is_empty_for_new_data,
    find_blank_rows_before_total,
)

from gk_pilepro.ui.gk_icons import (
    rounded_icon_image,
    sharp_icon_image,
    sharp_icon_image_small,
    get_windows_work_area,
    get_windows_dpi,
    build_simplified_taskbar_icon,
    build_detailed_app_icon,
    build_icon_variant,
)

from gk_pilepro.ui import gk_ui

from gk_pilepro.ui.gk_ui import (
    UI_BG,
    UI_SURFACE,
    UI_SURFACE_2,
    UI_BORDER,
    UI_TEXT,
    UI_MUTED,
    UI_PRIMARY,
    UI_PRIMARY_ACTIVE,
    UI_SUCCESS,
    UI_SUCCESS_ACTIVE,
    UI_WARN,
    UI_ERROR,
    configure_ui_metrics,
    scale_px,
    ui_font,
    RoundedButton,
    ui_button,
    RoundedMappingLabel,
    RoundedMappingDropdown,
    RoundedMappingEntry,
)

from gk_pilepro.ui.gk_editors import MappingEditor, TableEditor





APP_TITLE = "GK PilePro"

APP_LOGO_PNG = Path("assets") / "gk_logo.png"

APP_TASKBAR_PNG = Path("assets") / "gk_app_icon.png"

APP_ICON_ICO = Path("assets") / "gk_app_icon.ico"

APP_SPLASH_LOGO_PNG = Path("assets") / "gk_splash_logo.png"

APP_SPLASH_BG_PNG = Path("assets") / "gk_splash_bg.png"

APP_UI_ICON_DIR = Path("assets") / "GK_PilePro_icon_files_no_bg" / "transparent_png"

APP_DECOR_BOTTOM_RIGHT = Path("assets") / "gk_footer_decoration.png"

APP_DECOR_SIDEBAR_BOTTOM = Path("assets") / "gk_sidebar_decoration.png"

SERVER_OWNER_MACHINE_CODE = os.getenv("GK_PILEPRO_SERVER_OWNER", "").strip().upper()






class App:

    def __init__(self, root):

        self.root = root

        root.title(APP_TITLE)

        env = load_env_values()

        self.api_key_var = tk.StringVar(value=env["GEMINI_API_KEY"])

        self.model_var = tk.StringVar(value=env["GEMINI_MODEL"] or DEFAULT_MODEL)

        self.screen_profile_var = tk.StringVar(value=env.get("SCREEN_PROFILE") or "auto")

        self.presence_server_var = tk.StringVar(value=resolve_presence_server_url(env.get("PRESENCE_SERVER_URL") or DEFAULT_PRESENCE_SERVER_URL))

        self._setup_responsive_metrics()

        self.app_logo_img = None



        self.template_var = tk.StringVar(value="Bảng bất kỳ - tự nhận cột")

        self.image_path = None
        self.image_paths = []
        self.preview_image_index = 0

        self.excel_path = None

        self.tk_img = None

        self.workbook = None

        self.excel_folder = None

        self.sheet_var = tk.StringVar()

        self.header_row = None

        self.excel_headers = []

        self.tables = []

        self.selected_excel_files = load_selected_excel_files()

        self.excel_recent_listbox = None

        self.excel_recent_panel = None

        self.filters_card = None

        self.home_page = None

        self.excel_page = None

        self.history_page = None

        self.mapping_page = None

        self.settings_panel = None

        self.mapping_templates_inner = None

        self.current_page = "home"

        self.excel_recent_selected_key = None

        self.current_workflow_id = None

        self.current_workflow_date = None

        self.current_workflow_label = None

        self.current_doc_kind = None

        self.history_selected_entry = None

        self.mapping_templates = load_mapping_templates()

        self.nav_widgets = {}

        self.user_name, self.user_role = current_user_role_labels()

        if not is_admin_build():

            self.user_name = resolve_member_display_name(get_machine_code())

            self.user_role = "Thành viên"

        self.user_role_var = tk.StringVar(value=self.user_role)

        self.approval_dialog_open = False

        self.admin_approval_panel = None

        self._presence_machine_cache = []
        self._presence_cache_lock = threading.Lock()
        self._presence_ping_inflight = False
        self._presence_poll_inflight = False
        self._presence_server_check_inflight = False
        self._presence_server_online = None
        self._presence_server_down_notified = False
        self._presence_server_process = None
        self._presence_server_db_path = app_data_path("presence_state.db")
        self._presence_server_pid_path = app_data_path("presence_server.pid")
        self._presence_server_runtime_lock = threading.Lock()
        self._user_minimized = False
        self._presence_server_down_dialog_shown = False
        self._presence_server_check_interval_ms = 150
        self._update_check_inflight = False
        self._update_installing = False
        self._last_status_text = "Sẵn sàng"
        self._last_status_tone = "success"
        self._ui_images = []

        if not is_admin_build():
            try:
                self._presence_server_online = check_presence_server_alive(
                    getattr(self, "presence_server_var", tk.StringVar(value="")).get(),
                    timeout=0.2,
                )
            except Exception:
                self._presence_server_online = False

        if not is_admin_build() and self._presence_server_online is not True:
            self.member_locked = True
            try:
                self.root.withdraw()
            except Exception:
                pass
            self._presence_server_down_notified = True
            try:
                self.root.after(0, self._notify_presence_server_down)
            except Exception:
                pass
            return

        self.member_locked = (not is_admin_build()) and (not is_machine_approved())

        if self.member_locked:

            try:

                self.root.withdraw()

            except Exception:

                pass



        self.build_ui()

        self._apply_user_visibility()

        if not is_admin_build() and self._presence_server_online is not True:
            try:
                self._presence_server_down_notified = True
                self.root.after(0, self._notify_presence_server_down)
            except Exception:
                pass

        self.root.bind_all("<Control-v>", self.paste_image_from_clipboard)

        self.root.bind_all("<Control-V>", self.paste_image_from_clipboard)

        self.root.bind_all("<F5>", self.refresh_ui)

        self.root.after(300, self._check_member_approval_loop)
        self.root.after(1500, self._machine_presence_ping_loop)
        self.root.after(1800, self._presence_cache_poll_loop)
        self.root.after(0, self._presence_server_monitor_loop)
        self.root.after(900, self._user_update_check_loop)
        if is_admin_build():
            self.root.after(250, self._ensure_presence_server_on_start)
            try:
                self.root.protocol("WM_DELETE_WINDOW", self._on_admin_window_close)
            except Exception:
                pass
        try:
            self.root.bind("<Unmap>", self._track_window_state)
            self.root.bind("<Map>", self._clear_user_minimized_state)
        except Exception:
            pass



    def refresh_ui(self, event=None):

        try:

            self._set_status("Dang tai lai ung dung...", "warn")

        except Exception:

            pass

        self.root.after(80, self._restart_process)

        return "break"


    def _set_status(self, text, tone=None):

        try:

            message = str(text or "")

            if tone is None:

                lowered = message.lower()

                if any(key in lowered for key in ("lỗi", "không", "bảo trì", "đang tắt", "đã tắt", "chưa", "khong")):

                    tone = "error"

                elif any(key in lowered for key in ("đang", "dang", "chờ", "cho")):

                    tone = "warn"

                else:

                    tone = "success"

            color = {

                "error": UI_ERROR,

                "warn": UI_WARN,

                "success": UI_SUCCESS,

            }.get(tone, UI_SUCCESS)

            self._last_status_text = message
            self._last_status_tone = tone

            if hasattr(self, "status") and self.status is not None:
                self.status.config(text=message, fg=color)

            footer_status = getattr(self, "footer_status_var", None)
            if footer_status is not None:
                footer_status.set(message)

            filter_status = getattr(self, "filter_status_var", None)
            if filter_status is not None:
                filter_status.set(message)

            footer_dot = getattr(self, "footer_status_dot", None)
            if footer_dot is not None:
                try:
                    footer_dot.config(fg=color)
                except Exception:
                    pass

        except Exception:

            pass

    def _send_log_to_admin(self):
        self._set_status("Đang gửi log...", "warn")

        def _do_send():
            try:
                log_path = role_log_path()
                log_path.parent.mkdir(parents=True, exist_ok=True)
                if not log_path.exists():
                    log_path.write_text("Chưa có lỗi nào được ghi nhận.\n", encoding="utf-8")
                try:
                    log_text = log_path.read_text(encoding="utf-8", errors="replace")
                except Exception:
                    log_text = "Không đọc được nội dung log."
                if not log_text.strip():
                    log_text = "Chưa có lỗi nào được ghi nhận."
                log_text = clean_role_log_text_for_report(log_text)

                server_url = presence_server_url_from_env()
                payload = {
                    "machine_code": get_machine_code(),
                    "user_name": resolve_member_display_name(get_machine_code()),
                    "windows_user": os.environ.get("USERNAME", ""),
                    "computer_name": os.environ.get("COMPUTERNAME", socket.gethostname()),
                    "role": "Admin" if is_admin_build() else "User",
                    "app_kind": "admin" if is_admin_build() else "user",
                    "message": "User gửi log lỗi" if not is_admin_build() else "Admin gửi log lỗi",
                    "log_text": log_text[-60000:],
                    "log_path": str(log_path),
                }

                if not check_presence_server_alive(server_url, timeout=0.7):
                    def _on_server_down():
                        self._set_status("Không gửi được log: server chưa bật.", "error")
                        messagebox.showwarning(
                            "Gửi log cho Admin",
                            "Chưa gửi được log vì server Admin chưa bật.\n\nVui lòng báo Admin bật server rồi bấm gửi lại.",
                        )
                    self.root.after(0, _on_server_down)
                    return

                if not send_presence_error_log(server_url, payload, timeout=5):
                    def _on_send_fail():
                        self._set_status("Không gửi được log lên server.", "error")
                        messagebox.showerror(
                            "Gửi log cho Admin",
                            "Không gửi được log lên server.\n\nVui lòng thử lại hoặc báo Admin kiểm tra server.",
                        )
                    self.root.after(0, _on_send_fail)
                    return

                def _on_success():
                    self._set_status("Đã gửi log cho Admin.", "success")
                    messagebox.showinfo(
                        "Gửi log cho Admin",
                        "Admin đã nhận được log.\n\nThông tin đã gửi gồm mã máy, tên máy Windows, user Windows và nội dung lỗi.",
                    )
                self.root.after(0, _on_success)

            except Exception as exc:
                write_role_error_log("send_log_to_admin", exc)
                def _on_exc(e=exc):
                    self._set_status("Không chuẩn bị được log.", "error")
                    messagebox.showerror("Gửi log cho Admin", f"Không chuẩn bị được file log:\n{e}")
                self.root.after(0, _on_exc)

        threading.Thread(target=_do_send, daemon=True).start()


    def _restart_process(self):

        try:

            self.root.destroy()

        except Exception:

            pass

        try:

            os.execv(sys.executable, [sys.executable] + sys.argv)

        except Exception:

            messagebox.showerror("Khong tai lai duoc", traceback.format_exc())

            return

    def _on_admin_window_close(self):

        if is_admin_build():
            # Đóng app admin không tắt server; chỉ nút "Tắt server" mới làm việc đó.
            try:
                self._set_status("Đóng ứng dụng admin, server vẫn tiếp tục chạy.", "success")
            except Exception:
                pass

        try:

            self.root.destroy()

        except Exception:

            pass

    def _set_presence_machine_cache(self, rows):

        try:

            normalized = []

            if isinstance(rows, list):

                for item in rows:

                    if isinstance(item, dict):

                        normalized.append(dict(item))

            with self._presence_cache_lock:

                self._presence_machine_cache = normalized

        except Exception:

            pass

    def _get_presence_machine_cache(self):

        try:

            with self._presence_cache_lock:

                return [dict(item) for item in self._presence_machine_cache if isinstance(item, dict)]

        except Exception:

            return []

    def _presence_server_is_running(self):

        try:

            proc = getattr(self, "_presence_server_process", None)
            if proc is not None:
                try:
                    if proc.poll() is None:
                        return True
                except Exception:
                    pass

            pid = self._read_presence_server_pid()
            if pid is not None and self._is_pid_alive(pid):
                return True

            return False

        except Exception:

            return False

    def _read_presence_server_pid(self):

        try:

            path = Path(self._presence_server_pid_path)
            if not path.exists():
                return None
            raw = path.read_text(encoding="utf-8").strip()
            pid = int(raw)
            return pid if pid > 0 else None

        except Exception:

            return None

    def _write_presence_server_pid(self, pid):

        try:

            Path(self._presence_server_pid_path).write_text(str(int(pid)), encoding="utf-8")

        except Exception:

            pass

    def _clear_presence_server_pid(self):

        try:

            Path(self._presence_server_pid_path).unlink(missing_ok=True)

        except Exception:

            pass

    def _is_pid_alive(self, pid):

        try:

            os.kill(int(pid), 0)
            return True

        except Exception:

            return False

    def _terminate_all_presence_server_processes(self):

        try:
            subprocess.run(
                ["taskkill", "/IM", "presence_server.exe", "/T", "/F"],
                capture_output=True,
            )
        except Exception:
            pass

    def _ensure_presence_server_on_start(self):

        if not is_admin_build():

            return

        if not is_server_owner_machine():
            try:
                self._set_status("May nay khong phai may chu so huu server.", "error")
            except Exception:
                pass
            return

        try:

            if self._presence_server_is_running():

                try:

                    self._set_status("Server trạng thái máy đang chạy.", "success")

                except Exception:

                    pass

                return

            ok, msg = self._start_presence_server()

            try:

                self._set_status(msg)

            except Exception:

                pass

            if not ok:

                try:

                    messagebox.showerror("Không khởi động được server", msg)

                except Exception:

                    pass

        except Exception:

            pass

    def _presence_server_bind_from_url(self, server_url=None):
        url = normalize_presence_server_url(
            server_url
            if server_url is not None
            else getattr(self, "presence_server_var", tk.StringVar(value=DEFAULT_PRESENCE_SERVER_URL)).get()
        )
        host = "0.0.0.0"
        port = 8765
        try:
            parsed = urllib_parse.urlsplit(url)
            if parsed.port:
                port = int(parsed.port)
        except Exception:
            pass
        return host, port

    def _local_ipv4_addresses(self):
        ips = {"127.0.0.1", "localhost"}
        try:
            name = socket.gethostname()
            for item in socket.getaddrinfo(name, None, socket.AF_INET):
                ip = item[4][0]
                if ip:
                    ips.add(str(ip))
        except Exception:
            pass
        try:
            detected = _detect_local_ip_address()
            if detected:
                ips.add(str(detected))
        except Exception:
            pass
        return ips

    def _validate_presence_server_url_for_admin(self, server_url):
        url = normalize_presence_server_url(server_url)
        if not url:
            return False, "Chưa nhập URL server."
        try:
            parsed = urllib_parse.urlsplit(url)
            host = (parsed.hostname or "").strip()
            if host and host not in self._local_ipv4_addresses():
                local_ips = sorted(ip for ip in self._local_ipv4_addresses() if ip not in {"127.0.0.1", "localhost"})
                hint = ", ".join(local_ips) if local_ips else _detect_local_ip_address()
                return False, f"IP {host} không thuộc máy admin. IP máy này: {hint}."
        except Exception:
            return False, "URL server không hợp lệ."
        return True, ""

    def _ensure_presence_server_for_url(self, server_url=None):
        if not is_admin_build():
            return True, "Bản user không điều khiển server."
        if not is_server_owner_machine():
            return False, "May nay khong duoc phep khoi dong server."

        url = resolve_presence_server_url(
            server_url
            if server_url is not None
            else getattr(self, "presence_server_var", tk.StringVar(value=DEFAULT_PRESENCE_SERVER_URL)).get()
        )
        ok_url, url_msg = self._validate_presence_server_url_for_admin(url)
        if not ok_url:
            return False, url_msg
        if url and check_presence_server_alive(url, timeout=0.5):
            self._presence_server_down_notified = False
            self._presence_server_down_dialog_shown = False
            return True, "Server trạng thái máy đang chạy."

        if self._presence_server_is_running():
            self._stop_presence_server()
        ok, msg = self._start_presence_server()
        if not ok:
            return ok, msg
        for _ in range(8):
            if check_presence_server_alive(url, timeout=0.35):
                return True, "Đã bật server."
            time.sleep(0.15)
        return False, f"Server đã chạy nhưng URL {url} chưa truy cập được."

    def _start_presence_server(self):

        if not is_admin_build():

            return False, "Chỉ bản admin mới được điều khiển server."

        if not is_server_owner_machine():
            return False, "May nay khong duoc phep khoi dong server."

        with self._presence_server_runtime_lock:

            if self._presence_server_is_running():

                return True, "Server đã đang chạy."

            try:

                db_path = Path(self._presence_server_db_path).resolve()
                bind_host, bind_port = self._presence_server_bind_from_url()
                try:
                    import presence_server as presence_mod
                    presence_mod.init_db(db_path)
                except Exception:
                    pass

                if getattr(sys, "frozen", False):
                    server_exe = app_dir() / "presence_server.exe"
                    if server_exe.exists():
                        cmd = [
                            str(server_exe),
                            "--host",
                            bind_host,
                            "--port",
                            str(bind_port),
                            "--db",
                            str(db_path),
                            "--timeout",
                            "20",
                        ]
                    else:
                        cmd = [
                            sys.executable,
                            "--presence-server",
                            "--host",
                            bind_host,
                            "--port",
                            str(bind_port),
                            "--db",
                            str(db_path),
                            "--timeout",
                            "20",
                        ]
                else:
                    cmd = [
                        sys.executable,
                        str(Path(__file__).resolve()),
                        "--presence-server",
                        "--host",
                        bind_host,
                        "--port",
                        str(bind_port),
                        "--db",
                        str(db_path),
                        "--timeout",
                        "20",
                    ]

                creationflags = 0
                for flag_name in ("DETACHED_PROCESS", "CREATE_NEW_PROCESS_GROUP", "CREATE_NO_WINDOW"):
                    creationflags |= int(getattr(subprocess, flag_name, 0) or 0)

                child_env = os.environ.copy()
                # Launch the server in a clean PyInstaller environment so the
                # admin onefile process can exit without keeping its _MEI
                # temporary directory pinned by inherited runtime state.
                for key in list(child_env.keys()):
                    if key.startswith("_PYI") or key.startswith("PYINSTALLER"):
                        child_env.pop(key, None)
                child_env["PYINSTALLER_RESET_ENVIRONMENT"] = "1"

                proc = subprocess.Popen(
                    cmd,
                    cwd=str(app_dir()),
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                    stdin=subprocess.DEVNULL,
                    env=child_env,
                    creationflags=creationflags,
                )

                self._presence_server_process = proc
                self._write_presence_server_pid(proc.pid)
                self._presence_server_down_notified = False
                self._presence_server_down_dialog_shown = False
                try:
                    self._set_status(f"Server trạng thái máy đang chạy tại {bind_host}:{bind_port}.", "success")
                except Exception:
                    pass
                return True, "Đã bật server."

            except Exception as exc:

                self._presence_server_process = None
                self._clear_presence_server_pid()
                return False, f"Không khởi động được server: {exc}"

    def _stop_presence_server(self):

        if not is_admin_build():

            return False, "Chỉ bản admin mới được điều khiển server."

        with self._presence_server_runtime_lock:

            proc = getattr(self, "_presence_server_process", None)
            pid = None
            had_known_process = False
            if proc is not None:
                try:
                    if proc.poll() is None:
                        pid = proc.pid
                        had_known_process = True
                    else:
                        proc = None
                except Exception:
                    proc = None

            if proc is None:
                pid = self._read_presence_server_pid()
                had_known_process = pid is not None and self._is_pid_alive(pid)

            try:

                if proc is not None:
                    proc.terminate()
                    try:
                        proc.wait(timeout=2)
                    except Exception:
                        pass
                if pid is not None:
                    subprocess.run(["taskkill", "/PID", str(int(pid)), "/T", "/F"], capture_output=True)

            except Exception:

                pass

            self._terminate_all_presence_server_processes()
            self._presence_server_process = None
            self._clear_presence_server_pid()
            self._presence_server_down_notified = True
            self._presence_server_down_dialog_shown = True

            try:

                self._set_status("Server trạng thái máy đã tắt.", "error")

            except Exception:

                pass

            return True, "Đã tắt server." if had_known_process else "Server đang bảo trì."

    def _apply_user_visibility(self):

        if is_admin_build():

            return

        can_show = (not self.member_locked) and (self._presence_server_online is True)

        try:

            if can_show:

                if getattr(self, "_user_minimized", False):
                    return

                self._user_minimized = False
                self.root.deiconify()

                self.root.lift()

                try:
                    pass

                except Exception:

                    pass

            else:

                self.root.withdraw()
                self._user_minimized = False

        except Exception:

            pass

    def _track_window_state(self, _event=None):
        try:
            if self.root.state() == "iconic":
                self._user_minimized = True
        except Exception:
            pass

    def _clear_user_minimized_state(self, _event=None):
        self._user_minimized = False

    def _notify_presence_server_down(self):

        if is_admin_build():

            return

        if getattr(self, "_presence_server_down_dialog_shown", False):
            return

        self._presence_server_down_dialog_shown = True

        try:

            self._set_status("Server trạng thái máy đang bảo trì.", "error")

        except Exception:

            pass

        try:

            messagebox.showwarning(
                "Server đang bảo trì",
                "Server đang bảo trì.\n\nVui lòng chờ admin mở lại server để tiếp tục sử dụng ứng dụng.",
            )

        except Exception:

            pass

        try:
            self.root.after(700, self.root.destroy)
        except Exception:
            try:
                self.root.destroy()
            except Exception:
                pass

    def _exit_due_to_presence_server_down(self):

        if is_admin_build():

            return

        try:

            self._set_status("Server trạng thái máy đang bảo trì. Ứng dụng sẽ đóng.", "error")

        except Exception:

            pass

        # Kept for compatibility; actual destroy is scheduled in _notify_presence_server_down().





























    def _record_history(self, action, status="success", file_path=None, sheet=None, rows=None, message="", extra=None, workflow_id=None, workflow_label=None):

        try:

            now = datetime.now().astimezone()

            entry = {

                "timestamp": now.isoformat(timespec="seconds"),

                "date": now.strftime("%Y-%m-%d"),

                "time": now.strftime("%H:%M:%S"),

                "action": str(action or "").strip(),

                "status": str(status or "success").strip(),

                "file_path": str(file_path or "").strip(),

                "file_name": Path(file_path).name if file_path else "",

                "sheet": str(sheet or "").strip(),

                "rows": rows,

                "message": str(message or "").strip(),

                "workflow_id": str(workflow_id or self.current_workflow_id or "").strip(),

                "workflow_label": str(workflow_label or self.current_workflow_label or "").strip(),

            }

            if isinstance(extra, dict) and extra:

                entry["extra"] = extra

            append_history_entry(entry)

            self._sync_history_view()

        except Exception:

            pass

























































    def show_page(self, page_name):

        page_name = page_name if page_name in {"home", "excel", "history", "mapping"} else "home"

        self.current_page = page_name

        try:

            if self.home_page is not None:

                if page_name == "home":

                    self.home_page.pack(fill="both", expand=True)

                else:

                    self.home_page.pack_forget()

            if self.excel_page is not None:

                if page_name == "excel":

                    self.excel_page.pack(fill="both", expand=True)

                    self._sync_excel_recent_sidebar()

                else:

                    self.excel_page.pack_forget()

            if self.history_page is not None:

                if page_name == "history":

                    self.history_page.pack(fill="both", expand=True)

                    self._sync_history_view()

                else:

                    self.history_page.pack_forget()

            if self.mapping_page is not None:

                if page_name == "mapping":

                    self.mapping_page.pack(fill="both", expand=True)

                    self._render_mapping_templates()

                else:

                    self.mapping_page.pack_forget()

        except Exception:

            pass

        self._refresh_nav_state()

        try:
            self._scroll_main_content_to_top()
        except Exception:
            pass

        return "break"



    def show_home_page(self, event=None):

        return self.show_page("home")









    def _scroll_main_content_to_top(self):
        canvas = getattr(self, "home_body_canvas", None) or getattr(self, "content_canvas", None)
        if canvas is not None:
            canvas.yview_moveto(0)


    def _on_main_content_mousewheel(self, event):
        canvas = getattr(self, "home_body_canvas", None) or getattr(self, "content_canvas", None)
        if canvas is None:
            return
        try:
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        except Exception:
            pass


    def _bind_main_content_mousewheel(self, event=None):
        try:
            canvas = getattr(self, "home_body_canvas", None) or getattr(self, "content_canvas", None)
            if canvas is not None:
                canvas.bind_all("<MouseWheel>", self._on_main_content_mousewheel)
        except Exception:
            pass


    def _unbind_main_content_mousewheel(self, event=None):
        try:
            canvas = getattr(self, "home_body_canvas", None) or getattr(self, "content_canvas", None)
            if canvas is not None:
                canvas.unbind_all("<MouseWheel>")
        except Exception:
            pass












    def _recent_widget_contains(self, widget):
        targets = {
            getattr(self, "excel_recent_canvas", None),
            getattr(self, "excel_recent_inner", None),
            getattr(self, "excel_recent_panel", None),
        }
        while widget is not None:
            if widget in targets:
                return True
            widget = getattr(widget, "master", None)
        return False


    def _on_recent_mousewheel(self, event):
        canvas = getattr(self, "excel_recent_canvas", None)
        if canvas is None:
            return
        if not self._recent_widget_contains(getattr(event, "widget", None)):
            return
        try:
            delta = getattr(event, "delta", 0) or 0
            if delta:
                canvas.yview_scroll(int(-1 * (delta / 120)), "units")
                return
            num = getattr(event, "num", None)
            if num == 4:
                canvas.yview_scroll(-1, "units")
            elif num == 5:
                canvas.yview_scroll(1, "units")
        except Exception:
            pass


    def _bind_recent_mousewheel_recursive(self, widget=None):
        widget = widget or getattr(self, "excel_recent_inner", None)
        if widget is None:
            return
        try:
            widget.bind("<MouseWheel>", self._on_recent_mousewheel, add="+")
            widget.bind("<Button-4>", self._on_recent_mousewheel, add="+")
            widget.bind("<Button-5>", self._on_recent_mousewheel, add="+")
        except Exception:
            pass
        try:
            for child in widget.winfo_children():
                self._bind_recent_mousewheel_recursive(child)
        except Exception:
            pass




















    def _refresh_nav_state(self):

        active_page = getattr(self, "current_page", "home")

        for page_name, widgets in getattr(self, "nav_widgets", {}).items():

            row = widgets.get("row")

            inner = widgets.get("inner")

            accent = widgets.get("accent")

            icon = widgets.get("icon")

            label = widgets.get("label")

            if row is None or icon is None or label is None:

                continue

            is_active = page_name == active_page

            bg = "#0f8d6d" if is_active else "#053f32"

            fg = "#ffffff" if is_active else "#d8f3e8"

            try:

                row.config(bg=bg)

                row.config(highlightbackground="#0f8d6d" if is_active else "#0b5a45")

                if inner is not None:

                    inner.config(bg=bg)

                if accent is not None:

                    accent.config(bg="#f4c542" if is_active else bg)

                icon.config(bg=bg, fg=fg)

                label.config(bg=bg, fg=fg, font=ui_font(11, bold=is_active))

            except Exception:

                pass


    def _activate_nav_item(self, page_name, callback=None):
        if page_name in {"settings", "help", "about"}:
            self._dialog_return_page = getattr(self, "current_page", "home")
        else:
            self._dialog_return_page = None
        self.current_page = page_name
        self._refresh_nav_state()
        if callback is not None:
            return callback()
        return None












    def _check_member_approval_loop(self):

        if is_admin_build():

            return

        try:

            approved = is_machine_approved()

            if not approved:

                self.member_locked = True

                if self._presence_server_online is not True:
                    self._notify_presence_server_down()
                    return

                try:

                    self.root.withdraw()

                except Exception:

                    pass

                if not self.approval_dialog_open:

                    self.show_member_approval_dialog()

            elif self.member_locked:

                self.member_locked = False

                try:

                    if not is_admin_build():

                        self.user_role = resolve_member_display_name(get_machine_code())

                        self.user_role_var.set(self.user_role)

                        try:

                            self.root.update_idletasks()

                        except Exception:

                            pass

                    self._set_status("Máy đã được duyệt.", "success")

                except Exception:

                    pass

                self._apply_user_visibility()

        finally:

            try:

                self.root.after(3000, self._check_member_approval_loop)

            except Exception:

                pass


    def _machine_presence_ping_loop(self):

        try:

            if not self.member_locked and self._presence_server_online is not False:
                server_url = resolve_presence_server_url(getattr(self, "presence_server_var", tk.StringVar(value="")).get())
                if server_url and not self._presence_ping_inflight:
                    self._presence_ping_inflight = True

                    def worker():

                        try:

                            payload = {
                                "machine_code": get_machine_code(),
                                "user_name": getattr(self, "user_name", "") or current_os_username(),
                                "role": getattr(self, "user_role", ""),
                                "app_kind": "admin" if is_admin_build() else "user",
                                "status": "online",
                            }
                            if send_presence_heartbeat(server_url, payload):
                                update_admin_machine_last_seen(get_machine_code())

                        finally:

                            self._presence_ping_inflight = False

                    threading.Thread(target=worker, daemon=True).start()

        except Exception:

            pass

        try:

            self.root.after(5000, self._machine_presence_ping_loop)

        except Exception:

            pass

    def _presence_server_monitor_loop(self):

        try:

            if not is_admin_build():

                server_url = resolve_presence_server_url(getattr(self, "presence_server_var", tk.StringVar(value="")).get())

                if server_url and not self._presence_server_check_inflight:

                    self._presence_server_check_inflight = True

                    def worker():

                        online = False

                        try:

                            online = check_presence_server_alive(server_url, timeout=0.2)

                        finally:

                            previous = self._presence_server_online

                            self._presence_server_online = online

                            self._presence_server_check_inflight = False

                            if previous is not online:

                                if not online:

                                    if not self._presence_server_down_notified:

                                        self._presence_server_down_notified = True

                                        try:
                                            self.root.after(0, self._notify_presence_server_down)
                                        except Exception:
                                            pass

                                else:

                                    self._presence_server_down_notified = False
                                    self._presence_server_down_dialog_shown = False
                                    try:
                                        self.root.after(0, lambda: self._set_status("Máy đã kết nối server trạng thái.", "success"))
                                    except Exception:
                                        pass

                            self._apply_user_visibility()

                            if not online and not self._presence_server_down_dialog_shown:

                                self._presence_server_down_notified = True

                                try:
                                    self.root.after(0, self._notify_presence_server_down)
                                except Exception:
                                    pass

                    threading.Thread(target=worker, daemon=True).start()

        except Exception:

            pass

        try:

            self.root.after(self._presence_server_check_interval_ms, self._presence_server_monitor_loop)

        except Exception:

            pass

    def _show_update_progress_dialog(self):
        """Show a floating progress window in the bottom-right corner when downloading an update."""
        try:
            if getattr(self, "_update_progress_dialog", None):
                try:
                    self._update_progress_dialog.destroy()
                except Exception:
                    pass
            dlg = tk.Toplevel(self.root)
            dlg.title("")
            dlg.overrideredirect(True)
            dlg.attributes("-topmost", True)
            dlg.configure(bg=UI_BORDER)
            w, h = scale_px(330), scale_px(90)
            sw = self.root.winfo_screenwidth()
            sh = self.root.winfo_screenheight()
            x = sw - w - scale_px(20)
            y = sh - h - scale_px(56)
            dlg.geometry(f"{w}x{h}+{x}+{y}")
            inner = tk.Frame(dlg, bg=UI_SURFACE, padx=scale_px(14), pady=scale_px(10))
            inner.pack(fill="both", expand=True, padx=1, pady=1)
            tk.Label(
                inner, text="\u1f504  \u0110ang t\u1ea3i b\u1ea3n c\u1eadp nh\u1eadt...",
                font=ui_font(10, bold=True), bg=UI_SURFACE, fg=UI_TEXT, anchor="w"
            ).pack(fill="x")
            pb = ttk.Progressbar(inner, mode="indeterminate", length=scale_px(292))
            pb.pack(fill="x", pady=(scale_px(6), 0))
            pb.start(12)
            info_lbl = tk.Label(
                inner, text="\u0110ang k\u1ebft n\u1ed1i t\u1edbi m\u00e1y ch\u1ee7...",
                font=ui_font(9), bg=UI_SURFACE, fg=UI_MUTED, anchor="w"
            )
            info_lbl.pack(fill="x", pady=(scale_px(3), 0))
            dlg._pb = pb
            dlg._info_lbl = info_lbl
            dlg._pb_determinate = False
            self._update_progress_dialog = dlg
        except Exception:
            self._update_progress_dialog = None

    def _update_update_progress(self, downloaded, total):
        """Update the progress bar. Must be called on the main (tkinter) thread."""
        try:
            dlg = getattr(self, "_update_progress_dialog", None)
            mb_done = downloaded / 1024 / 1024
            if total > 0:
                pct = min(100, int(downloaded * 100 / total))
                mb_total = total / 1024 / 1024
                if dlg:
                    pb = dlg._pb
                    info_lbl = dlg._info_lbl
                    if not dlg._pb_determinate:
                        pb.stop()
                        pb.config(mode="determinate", maximum=100)
                        dlg._pb_determinate = True
                    pb.config(value=pct)
                    info_lbl.config(text=f"{mb_done:.1f} / {mb_total:.1f} MB  \u2014  {pct}%")
            else:
                if dlg:
                    dlg._info_lbl.config(text=f"\u0110\u00e3 t\u1ea3i {mb_done:.1f} MB...")
        except Exception:
            pass

    def _close_update_progress_dialog(self):
        """Destroy the floating update progress window."""
        try:
            dlg = getattr(self, "_update_progress_dialog", None)
            if dlg:
                dlg.destroy()
        except Exception:
            pass
        self._update_progress_dialog = None

    def _user_update_check_loop(self):

        try:

            if (
                not is_admin_build()
                and getattr(sys, "frozen", False)
                and not self._update_installing
                and not self._update_check_inflight
                and self._presence_server_online is True
            ):

                server_url = resolve_presence_server_url(getattr(self, "presence_server_var", tk.StringVar(value="")).get())

                if server_url:

                    self._update_check_inflight = True

                    def worker():

                        update_info = None
                        update_path = None

                        try:

                            update_info = fetch_user_update_info(server_url, timeout=3)

                            if update_info:

                                remote_sha = str(update_info.get("sha256") or "").strip().lower()
                                local_sha = file_sha256(sys.executable).lower()

                                if remote_sha and remote_sha != local_sha:

                                    # Show download progress dialog
                                    try:
                                        self.root.after(0, self._show_update_progress_dialog)
                                    except Exception:
                                        pass

                                    def _on_progress(downloaded, total):
                                        try:
                                            self.root.after(0, lambda d=downloaded, t=total: self._update_update_progress(d, t))
                                        except Exception:
                                            pass

                                    update_path = download_user_update(server_url, update_info, progress_callback=_on_progress)

                        except Exception:

                            update_path = None

                        finally:

                            self._update_check_inflight = False

                            if update_path:

                                try:
                                    self.root.after(0, lambda p=update_path: self._install_user_update(p))
                                except Exception:
                                    pass

                            else:

                                # Close progress dialog if download failed or was not needed
                                try:
                                    self.root.after(0, self._close_update_progress_dialog)
                                except Exception:
                                    pass

                    threading.Thread(target=worker, daemon=True).start()

        except Exception:

            pass

        try:

            self.root.after(60000, self._user_update_check_loop)

        except Exception:

            pass

    def _install_user_update(self, update_path):

        if is_admin_build() or self._update_installing:

            # Close any lingering progress dialog
            self._close_update_progress_dialog()

            return

        self._update_installing = True

        # Close the download progress dialog before showing the install message
        self._close_update_progress_dialog()

        try:

            messagebox.showinfo(
                "Cập nhật ứng dụng",
                "Đã có bản cập nhật mới.\n\nỨng dụng sẽ tự cập nhật và mở lại ngay.",
            )

        except Exception:

            pass

        try:

            current_exe = Path(sys.executable).resolve()
            update_path = Path(update_path).resolve()
            bat_path = update_path.parent / "install_update.bat"
            pid = os.getpid()
            bat = f"""@echo off
set "SRC={update_path}"
set "DST={current_exe}"
set "PID={pid}"
:wait
tasklist /FI "PID eq %PID%" | find "%PID%" >nul
if not errorlevel 1 (
  timeout /t 1 /nobreak >nul
  goto wait
)
copy /Y "%SRC%" "%DST%" >nul
start "" "%DST%"
del "%SRC%" >nul 2>nul
del "%~f0" >nul 2>nul
"""
            bat_path.write_text(bat, encoding="utf-8")
            creationflags = 0
            for flag_name in ("CREATE_NEW_PROCESS_GROUP", "CREATE_NO_WINDOW"):
                creationflags |= int(getattr(subprocess, flag_name, 0) or 0)
            subprocess.Popen(
                ["cmd", "/c", str(bat_path)],
                cwd=str(current_exe.parent),
                stdin=subprocess.DEVNULL,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                creationflags=creationflags,
            )

        except Exception as exc:

            self._update_installing = False

            try:

                messagebox.showerror("Không cập nhật được", f"Không thể tự cập nhật ứng dụng:\n{exc}")

            except Exception:

                pass

            return

        try:

            self.root.destroy()

        except Exception:

            os._exit(0)

    def _presence_cache_poll_loop(self):

        try:

            server_url = resolve_presence_server_url(getattr(self, "presence_server_var", tk.StringVar(value="")).get())

            if server_url and not self._presence_poll_inflight:

                self._presence_poll_inflight = True

                def worker():

                    rows = []

                    try:

                        rows = fetch_presence_machines(server_url, timeout=2)

                    finally:

                        self._set_presence_machine_cache(rows)

                        self._presence_poll_inflight = False

                threading.Thread(target=worker, daemon=True).start()

        except Exception:

            pass

        try:

            self.root.after(5000, self._presence_cache_poll_loop)

        except Exception:

            pass



    def _center_dialog_on_screen(self, win):

        try:

            win.update_idletasks()

            width = win.winfo_width()

            height = win.winfo_height()

            screen_w = win.winfo_screenwidth()

            screen_h = win.winfo_screenheight()

            x = max(0, (screen_w - width) // 2)

            y = max(0, (screen_h - height) // 2)

            win.geometry(f"+{x}+{y}")

        except Exception:

            pass

    def _fit_dialog_to_screen(self, win, preferred_w, preferred_h, min_w=720, min_h=520, max_ratio=0.8, lock_size=False):
        try:
            screen_w = int(getattr(self, "screen_w", 0) or win.winfo_screenwidth() or 1366)
            screen_h = int(getattr(self, "screen_h", 0) or win.winfo_screenheight() or 768)
            max_w = max(640, int(screen_w * max_ratio))
            max_h = max(480, int(screen_h * max_ratio))
            width = max(min_w, min(int(preferred_w), max_w))
            height = max(min_h, min(int(preferred_h), max_h))
            win.geometry(f"{width}x{height}")
            win.minsize(min(width, max_w), min(height, max_h))
            if lock_size:
                win.maxsize(width, height)
        except Exception:
            try:
                win.geometry(f"{preferred_w}x{preferred_h}")
            except Exception:
                pass



    def show_member_approval_dialog(self):

        if self.approval_dialog_open:

            return

        self.approval_dialog_open = True

        machine_code = get_machine_code()

        win = tk.Toplevel(self.root)

        win.title("Yêu cầu duyệt sử dụng")

        win.configure(bg=UI_SURFACE)

        win.resizable(False, False)

        try:

            if self.root.state() != "withdrawn":

                win.transient(self.root)

        except Exception:

            pass

        win.grab_set()

        try:

            win.protocol("WM_DELETE_WINDOW", self.root.destroy)

        except Exception:

            pass

        try:

            win.bind("<Destroy>", lambda _e: setattr(self, "approval_dialog_open", False))

        except Exception:

            pass



        body = tk.Frame(win, bg=UI_SURFACE, padx=22, pady=18)
        body.pack(fill="both", expand=True)
        body.grid_columnconfigure(0, weight=1)

        read_box = tk.Frame(body, bg="#fbfdff", highlightthickness=1, highlightbackground=UI_BORDER, padx=18, pady=16)
        read_box.grid(row=0, column=0, sticky="ew", pady=(0, 16))

        tk.Label(read_box, text="Phần đọc mã", bg="#fbfdff", fg=UI_TEXT, font=ui_font(11, bold=True)).pack(anchor="w")
        tk.Label(read_box, text="Gửi mã máy bên dưới cho Admin để nhận mã duyệt.", bg="#fbfdff", fg=UI_MUTED, font=ui_font(10)).pack(anchor="w", pady=(2, 10))

        tk.Label(read_box, text="Mã máy", bg="#fbfdff", fg=UI_TEXT, font=ui_font(11, bold=True)).pack(anchor="w")

        machine_var = tk.StringVar(value=machine_code)

        machine_entry = tk.Entry(read_box, textvariable=machine_var, width=46, relief="solid", bd=1, font=ui_font(11))

        machine_entry.pack(fill="x", pady=(4, 10))

        machine_entry.configure(state="readonly")



        tk.Label(read_box, text="Mã duyệt", bg="#fbfdff", fg=UI_TEXT, font=ui_font(11, bold=True)).pack(anchor="w")

        code_var = tk.StringVar()

        code_entry = tk.Entry(read_box, textvariable=code_var, width=46, relief="solid", bd=1, font=ui_font(11))

        code_entry.pack(fill="x", pady=(4, 0))

        code_entry.focus_set()

        member_box = tk.Frame(body, bg=UI_SURFACE, highlightthickness=1, highlightbackground=UI_BORDER, padx=18, pady=16)
        member_box.grid(row=1, column=0, sticky="ew", pady=(0, 16))

        tk.Label(member_box, text="Khối thành viên", bg=UI_SURFACE, fg=UI_TEXT, font=ui_font(11, bold=True)).pack(anchor="w")
        tk.Label(member_box, textvariable=self.user_role_var, bg=UI_SURFACE, fg=UI_MUTED, font=ui_font(10)).pack(anchor="w", pady=(3, 0))



        actions = tk.Frame(body, bg=UI_SURFACE)
        actions.grid(row=2, column=0, sticky="ew")



        def copy_machine():

            self.root.clipboard_clear()

            self.root.clipboard_append(machine_code)

            try:

                self._set_status("Đã copy mã máy.", "success")

            except Exception:

                pass



        def approve():

            if self._presence_server_online is not True:
                try:
                    self._set_status("Server trạng thái máy đang bảo trì.", "error")
                except Exception:
                    pass
                messagebox.showerror(
                    "Server đang bảo trì",
                    "Không thể duyệt máy khi server đang bảo trì.",
                )
                return

            if save_machine_approval(code_var.get()):

                if not is_admin_build():

                    self.user_role = resolve_member_display_name(machine_code)

                    self.user_role_var.set(self.user_role)

                win.destroy()

                self.member_locked = False

                try:

                    self.root.deiconify()

                    self.root.lift()

                    self._set_status("Máy đã được duyệt.", "success")

                    self.root.update_idletasks()

                except Exception:

                    pass

            else:

                messagebox.showerror("Mã không đúng", "Mã duyệt không hợp lệ. Kiểm tra lại mã Admin gửi.")



        ui_button(actions, "Copy mã máy", copy_machine, width=12, variant="soft").pack(side="left", padx=(0, 8))

        ui_button(actions, "Xác nhận", approve, width=11, variant="success").pack(side="left")

        ui_button(actions, "Thoát", self.root.destroy, width=9).pack(side="right")



        self._fit_dialog_to_screen(win, 680, 460, min_w=680, min_h=460, max_ratio=0.70, lock_size=True)
        self._center_dialog_on_screen(win)

        try:

            win.lift()

            win.focus_force()

            code_entry.focus_force()

            win.attributes("-topmost", True)

            win.after(700, lambda: win.attributes("-topmost", False))

        except Exception:

            pass

        self.root.wait_window(win)




















    def _setup_responsive_metrics(self):
        profile_key = "auto"
        try:
            profile_var = getattr(self, "screen_profile_var", None)
            if profile_var is not None:
                profile_key = str(profile_var.get()).strip().lower() or "auto"
        except Exception:
            profile_key = "auto"

        work_area = get_windows_work_area() or {}
        try:
            sw = int(work_area.get("width") or self.root.winfo_screenwidth() or 1500)
            sh = int(work_area.get("height") or self.root.winfo_screenheight() or 900)
        except Exception:
            sw, sh = 1500, 900

        actual_hwnd = None
        try:
            self.root.update_idletasks()
            actual_hwnd = self.root.winfo_id()
        except Exception:
            actual_hwnd = None

        actual_dpi = get_windows_dpi(actual_hwnd)
        simulated_dpi = None
        forced_size = None
        custom_match = re.match(r"^custom:(\d+)x(\d+)(?:@(\d+))?$", profile_key)
        if profile_key != "auto":
            profile_sizes = {
                "laptop_156": (1366, 768, 240),
                "laptop_16": (1600, 900, 240),
                "display_1920x1080": (1920, 1080, 280),
                "display_1600x900": (1600, 900, 240),
                "display_1366x768": (1366, 768, 240),
                "display_1280x720": (1280, 720, 240),
                "display_960x540": (960, 540, 160),
                "display_1080x1920": (1080, 1920, 280),
                "display_900x1600": (900, 1600, 240),
                "display_768x1366": (768, 1366, 240),
                "display_720x1280": (720, 1280, 240),
                "display_540x960": (540, 960, 160),
                "display_2560x1080": (2560, 1080, 240),
                "display_3440x1440": (3440, 1440, 280),
                "display_1920x800": (1920, 800, 220),
            }
            forced = profile_sizes.get(profile_key)
            if forced:
                forced_size = (forced[0], forced[1])
                simulated_dpi = int(forced[2])
            elif custom_match:
                forced_size = (int(custom_match.group(1)), int(custom_match.group(2)))
                simulated_dpi = int(custom_match.group(3) or 240)

        if forced_size:
            sw, sh = forced_size
        dpi_for_scale = simulated_dpi if simulated_dpi is not None else actual_dpi

        self.screen_w = sw
        self.screen_h = sh
        self.screen_dpi = dpi_for_scale
        self.screen_profile_mode = "test" if profile_key != "auto" else "auto"

        self.compact_ui = sw < 1700 or sh < 950
        self.tiny_ui = sw <= 1366 or sh <= 820
        self.short_ui = sh <= 820
        self.dense_ui = sw <= 1600 or sh <= 900
        self.micro_ui = sw <= 1366 or sh <= 768
        self.main_content_scroll = bool(sw <= 1600 or sh <= 900)

        try:
            if self.screen_profile_mode == "auto":
                if sw <= 1366 or sh <= 768:
                    gk_ui.UI_SCALE = max(1.00, min(1.12, round(actual_dpi / 120.0, 2)))
                elif sw <= 1600 or sh <= 900:
                    gk_ui.UI_SCALE = max(1.03, min(1.16, round(actual_dpi / 108.0, 2)))
                else:
                    gk_ui.UI_SCALE = max(1.0, min(1.20, round(actual_dpi / 96.0, 2)))
            else:
                gk_ui.UI_SCALE = max(1.00, min(1.20, round(dpi_for_scale / 240.0, 2)))
            gk_ui.UI_READABILITY_MODE = True
            gk_ui.UI_FONT_BONUS = 1
            gk_ui.UI_FONT_MIN_SIZE = 12
        except Exception:
            gk_ui.UI_SCALE = 1.0
            gk_ui.UI_FONT_BONUS = 1
            gk_ui.UI_FONT_MIN_SIZE = 12
            gk_ui.UI_READABILITY_MODE = True

        try:
            self.root.tk.call("tk", "scaling", gk_ui.UI_SCALE)
        except Exception:
            pass

        if self.micro_ui:
            min_win_w = min(980, max(860, sw - 110))
            min_win_h = min(620, max(560, sh - 110))
        elif self.dense_ui:
            min_win_w = min(1040, max(880, sw - 90))
            min_win_h = min(650, max(580, sh - 100))
        else:
            min_win_w = min(1080, max(900, sw - 80))
            min_win_h = min(700, max(600, sh - 80))

        if self.screen_profile_mode == "auto":
            win_w = sw
            win_h = sh
        else:
            win_w = min(sw, max(min_win_w, int(sw * 0.98)))
            win_h = min(sh, max(min_win_h, int(sh * 0.985)))
        try:
            self.root.geometry(f"{win_w}x{win_h}+0+0")
            self.root.minsize(min_win_w, min_win_h)
            if self.screen_profile_mode == "auto":
                try:
                    self.root.state("zoomed")
                except Exception:
                    pass
        except Exception:
            pass

        if self.micro_ui:
            self.sidebar_w = scale_px(190)
            self.main_padx = scale_px(7)
            self.main_pady = scale_px(6)
            self.card_padx = scale_px(8)
            self.card_pady = scale_px(6)
            self.workspace_mins = (scale_px(250), scale_px(470), scale_px(320))
            self.image_drop_h = scale_px(118)
            self.mapping_canvas_h = scale_px(150)
            self.logo_max = (scale_px(155), scale_px(180))
        elif self.tiny_ui:
            self.sidebar_w = scale_px(200)
            self.main_padx = scale_px(8)
            self.main_pady = scale_px(8)
            self.card_padx = scale_px(8)
            self.card_pady = scale_px(7)
            self.workspace_mins = (scale_px(260), scale_px(500), scale_px(330))
            self.image_drop_h = scale_px(126)
            self.mapping_canvas_h = scale_px(162)
            self.logo_max = (scale_px(165), scale_px(190))
        elif self.dense_ui:
            self.sidebar_w = scale_px(200)
            self.main_padx = scale_px(11)
            self.main_pady = scale_px(9)
            self.card_padx = scale_px(10)
            self.card_pady = scale_px(8)
            self.workspace_mins = (scale_px(270), scale_px(540), scale_px(350))
            self.image_drop_h = scale_px(138)
            self.mapping_canvas_h = scale_px(176)
            self.logo_max = (scale_px(170), scale_px(200))
        elif self.compact_ui:
            self.sidebar_w = scale_px(205)
            self.main_padx = scale_px(14)
            self.main_pady = scale_px(12)
            self.card_padx = scale_px(11)
            self.card_pady = scale_px(9)
            self.workspace_mins = (scale_px(280), scale_px(580), scale_px(370))
            self.image_drop_h = scale_px(150)
            self.mapping_canvas_h = scale_px(190)
            self.logo_max = (scale_px(175), scale_px(205))
        else:
            self.sidebar_w = scale_px(210)
            self.main_padx = scale_px(22)
            self.main_pady = scale_px(18)
            self.card_padx = scale_px(14)
            self.card_pady = scale_px(12)
            self.workspace_mins = (scale_px(300), scale_px(680), scale_px(400))
            self.image_drop_h = scale_px(158)
            self.mapping_canvas_h = scale_px(198)
            self.logo_max = (scale_px(185), scale_px(215))



    def setup_window_icon(self):

        icon_file = resource_path(*APP_ICON_ICO.parts)

        self._apply_window_icon(icon_file)

        try:

            self.root.after(250, lambda: self._apply_window_icon(icon_file))

        except Exception:

            pass



    def _apply_window_icon(self, icon_file):

        try:

            if icon_file.exists():

                self.root.iconbitmap(default=str(icon_file))

                self.root.wm_iconbitmap(default=str(icon_file))

        except Exception:

            pass

        try:

            logo_path = resource_path(*APP_LOGO_PNG.parts)
            detailed_base = build_detailed_app_icon(logo_path, 256) if logo_path.exists() else build_simplified_taskbar_icon(256)

            icon_imgs = [

                ImageTk.PhotoImage(build_icon_variant(logo_path, 16) if logo_path.exists() else sharp_icon_image_small(detailed_base, (16, 16))),

                ImageTk.PhotoImage(build_icon_variant(logo_path, 32) if logo_path.exists() else sharp_icon_image_small(detailed_base, (32, 32))),

                ImageTk.PhotoImage(build_icon_variant(logo_path, 48) if logo_path.exists() else sharp_icon_image_small(detailed_base, (48, 48))),

                ImageTk.PhotoImage(build_icon_variant(logo_path, 64) if logo_path.exists() else sharp_icon_image_small(detailed_base, (64, 64))),

                ImageTk.PhotoImage(build_icon_variant(logo_path, 128) if logo_path.exists() else sharp_icon_image_small(detailed_base, (128, 128))),

                ImageTk.PhotoImage(build_icon_variant(logo_path, 256) if logo_path.exists() else sharp_icon_image(detailed_base, (256, 256))),

            ]

            self.root.iconphoto(True, *icon_imgs)

            self.window_icon_imgs = icon_imgs

        except Exception:

            pass



    def setup_theme(self):

        self.root.configure(bg=UI_BG)

        try:

            row_height = scale_px(30) if getattr(self, "micro_ui", False) else (scale_px(31) if getattr(self, "dense_ui", False) else scale_px(32))

            heading_pad = (scale_px(6), scale_px(5)) if getattr(self, "micro_ui", False) else ((scale_px(7), scale_px(6)) if getattr(self, "dense_ui", False) else (scale_px(8), scale_px(7)))

            style = ttk.Style(self.root)

            style.theme_use("clam")

            try:
                self.root.option_add("*Font", ui_font(11))
                self.root.option_add("*Entry.Font", ui_font(11))
                self.root.option_add("*Text.Font", ui_font(11))
                self.root.option_add("*Menu.Font", ui_font(11))
            except Exception:
                pass

            style.configure(".", font=ui_font(11), background=UI_BG, foreground=UI_TEXT)

            style.configure(

                "TCombobox",

                fieldbackground=UI_SURFACE,

                background=UI_SURFACE,

                foreground=UI_TEXT,

                font=ui_font(11),

                bordercolor=UI_BORDER,

                lightcolor=UI_BORDER,

                darkcolor=UI_BORDER,

                arrowcolor=UI_MUTED,

                relief="flat",

                padding=(scale_px(10), scale_px(7)),

                arrowsize=14,

            )

            style.map(

                "TCombobox",

                fieldbackground=[("readonly", UI_SURFACE), ("focus", "#fbfdff")],

                bordercolor=[("focus", UI_PRIMARY), ("!focus", UI_BORDER)],

                arrowcolor=[("active", UI_PRIMARY), ("!active", UI_MUTED)],

            )

            style.configure(

                "SoftBlue.TCombobox",

                fieldbackground="#f8fbff",

                background="#edf6ff",

                foreground=UI_TEXT,

                bordercolor="#bcd2ee",

                font=ui_font(11),

                lightcolor="#bcd2ee",

                darkcolor="#bcd2ee",

                arrowcolor="#5f728b",

                relief="flat",

                padding=(10, 7),

                arrowsize=14,

            )

            style.map(

                "SoftBlue.TCombobox",

                fieldbackground=[("readonly", "#f8fbff"), ("focus", "#eef7ff")],

                background=[("readonly", "#edf6ff"), ("active", "#dceeff")],

                bordercolor=[("focus", UI_PRIMARY), ("!focus", "#bcd2ee")],

                arrowcolor=[("active", UI_PRIMARY), ("!active", "#5f728b")],

                selectbackground=[("readonly", "#dbeafe")],

                selectforeground=[("readonly", UI_TEXT)],

            )

            style.configure(

                "TEntry",

                fieldbackground=UI_SURFACE,

                foreground=UI_TEXT,

                bordercolor=UI_BORDER,

                font=ui_font(11),

                lightcolor=UI_BORDER,

                darkcolor=UI_BORDER,

                relief="flat",

                padding=(10, 7),

            )

            style.map("TEntry", bordercolor=[("focus", UI_PRIMARY), ("!focus", UI_BORDER)])

            style.configure(

                "Treeview",

                background=UI_SURFACE,

                fieldbackground=UI_SURFACE,

                foreground=UI_TEXT,

                rowheight=row_height,

                font=ui_font(11),

                bordercolor=UI_BORDER,

                lightcolor=UI_BORDER,

                darkcolor=UI_BORDER,

            )

            style.configure(

                "Treeview.Heading",

                background=UI_SURFACE_2,

                foreground=UI_TEXT,

                font=ui_font(11, bold=True),

                relief="flat",

                padding=heading_pad,

            )

            style.map("Treeview", background=[("selected", "#dbeafe")], foreground=[("selected", UI_TEXT)])

            style.configure(

                "Preview.Treeview",

                background="#f7f7f3",

                fieldbackground="#f7f7f3",

                foreground="#1f2933",

                font=ui_font(11),

                rowheight=row_height,

                bordercolor="#cfd7dd",

                lightcolor="#cfd7dd",

                darkcolor="#cfd7dd",

            )

            style.configure(

                "Preview.Treeview.Heading",

                background="#6b6358",

                foreground="#ffffff",

                font=ui_font(11, bold=True),

                relief="flat",

                padding=heading_pad,

            )

            style.map(

                "Preview.Treeview",

                background=[("selected", "#cfe5ff")],

                foreground=[("selected", UI_TEXT)],

            )

            style.map(

                "Preview.Treeview.Heading",

                background=[("active", "#5d554c"), ("!active", "#6b6358")],

                foreground=[("active", "#ffffff"), ("!active", "#ffffff")],

            )

            style.configure(

                "Vertical.TScrollbar",

                gripcount=0,

                background="#c8d5e6",

                darkcolor="#c8d5e6",

                lightcolor="#c8d5e6",

                troughcolor="#eef3f8",

                bordercolor="#eef3f8",

                arrowcolor=UI_MUTED,

                relief="flat",

                width=14,

            )

            style.configure(

                "Horizontal.TScrollbar",

                gripcount=0,

                background="#c8d5e6",

                darkcolor="#c8d5e6",

                lightcolor="#c8d5e6",

                troughcolor="#eef3f8",

                bordercolor="#eef3f8",

                arrowcolor=UI_MUTED,

                relief="flat",

                width=14,

            )

            style.map(

                "Vertical.TScrollbar",

                background=[("active", "#9fb2cc"), ("pressed", "#7f96b5")],

                arrowcolor=[("active", UI_PRIMARY), ("!active", UI_MUTED)],

            )

            style.map(

                "Horizontal.TScrollbar",

                background=[("active", "#9fb2cc"), ("pressed", "#7f96b5")],

                arrowcolor=[("active", UI_PRIMARY), ("!active", UI_MUTED)],

            )

        except Exception:

            pass


    def _ui_asset_image(self, relative_path, size=None, alpha=1.0):
        try:
            path = resource_path(*Path(relative_path).parts)
            if not path.exists():
                return None
            img = Image.open(path).convert("RGBA")
            bbox = img.getchannel("A").getbbox()
            if bbox:
                img = img.crop(bbox)
            if size:
                img.thumbnail((scale_px(size[0]), scale_px(size[1])), Image.Resampling.LANCZOS)
            if alpha < 1:
                channel = img.getchannel("A").point(lambda v: int(v * alpha))
                img.putalpha(channel)
            tk_img = ImageTk.PhotoImage(img)
            self._ui_images.append(tk_img)
            return tk_img
        except Exception:
            return None

    def _ui_icon(self, filename, size=22):
        return self._ui_asset_image(APP_UI_ICON_DIR / filename, (size, size))

    def _footer_server_state(self):
        try:
            url = resolve_presence_server_url(getattr(self, "presence_server_var", tk.StringVar(value=DEFAULT_PRESENCE_SERVER_URL)).get())
        except Exception:
            url = ""
        try:
            if is_admin_build() and self._presence_server_is_running():
                return f"Đang chạy - {url}" if url else "Đang chạy"
            online = getattr(self, "_presence_server_online", None)
            if online is True:
                return f"Đang chạy - {url}" if url else "Đang chạy"
            if online is False:
                return f"Bảo trì - {url}" if url else "Bảo trì"
            return f"Đang kiểm tra - {url}" if url else "Đang kiểm tra"
        except Exception:
            return url or "Không rõ"

    def _refresh_footer_clock(self):
        try:
            now = datetime.now()
            if getattr(self, "footer_time_var", None) is not None:
                self.footer_time_var.set(now.strftime("%H:%M:%S"))
            if getattr(self, "footer_date_var", None) is not None:
                self.footer_date_var.set(now.strftime("%d/%m/%Y"))
            if getattr(self, "footer_server_var", None) is not None:
                self.footer_server_var.set(self._footer_server_state())
            if getattr(self, "footer_status_var", None) is not None:
                self.footer_status_var.set(getattr(self, "_last_status_text", "Sẵn sàng"))
        except Exception:
            pass
        try:
            self.root.after(1000, self._refresh_footer_clock)
        except Exception:
            pass

    def _refresh_mapping_stats(self):
        try:
            vars_ = list(getattr(getattr(self, "mapping_editor", None), "mapping_vars", []) or [])
            total = len(vars_)
            mapped = sum(1 for var in vars_ if str(var.get() or "").strip())
            unmapped = max(0, total - mapped)
            if getattr(self, "mapping_stat_mapped_var", None) is not None:
                self.mapping_stat_mapped_var.set(str(mapped))
            if getattr(self, "mapping_stat_unmapped_var", None) is not None:
                self.mapping_stat_unmapped_var.set(str(unmapped))
            if getattr(self, "mapping_stat_total_var", None) is not None:
                self.mapping_stat_total_var.set(str(total))
        except Exception:
            pass
        try:
            self.root.after(1000, self._refresh_mapping_stats)
        except Exception:
            pass



    def build_ui(self):

        self.setup_window_icon()

        self.setup_theme()

        self._ui_images = []
        sidebar_bg = "#053f32"
        sidebar_bg_2 = "#0b5a45"
        sidebar_active = "#0f8d6d"
        sidebar_text = "#d8f3e8"
        sidebar_muted = "#9fcec0"
        main_bg = "#eef7fb"

        def image_label(parent, image, bg, **kwargs):
            if image is None:
                return tk.Label(parent, bg=bg, **kwargs)
            return tk.Label(parent, image=image, bg=bg, **kwargs)

        def icon_button(parent, text, command, icon_file=None, variant="default", width=0):
            btn = ui_button(parent, text, command, width=width, variant=variant)
            if icon_file:
                try:
                    btn.configure(image=self._ui_icon(icon_file, 18), compound="left")
                except Exception:
                    pass
            return btn

        def panel_heading(parent, icon_file, title, bg=UI_SURFACE):
            row = tk.Frame(parent, bg=bg)
            row.pack(fill="x")
            icon = self._ui_icon(icon_file, 22) if icon_file else None
            image_label(row, icon, bg).pack(side="left", padx=(0, 8))
            tk.Label(row, text=title, font=ui_font(10, bold=True), bg=bg, fg=UI_TEXT).pack(side="left")
            return row



        def card(parent, padx=None, pady=None):

            if padx is None:

                padx = self.card_padx

            if pady is None:

                pady = self.card_pady

            frame = tk.Frame(

                parent,

                bg=UI_SURFACE,

                padx=padx,

                pady=pady,

                highlightthickness=1,

                highlightbackground="#e6edf5",

            )

            return frame



        def section_title(parent, title, subtitle=None, title_size=10):

            tk.Label(parent, text=title, font=ui_font(title_size, bold=True), bg=UI_SURFACE, fg=UI_TEXT).pack(anchor="w")

            if subtitle:

                tk.Label(parent, text=subtitle, font=ui_font(10), bg=UI_SURFACE, fg=UI_MUTED).pack(anchor="w", pady=(2, 0))



        shell = tk.Frame(self.root, bg=main_bg)

        shell.pack(fill="both", expand=True)



        sidebar = tk.Frame(

            shell,

            width=self.sidebar_w,

            bg=sidebar_bg,

            padx=6 if self.tiny_ui else (8 if self.compact_ui else 10),

            pady=8 if self.tiny_ui else (10 if self.compact_ui else 14),

            highlightthickness=1,

            highlightbackground=sidebar_bg_2,

        )

        sidebar.pack(side="left", fill="y")

        sidebar.pack_propagate(False)



        brand = tk.Frame(sidebar, bg=sidebar_bg)

        brand.pack(fill="x", pady=(0, 14 if self.tiny_ui else 24))

        logo_file = resource_path(*APP_LOGO_PNG.parts)

        try:

            if logo_file.exists():

                logo_source = Image.open(logo_file).convert("RGBA")

                bbox = logo_source.getchannel("A").getbbox()

                if bbox:

                    logo_source = logo_source.crop(bbox)

                logo_source.thumbnail(self.logo_max, Image.LANCZOS)

                self.app_logo_img = ImageTk.PhotoImage(logo_source)

                tk.Label(brand, image=self.app_logo_img, bg=sidebar_bg).pack(anchor="center")

        except Exception:

            self.app_logo_img = None



        nav_items = [

            ("home", "04_sidebar_home.png", "Trang chủ", True),

            ("excel", "05_sidebar_excel.png", "Excel", False),

            ("history", "06_sidebar_history.png", "Lịch sử", False),

            ("mapping", "07_sidebar_mapping.png", "Mẫu mapping", False),

            ("settings", "08_sidebar_settings.png", "Cài đặt", False),

            ("help", "09_sidebar_help.png", "Trợ giúp", False),

            ("about", "10_sidebar_info.png", "Giới thiệu", False),

        ]

        for page_id, icon, text, active in nav_items:

            bg = sidebar_active if active else sidebar_bg

            fg = "#ffffff" if active else sidebar_text

            row = tk.Frame(sidebar, bg=bg, padx=0, pady=0, highlightthickness=1, highlightbackground=sidebar_active if active else sidebar_bg_2)

            row.pack(fill="x", pady=2)

            inner = tk.Frame(row, bg=bg, padx=8, pady=8)

            inner.pack(fill="both", expand=True)

            accent = tk.Frame(inner, bg="#f4c542" if active else bg, width=4, height=22)

            accent.pack(side="left", padx=(0, 8), fill="y")

            icon_img = self._ui_icon(icon, 20)
            icon_label = image_label(inner, icon_img, bg, width=24)

            icon_label.pack(side="left")

            text_label = tk.Label(
                inner,
                text=text,
                bg=bg,
                fg=fg,
                font=ui_font(12, bold=active),
                anchor="w",
                justify="left",
                wraplength=max(96, self.sidebar_w - 60),
            )

            text_label.pack(side="left", padx=(4 if self.tiny_ui else 6, 0), fill="x", expand=True)

            self.nav_widgets[page_id] = {"row": row, "inner": inner, "accent": accent, "icon": icon_label, "label": text_label}

            if page_id == "excel":

                for widget in (row, inner, accent, icon_label, text_label):

                    widget.bind("<Button-1>", lambda _e, p=page_id: self._activate_nav_item(p, self.show_excel_page))

            elif page_id == "home":

                for widget in (row, inner, accent, icon_label, text_label):

                    widget.bind("<Button-1>", lambda _e, p=page_id: self._activate_nav_item(p, self.show_home_page))

            elif page_id == "history":

                for widget in (row, inner, accent, icon_label, text_label):

                    widget.bind("<Button-1>", lambda _e, p=page_id: self._activate_nav_item(p, self.show_history_page))

            elif page_id == "mapping":

                for widget in (row, inner, accent, icon_label, text_label):

                    widget.bind("<Button-1>", lambda _e, p=page_id: self._activate_nav_item(p, self.show_mapping_page))

            elif page_id == "settings":

                for widget in (row, inner, accent, icon_label, text_label):

                    widget.bind("<Button-1>", lambda _e, p=page_id: self._activate_nav_item(p, self.show_settings_dialog))

            elif page_id == "help":

                for widget in (row, inner, accent, icon_label, text_label):

                    widget.bind("<Button-1>", lambda _e, p=page_id: self._activate_nav_item(p, self.show_help_dialog))

            elif page_id == "about":

                for widget in (row, inner, accent, icon_label, text_label):

                    widget.bind("<Button-1>", lambda _e, p=page_id: self._activate_nav_item(p, self.show_about_dialog))



        status_card = tk.Frame(sidebar, bg="#0b5a45", highlightthickness=1, highlightbackground="#19765d")
        status_card.pack(fill="x", pady=(10, 0))
        status_inner = tk.Frame(status_card, bg="#0b5a45", padx=8, pady=8)
        status_inner.pack(fill="both", expand=True)
        tk.Label(status_inner, text="Kết nối server", bg="#0b5a45", fg=sidebar_muted, font=ui_font(9, bold=True)).pack(anchor="w", pady=(0, 4))
        self.status = tk.Label(
            status_inner,
            text="● Sẵn sàng",
            anchor="center",
            fg=UI_SUCCESS,
            bg="#0b5a45",
            font=ui_font(8 if (self.tiny_ui or self.micro_ui) else 9, bold=True),
            wraplength=max(88, self.sidebar_w - 62),
            justify="center",
        )
        self.status.pack(fill="both", expand=True)

        sidebar_spacer = tk.Frame(sidebar, bg=sidebar_bg, height=8 if (self.tiny_ui or self.micro_ui) else 12)

        sidebar_spacer.pack(fill="x")

        user_box = card(sidebar, padx=4 if (self.tiny_ui or self.micro_ui) else 5, pady=3 if (self.tiny_ui or self.micro_ui) else 5)

        user_box.configure(bg="#0b5a45", highlightbackground="#19765d")

        user_box.pack(fill="x", pady=(8 if is_admin_build() else 6, 0))
        user_box.pack_propagate(False)

        user_inner = tk.Frame(user_box, bg="#0b5a45")

        user_inner.pack(fill="both", expand=True)

        member_content = tk.Frame(user_inner, bg="#0b5a45")

        member_content.pack(fill="both", expand=True, pady=(5 if (self.tiny_ui or self.micro_ui) else 7, 4))

        if self.tiny_ui or self.micro_ui:
            self.sidebar_member_role_label = tk.Label(
                member_content,
                textvariable=self.user_role_var,
                font=ui_font(10, bold=True),
                bg="#0b5a45",
                fg="#ffffff",
                justify="center",
                wraplength=max(100, self.sidebar_w - 44),
            )
            self.sidebar_member_role_label.pack(anchor="center")

            self.sidebar_member_name_label = tk.Label(
                member_content,
                text=self.user_name,
                font=ui_font(9),
                bg="#0b5a45",
                fg=sidebar_muted,
                justify="center",
                wraplength=max(100, self.sidebar_w - 44),
            )
            self.sidebar_member_name_label.pack(anchor="center", pady=(2, 0))
        else:
            tk.Label(
                member_content,
                textvariable=self.user_role_var,
                font=ui_font(11, bold=True),
                bg="#0b5a45",
                fg="#ffffff",
                justify="center",
                wraplength=max(110, self.sidebar_w - 44),
            ).pack(anchor="center")

            tk.Label(
                member_content,
                text=self.user_name,
                font=ui_font(10),
                bg="#0b5a45",
                fg=sidebar_muted,
                justify="center",
                wraplength=max(110, self.sidebar_w - 44),
            ).pack(anchor="center", pady=(4, 0))

        if is_admin_build():

            admin_btn_row = tk.Frame(member_content, bg="#0b5a45")

            admin_btn_row.pack(anchor="center", pady=(6 if (self.tiny_ui or self.micro_ui) else 8, 0))

            ui_button(
                admin_btn_row,
                "Duyệt máy",
                self.open_admin_approval_panel,
                width=10 if (self.tiny_ui or self.micro_ui) else 11,
                variant="warn",
            ).pack(anchor="center")

        if not is_admin_build():
            log_btn_row = tk.Frame(member_content, bg="#0b5a45")
            log_btn_row.pack(anchor="center", pady=(5 if (self.tiny_ui or self.micro_ui) else 7, 0))
            self._log_btn = ui_button(
                log_btn_row,
                "Gửi log",
                self._send_log_to_admin,
                width=9 if (self.tiny_ui or self.micro_ui) else 10,
                variant="soft",
            )
            self._log_btn.pack(anchor="center")

        sidebar_decor = self._ui_asset_image(APP_DECOR_SIDEBAR_BOTTOM, (self.sidebar_w - 12, 180), alpha=0.9)
        if sidebar_decor is not None:
            tk.Label(sidebar, image=sidebar_decor, bg=sidebar_bg, bd=0).pack(side="bottom", fill="x", pady=(8, 0))

        main = tk.Frame(shell, bg=main_bg, padx=self.main_padx, pady=self.main_pady)

        main.pack(side="left", fill="both", expand=True)



        header = tk.Frame(main, bg="#f8fcff", padx=14, pady=10, highlightthickness=1, highlightbackground="#dbe8f1")

        header.pack(fill="x")

        title_box = tk.Frame(header, bg="#f8fcff")

        title_box.pack(side="left", fill="x", expand=True)

        if APP_TITLE:

            tk.Label(title_box, text=APP_TITLE, font=ui_font(16, bold=True), bg="#f8fcff", fg=UI_TEXT).pack(anchor="w")

        tk.Label(title_box, text="Ứng dụng Phục hồi & Quản lý Dữ liệu Cọc", font=ui_font(10), bg="#f8fcff", fg=UI_MUTED).pack(anchor="w", pady=(3, 0))

        if not self.tiny_ui:

            if is_admin_build():
                notify_button = tk.Canvas(
                    header,
                    width=38,
                    height=36,
                    bg="#f8fcff",
                    bd=0,
                    highlightthickness=0,
                    cursor="hand2",
                )
                notify_button.pack(side="left", padx=4)
                notify_button.create_oval(4, 4, 34, 34, fill="#ffffff", outline="#dbe3ef", width=1)
                bell_img = self._ui_icon("11_header_notification.png", 18)
                if bell_img is not None:
                    notify_button.create_image(19, 19, image=bell_img)
                else:
                    notify_button.create_text(19, 19, text="🔔", fill="#111827", font=("Segoe UI Emoji", 15))
                self.admin_log_notify_canvas = notify_button
                self.admin_log_badge_oval = notify_button.create_oval(
                    23,
                    1,
                    37,
                    15,
                    fill=UI_ERROR,
                    outline=UI_ERROR,
                    state="hidden",
                )
                self.admin_log_badge_text = notify_button.create_text(
                    30,
                    8,
                    text="",
                    fill="#ffffff",
                    font=ui_font(6, bold=True),
                    state="hidden",
                )
                notify_button.bind("<Button-1>", lambda _e: self.open_admin_log_panel())
                self.root.after(500, self._admin_log_badge_loop)

            avatar_img = self._ui_icon("12_header_avatar.png", 34)
            image_label(header, avatar_img, "#f8fcff", text="A", fg="#ffffff", font=ui_font(10, bold=True)).pack(side="left", padx=(10, 6))

            profile = tk.Frame(header, bg="#f8fcff")

            profile.pack(side="left")

            tk.Label(profile, textvariable=self.user_role_var, font=ui_font(10), bg="#f8fcff", fg=UI_MUTED, justify="center").pack(anchor="center")

            tk.Label(profile, text=self.user_name, font=ui_font(10, bold=True), bg="#f8fcff", fg=UI_TEXT, justify="center").pack(anchor="center", pady=(1, 0))
            dropdown_img = self._ui_icon("13_header_dropdown.png", 14)
            image_label(header, dropdown_img, "#f8fcff").pack(side="left", padx=(8, 0))


        self.content_canvas = None
        self._content_window_id = None
        content = tk.Frame(main, bg=main_bg)
        content.pack(fill="both", expand=True)

        footer = tk.Frame(main, bg="#f8fcff", padx=12, pady=7, highlightthickness=1, highlightbackground="#dbe8f1")
        footer.pack(fill="x", pady=(8, 0))
        self.footer_server_var = tk.StringVar(value=self._footer_server_state())
        self.footer_status_var = tk.StringVar(value=getattr(self, "_last_status_text", "Sẵn sàng"))
        self.footer_time_var = tk.StringVar(value=datetime.now().strftime("%H:%M:%S"))
        self.footer_date_var = tk.StringVar(value=datetime.now().strftime("%d/%m/%Y"))
        self.footer_status_dot = None

        def footer_item(label, variable=None, static_text=None, dot_file=None):
            item = tk.Frame(footer, bg="#f8fcff")
            item.pack(side="left", padx=(0, 18))
            if dot_file:
                dot_img = self._ui_icon(dot_file, 11)
                dot = image_label(item, dot_img, "#f8fcff", text="●", fg=UI_SUCCESS, font=ui_font(9, bold=True))
                dot.pack(side="left", padx=(0, 5))
                if dot_file == "31_footer_ready_status_dot.png":
                    self.footer_status_dot = dot
            tk.Label(item, text=f"{label}:", bg="#f8fcff", fg=UI_MUTED, font=ui_font(9, bold=True)).pack(side="left")
            if variable is not None:
                tk.Label(item, textvariable=variable, bg="#f8fcff", fg=UI_TEXT, font=ui_font(9)).pack(side="left", padx=(4, 0))
            else:
                tk.Label(item, text=static_text or "", bg="#f8fcff", fg=UI_TEXT, font=ui_font(9)).pack(side="left", padx=(4, 0))

        footer_item("Phiên bản", static_text="1.0.0")
        footer_item("Máy chủ", variable=self.footer_server_var, dot_file="30_footer_running_server_dot.png")
        footer_item("Trạng thái", variable=self.footer_status_var, dot_file="31_footer_ready_status_dot.png")
        footer_item("Thời gian", variable=self.footer_time_var)
        footer_item("Ngày", variable=self.footer_date_var)
        footer_decor = self._ui_asset_image(APP_DECOR_BOTTOM_RIGHT, (270, 92), alpha=0.62)
        if footer_decor is not None:
            tk.Label(footer, image=footer_decor, bg="#f8fcff", bd=0).pack(side="right", padx=(12, 0))
        self.root.after(250, self._refresh_footer_clock)

        self.home_page = tk.Frame(content, bg=main_bg)

        self.excel_page = tk.Frame(content, bg=main_bg)

        self.home_page.pack(fill="both", expand=True)



        toolbar = card(
            self.home_page,
            padx=6 if self.micro_ui else (8 if self.tiny_ui else 12),
            pady=5 if self.micro_ui else (7 if self.tiny_ui else 10),
        )
        toolbar.pack(fill="x", pady=(6 if self.micro_ui else (8 if self.tiny_ui else 12), 6 if self.micro_ui else (8 if self.tiny_ui else 10)))

        toolbar_inner = tk.Frame(toolbar, bg=UI_SURFACE)

        toolbar_inner.pack(fill="x")

        toolbar_inner.grid_columnconfigure(0, weight=5, uniform="toolbar")

        toolbar_inner.grid_columnconfigure(1, weight=0)

        toolbar_inner.grid_columnconfigure(2, weight=3, uniform="toolbar")

        toolbar_inner.grid_columnconfigure(3, weight=0)

        toolbar_inner.grid_columnconfigure(4, weight=2, uniform="toolbar")

        toolbar_gap = 2

        toolbar_group_pad = 4 if self.dense_ui else 6

        toolbar_sep_pad = 6 if self.dense_ui else 10



        def make_group(parent, column, title, btns, bg_color):

            group = tk.Frame(parent, bg=bg_color)

            group.grid(row=0, column=column, sticky="nsew", padx=(toolbar_group_pad, toolbar_group_pad))

            if title:

                tk.Label(group, text=title, font=ui_font(10, bold=True), fg=UI_MUTED, bg=bg_color).pack(side="top", anchor="center", pady=(0, 5))

            rows = [btns]

            for row_btns in rows:

                btn_row = tk.Frame(group, bg=bg_color)

                btn_row.pack(side="top", anchor="center", pady=(1, 0))

                for text, command, variant, icon_file in row_btns:

                    icon_button(btn_row, text, command, icon_file, width=0, variant=variant).pack(side="left", padx=(toolbar_gap, toolbar_gap))

            return group

        source_btns = [

            ("Chọn Excel", self.choose_excel, "primary", "14_action_import_excel.png"),

            ("Workbook" if self.compact_ui else "Đọc workbook", self.scan_current_workbook, "default", "15_action_import_workbook.png"),

            ("Từng sheet" if self.compact_ui else "Đọc từng sheet", self.read_each_sheet_content, "default", "16_action_read_sheet_list.png"),

            ("Công thức" if self.compact_ui else "Đọc công thức", self.read_current_excel_formulas, "default", "17_action_read_formula.png"),

            ("Đọc lại" if self.compact_ui else "Đọc lại Excel", self.refresh_excel_header_info, "soft", "18_action_read_excel.png"),

            ("Đặt lại", self.reset_current_session, "warn", "19_action_refresh.png"),

        ]

        make_group(toolbar_inner, 0, "NGUỒN DỮ LIỆU", source_btns, UI_SURFACE)
        
        separator1 = tk.Frame(toolbar_inner, bg="#e2e8f0", width=1)

        separator1.grid(row=0, column=1, sticky="ns", padx=(toolbar_sep_pad, toolbar_sep_pad), pady=4)

        process_btns = [

            ("Đọc bảng", self.run_gemini, "soft", "20_action_read_table.png"),

            ("Phiếu cọc" if self.compact_ui else "Đọc phiếu cọc", self.run_gemini_phieu_coc, "soft", "21_action_read_column.png"),

            ("Auto map", self.build_mapping, "warn", "22_action_auto_map.png"),

        ]

        make_group(toolbar_inner, 2, "XỬ LÝ DỮ LIỆU", process_btns, UI_SURFACE)
        
        separator2 = tk.Frame(toolbar_inner, bg="#e2e8f0", width=1)

        separator2.grid(row=0, column=3, sticky="ns", padx=(toolbar_sep_pad, toolbar_sep_pad), pady=4)

        export_btns = [

            ("Xem trước", self.preview_excel, "soft", "23_action_preview.png"),

            ("Xuất Excel" if self.compact_ui else "Xuất ra Excel", self.fill_excel, "success", "24_action_export_excel.png"),

        ]

        make_group(toolbar_inner, 4, "XEM & XUẤT", export_btns, UI_SURFACE)

        filters = card(self.home_page, padx=10 if self.tiny_ui else 14, pady=8 if self.tiny_ui else 9)

        self.filters_card = filters

        filters.pack(fill="x", pady=(0, 12))

        filter_top = tk.Frame(filters, bg=UI_SURFACE)

        filter_top.pack(fill="x")

        filter_bottom = filter_top



        tk.Label(filter_top, text="Sheet:", bg=UI_SURFACE, fg=UI_TEXT, font=ui_font(11, bold=True)).pack(side="left", padx=(4, 8))

        self.sheet_combo = RoundedMappingDropdown(

            filter_top,

            values=[],

            variable=self.sheet_var,

            bg_color="#f8fbff",

            border_color="#bcd2ee",

            width=150 if self.tiny_ui else (180 if self.compact_ui else 220),

            height=34,

            radius=8,

        )

        self.sheet_combo.pack(side="left", padx=(0, 10 if self.compact_ui else 18))

        self.sheet_combo.bind("<<ComboboxSelected>>", lambda e: self.refresh_excel_header_info())

        tk.Label(filter_bottom, text="Chế độ đọc bảng:", bg=UI_SURFACE, fg=UI_MUTED).pack(side="left", padx=(4, 8))

        self.template_combo = RoundedMappingDropdown(

            filter_bottom,

            values=["Bảng bất kỳ - tự nhận cột"],

            variable=self.template_var,

            bg_color="#f8fbff",

            border_color="#bcd2ee",

            width=220 if self.tiny_ui else (260 if self.compact_ui else 330),

            height=34,

            radius=8,

        )

        self.template_combo["values"] = ["Bảng bất kỳ - tự nhận cột"]

        self.template_combo.pack(side="left")

        status_pill_img = self._ui_icon("29_status_indicator_pill.png", 108)
        status_pill = tk.Frame(filter_top, bg=UI_SURFACE)
        status_pill.pack(side="right", padx=(12, 4))
        image_label(status_pill, status_pill_img, UI_SURFACE).pack(side="left", padx=(0, 6))
        tk.Label(status_pill, text="Trạng thái:", bg=UI_SURFACE, fg=UI_MUTED, font=ui_font(10, bold=True)).pack(side="left")
        self.filter_status_var = tk.StringVar(value=getattr(self, "_last_status_text", "Sẵn sàng"))
        tk.Label(status_pill, textvariable=self.filter_status_var, bg="#e8f7ef", fg=UI_SUCCESS, font=ui_font(10, bold=True), padx=10, pady=5).pack(side="left", padx=(6, 0))

        if self.main_content_scroll or self.short_ui or self.dense_ui or self.tiny_ui or self.micro_ui:
            home_body_shell = tk.Frame(self.home_page, bg=UI_BG)
            home_body_shell.pack(fill="both", expand=True)

            self.home_body_canvas = tk.Canvas(home_body_shell, bg=UI_BG, highlightthickness=0, bd=0)
            home_body_scroll = ttk.Scrollbar(home_body_shell, orient="vertical", command=self.home_body_canvas.yview)
            self.home_body_scrollbar = home_body_scroll
            self.home_body_canvas.configure(yscrollcommand=home_body_scroll.set)
            self.home_body_canvas.pack(side="left", fill="both", expand=True)

            home_body_content = tk.Frame(self.home_body_canvas, bg=UI_BG)
            self._home_body_window_id = self.home_body_canvas.create_window((0, 0), window=home_body_content, anchor="nw")

            def _sync_home_body_scrollregion(_event=None):
                try:
                    self.home_body_canvas.configure(scrollregion=self.home_body_canvas.bbox("all"))
                    bbox = self.home_body_canvas.bbox("all")
                    needs_scroll = bool(bbox and self.home_body_canvas.winfo_height() and (bbox[3] - bbox[1] > self.home_body_canvas.winfo_height() + 2))
                    if needs_scroll:
                        if not home_body_scroll.winfo_ismapped():
                            home_body_scroll.pack(side="right", fill="y")
                    else:
                        if home_body_scroll.winfo_ismapped():
                            home_body_scroll.pack_forget()
                except Exception:
                    pass

            def _sync_home_body_width(event):
                try:
                    self.home_body_canvas.itemconfigure(self._home_body_window_id, width=event.width)
                except Exception:
                    pass

            home_body_content.bind("<Configure>", _sync_home_body_scrollregion)
            self.home_body_canvas.bind("<Configure>", _sync_home_body_width)
            self.home_body_canvas.bind("<Enter>", self._bind_main_content_mousewheel)
            self.home_body_canvas.bind("<Leave>", self._unbind_main_content_mousewheel)
            self.root.after_idle(_sync_home_body_scrollregion)
        else:
            self.home_body_canvas = None
            self._home_body_window_id = None
            self.home_body_scrollbar = None
            home_body_content = tk.Frame(self.home_page, bg=UI_BG)
            home_body_content.pack(fill="both", expand=True)

        workspace = tk.Frame(home_body_content, bg=main_bg)

        workspace.pack(fill="both", expand=True)

        workflow = workspace

        left_min, center_min, right_min = self.workspace_mins

        workspace.grid_columnconfigure(0, weight=1, minsize=left_min)

        workspace.grid_columnconfigure(1, weight=4, minsize=center_min)

        workspace.grid_columnconfigure(2, weight=2, minsize=right_min)

        workspace.grid_rowconfigure(0, weight=1)



        left_col = tk.Frame(workspace, bg=main_bg)

        left_col.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        left_col.grid_columnconfigure(0, weight=1)
        left_col.grid_rowconfigure(0, weight=1, uniform="left_cards")
        left_col.grid_rowconfigure(1, weight=1, uniform="left_cards")

        image_card = card(left_col)

        image_card.grid(row=0, column=0, sticky="nsew", pady=(0, 8))

        panel_heading(image_card, "25_panel_image_upload.png", "ẢNH OCR")

        upload_box = tk.Frame(
            image_card,
            bg="#fbfdff",
            highlightthickness=1,
            highlightbackground="#d7e3f2",
            padx=8,
            pady=8,
            height=self.image_drop_h,
        )
        upload_box.pack(fill="both", expand=True, pady=(8, 0))
        upload_box.pack_propagate(False)
        upload_box.grid_columnconfigure(0, weight=1)
        upload_box.grid_rowconfigure(0, weight=1)
        upload_box.grid_rowconfigure(1, weight=0)

        preview_shell = tk.Frame(upload_box, bg="#fbfdff")
        preview_shell.grid(row=0, column=0, sticky="nsew", pady=(0, 12))
        preview_shell.grid_columnconfigure(0, weight=1)
        preview_shell.grid_rowconfigure(0, weight=1)

        preview_w = 250 if not (self.tiny_ui or self.micro_ui) else 210
        preview_h = max(96, int(self.image_drop_h) - scale_px(58))

        self.preview_frame = tk.Frame(
            preview_shell,
            bg="#f8fbff",
            highlightthickness=1,
            highlightbackground="#d7e3f2",
            width=preview_w,
            height=preview_h,
        )
        self.preview_frame.grid(row=0, column=0, sticky="nsew")
        self.preview_frame.pack_propagate(False)

        self.img_label = tk.Label(
            self.preview_frame,
            text="Kéo thả ảnh vào đây\nhoặc chọn file từ máy tính",
            bg="#f8fbff",
            fg=UI_TEXT,
            font=ui_font(10, bold=True),
            justify="center",
            cursor="hand2",
        )
        self.img_label.pack(fill="both", expand=True)
        self.img_label.bind("<Button-1>", self.open_current_preview_image)
        self.img_label.bind("<Control-v>", self.paste_image_from_clipboard)
        self.img_label.bind("<Control-V>", self.paste_image_from_clipboard)

        self.preview_prev_btn = tk.Button(
            self.preview_frame,
            text="‹",
            command=lambda: self.move_preview_image(-1),
            bg="#ffffff",
            fg=UI_TEXT,
            activebackground="#e8f1ff",
            activeforeground=UI_TEXT,
            relief="flat",
            bd=0,
            font=ui_font(18, bold=True),
            cursor="hand2",
            width=2,
            height=1,
        )
        self.preview_next_btn = tk.Button(
            self.preview_frame,
            text="›",
            command=lambda: self.move_preview_image(1),
            bg="#ffffff",
            fg=UI_TEXT,
            activebackground="#e8f1ff",
            activeforeground=UI_TEXT,
            relief="flat",
            bd=0,
            font=ui_font(18, bold=True),
            cursor="hand2",
            width=2,
            height=1,
        )
        self.preview_prev_btn.place(relx=0.03, rely=0.5, anchor="w")
        self.preview_next_btn.place(relx=0.97, rely=0.5, anchor="e")
        self.preview_prev_btn.lower()
        self.preview_next_btn.lower()

        self._preview_hover_hide_job = None

        def _preview_pointer_inside():
            try:
                x, y = self.preview_frame.winfo_pointerxy()
                widget = self.preview_frame.winfo_containing(x, y)
                while widget is not None:
                    if widget is self.preview_frame:
                        return True
                    widget = getattr(widget, "master", None)
            except Exception:
                pass
            return False

        def _preview_hover(_event=None):
            try:
                if self._preview_hover_hide_job:
                    self.preview_frame.after_cancel(self._preview_hover_hide_job)
                    self._preview_hover_hide_job = None
            except Exception:
                pass
            self.preview_prev_btn.lift()
            self.preview_next_btn.lift()

        def _preview_leave(_event=None):
            def _hide_if_outside():
                self._preview_hover_hide_job = None
                if not _preview_pointer_inside():
                    self.preview_prev_btn.lower()
                    self.preview_next_btn.lower()

            try:
                if self._preview_hover_hide_job:
                    self.preview_frame.after_cancel(self._preview_hover_hide_job)
                self._preview_hover_hide_job = self.preview_frame.after(140, _hide_if_outside)
            except Exception:
                _hide_if_outside()

        self.preview_frame.bind("<Enter>", _preview_hover)
        self.preview_frame.bind("<Leave>", _preview_leave)
        self.img_label.bind("<Enter>", _preview_hover)
        self.img_label.bind("<Leave>", _preview_leave)
        self.preview_prev_btn.bind("<Enter>", _preview_hover)
        self.preview_prev_btn.bind("<Leave>", _preview_leave)
        self.preview_next_btn.bind("<Enter>", _preview_hover)
        self.preview_next_btn.bind("<Leave>", _preview_leave)

        self.preview_counter_var = tk.StringVar(value="")
        self.preview_counter_label = tk.Label(
            upload_box,
            textvariable=self.preview_counter_var,
            bg="#fbfdff",
            fg=UI_MUTED,
            font=ui_font(10),
        )

        ui_button(upload_box, "Chọn ảnh / Tải lên", self.choose_image, width=18, variant="primary").grid(
            row=1,
            column=0,
            pady=(0, 2),
        )



        info = card(left_col)

        info.grid(row=1, column=0, sticky="nsew", pady=(8, 0))

        panel_heading(info, None, "THÔNG TIN EXCEL")

        self.excel_info = tk.Text(

            info,

            height=6 if self.micro_ui else (8 if self.short_ui or self.dense_ui else 10),

            wrap="word",

            bg="#fbfdff",

            fg=UI_TEXT,

            relief="flat",

            bd=0,

            highlightthickness=1,

            highlightbackground=UI_BORDER,

            font=ui_font(10),

        )

        self.excel_info.pack(fill="both", expand=True, pady=(10, 0))

        icon_button(info, "Chọn file Excel", self.choose_excel, "14_action_import_excel.png", width=18, variant="primary").pack(fill="x", pady=(10, 0))



        center_col = card(workspace)

        center_col.grid(row=0, column=1, sticky="nsew", padx=(0, 8))

        panel_heading(center_col, "26_panel_table_grid.png", "PREVIEW BẢNG")
        tk.Label(center_col, text="Kiểm tra và sửa dữ liệu trước khi đưa vào Excel", font=ui_font(10), bg=UI_SURFACE, fg=UI_MUTED).pack(anchor="w", pady=(2, 8))

        self.table_editor = TableEditor(center_col)
        self.table_editor.on_change = self._refresh_daily_summary_panel



        right_col = card(workspace)

        right_col.mapping_canvas_h = self.mapping_canvas_h

        right_col.grid(row=0, column=2, sticky="nsew")

        panel_heading(right_col, "27_panel_mapping_document.png", "XÁC NHẬN MAPPING CỘT")
        tk.Label(right_col, text="Kéo thả để ánh xạ dữ liệu giữa 2 nguồn", font=ui_font(10), bg=UI_SURFACE, fg=UI_MUTED).pack(anchor="w", pady=(2, 8))

        summary_card = tk.Frame(right_col, bg=UI_SURFACE)
        summary_card.pack(fill="x", pady=(8, 0))

        tk.Label(
            summary_card,
            text="TỔNG HỢP THEO NGÀY",
            bg=UI_SURFACE,
            fg=UI_TEXT,
            font=ui_font(10, bold=True),
            anchor="w",
        ).pack(anchor="w")

        self.daily_summary_text = tk.Text(
            summary_card,
            height=3 if self.micro_ui else (4 if self.short_ui or self.dense_ui else 5),
            wrap="word",
            bg="#fbfdff",
            fg=UI_TEXT,
            relief="flat",
            bd=0,
            highlightthickness=1,
            highlightbackground=UI_BORDER,
            font=ui_font(10),
        )
        self.daily_summary_text.pack(fill="x", expand=False, pady=(6, 0))

        mapping_action_row = tk.Frame(right_col, bg=UI_SURFACE)

        mapping_action_row.pack(fill="x", pady=(8, 0))

        icon_button(mapping_action_row, "Lưu mẫu", self.save_current_mapping_template, "28_panel_save.png", width=10, variant="soft").pack(anchor="e")

        self.mapping_editor = MappingEditor(right_col)

        stats = tk.Frame(right_col, bg=UI_SURFACE, highlightthickness=1, highlightbackground=UI_BORDER)
        stats.pack(fill="x", pady=(10, 0))
        self.mapping_stat_mapped_var = tk.StringVar(value="0")
        self.mapping_stat_unmapped_var = tk.StringVar(value="0")
        self.mapping_stat_total_var = tk.StringVar(value="0")

        def stat_cell(col, value_var, label, color):
            cell = tk.Frame(stats, bg=UI_SURFACE)
            cell.grid(row=0, column=col, sticky="nsew", padx=1, pady=9)
            tk.Label(cell, textvariable=value_var, bg=UI_SURFACE, fg=color, font=ui_font(16, bold=True)).pack(anchor="center")
            tk.Label(cell, text=label, bg=UI_SURFACE, fg=UI_MUTED, font=ui_font(9)).pack(anchor="center", pady=(2, 0))

        for i in range(3):
            stats.grid_columnconfigure(i, weight=1)
        stat_cell(0, self.mapping_stat_mapped_var, "Đã mapping", UI_SUCCESS)
        stat_cell(1, self.mapping_stat_unmapped_var, "Chưa mapping", UI_WARN)
        stat_cell(2, self.mapping_stat_total_var, "Tổng cột", UI_TEXT)
        self.root.after(300, self._refresh_mapping_stats)

        self._sidebar_member_spacer = sidebar_spacer
        self._sidebar_member_box = user_box
        self._workflow_anchor_card = workflow

        def _sync_sidebar_member_anchor(_event=None):
            try:
                spacer = getattr(self, "_sidebar_member_spacer", None)
                if is_admin_build():
                    try:
                        if spacer is not None:
                            spacer.configure(height=6)
                        member_box = getattr(self, "_sidebar_member_box", None)
                        if member_box is not None:
                            member_box.configure(height=122 if (self.tiny_ui or self.micro_ui) else 136)
                    except Exception:
                        pass
                    return
                member_box = getattr(self, "_sidebar_member_box", None)
                if spacer is None or member_box is None:
                    return
                self.root.update_idletasks()
                if not spacer.winfo_exists() or not member_box.winfo_exists():
                    return
                new_height = scale_px(10 if (self.tiny_ui or self.micro_ui) else 14)
                if new_height != int(str(spacer.cget("height") or 0)):
                    spacer.configure(height=new_height)
                if is_admin_build():
                    target_height = max(0, int(member_box.winfo_reqheight() or 0))
                    target_height = max(target_height, scale_px(138 if (self.tiny_ui or self.micro_ui) else 154))
                else:
                    target_height = scale_px(124 if (self.tiny_ui or self.micro_ui) else 138)
                if target_height and member_box.winfo_height() != target_height:
                    member_box.configure(height=target_height)
            except Exception:
                pass

        self._sync_sidebar_member_anchor = _sync_sidebar_member_anchor
        self.root.after_idle(self._sync_sidebar_member_anchor)
        self.root.bind("<Configure>", self._sync_sidebar_member_anchor, add="+")



        self.history_page = tk.Frame(content, bg=UI_BG, padx=self.main_padx, pady=self.main_pady)

        history_shell = card(self.history_page, padx=16 if self.compact_ui else 18, pady=14)

        history_shell.pack(fill="both", expand=True)

        history_top = tk.Frame(history_shell, bg=UI_SURFACE)

        history_top.pack(fill="x", pady=(0, 10))

        history_title = tk.Frame(history_top, bg=UI_SURFACE)

        history_title.pack(side="left", fill="x", expand=True)

        tk.Label(history_title, text="LỊCH SỬ XỬ LÝ", bg=UI_SURFACE, fg=UI_TEXT, font=ui_font(12, bold=True)).pack(anchor="w")

        tk.Label(

            history_title,

            text="Hiển thị theo ngày, có màu theo loại dữ liệu và trạng thái đọc / xuất Excel",

            bg=UI_SURFACE,

            fg=UI_MUTED,

            font=ui_font(11),

        ).pack(anchor="w", pady=(2, 0))

        filter_row = tk.Frame(history_top, bg=UI_SURFACE)
        filter_row.pack(anchor="w", pady=(10, 0))

        tk.Label(filter_row, text="Tìm ngày / tên:", bg=UI_SURFACE, fg=UI_MUTED, font=ui_font(11)).pack(side="left", padx=(0, 8))

        self.history_filter_var = tk.StringVar(value="")

        filter_shell = RoundedMappingEntry(
            filter_row,
            textvariable=self.history_filter_var,
            bg_color="#f8fbff",
            border_color="#bcd2ee",
            width=200,
            height=30,
            radius=8,
            font=ui_font(11),
        )
        filter_shell.pack(side="left")
        filter_entry = filter_shell.entry
        filter_entry.bind("<KeyRelease>", lambda _e: self._sync_history_view())
        if not getattr(self, "_history_filter_blur_bound", False):
            def _clear_history_filter_focus(event, shell=filter_shell):
                try:
                    if getattr(self, "current_page", "") != "history":
                        return
                    if str(event.widget).startswith(str(shell)):
                        return
                    self.root.focus_set()
                except Exception:
                    pass
            self.root.bind_all("<Button-1>", _clear_history_filter_focus, add="+")
            self._history_filter_blur_bound = True

        ui_button(filter_row, "Xóa lọc", lambda: (self.history_filter_var.set(""), self._sync_history_view()), width=10, variant="soft").pack(side="left", padx=(8, 0))

        tk.Label(filter_row, text="Hiển thị:", bg=UI_SURFACE, fg=UI_MUTED, font=ui_font(11)).pack(side="left", padx=(16, 8))

        self.history_kind_var = tk.StringVar(value="Cả hai")
        self.history_kind_combo = RoundedMappingDropdown(
            filter_row,
            values=["Cả hai", "Khối lượng", "Phiếu cọc"],
            variable=self.history_kind_var,
            bg_color="#f8fbff",
            border_color="#bcd2ee",
            width=150,
            height=30,
            radius=8,
        )
        self.history_kind_combo.pack(side="left")
        self.history_kind_combo.bind("<<ComboboxSelected>>", lambda _e: self._sync_history_view())

        self.history_summary_var = tk.StringVar(value="0 mục")

        tk.Label(

            history_top,

            textvariable=self.history_summary_var,

            bg="#eef4ff",

            fg=UI_PRIMARY,

            font=ui_font(11, bold=True),

            padx=10,

            pady=6,

        ).pack(side="right", padx=(8, 0))

        ui_button(history_top, "Làm mới", self._sync_history_view, width=10, variant="soft").pack(side="right")

        history_body = tk.Frame(history_shell, bg=UI_SURFACE)
        history_body.pack(fill="both", expand=True)

        history_list_panel = tk.Frame(history_body, bg=UI_SURFACE)
        history_list_panel.pack(side="left", fill="both", expand=True)
        self.history_list_panel = history_list_panel

        self.history_canvas = tk.Canvas(history_list_panel, bg=UI_SURFACE, highlightthickness=0, bd=0)
        history_scroll = ttk.Scrollbar(history_list_panel, orient="vertical", command=self.history_canvas.yview)
        self.history_canvas.configure(yscrollcommand=history_scroll.set)
        history_scroll.pack(side="right", fill="y")
        self.history_canvas.pack(side="left", fill="both", expand=True)

        self.history_inner = tk.Frame(self.history_canvas, bg=UI_SURFACE)
        self.history_canvas_window = self.history_canvas.create_window((0, 0), window=self.history_inner, anchor="nw")

        def _sync_history_scrollregion(_event=None):

            try:

                self.history_canvas.configure(scrollregion=self.history_canvas.bbox("all"))

            except Exception:

                pass

        self.history_inner.bind("<Configure>", _sync_history_scrollregion)
        self.history_canvas.bind("<Configure>", lambda e: self.history_canvas.itemconfigure(self.history_canvas_window, width=e.width))
        self._bind_history_mousewheel()

        detail_panel = card(history_body, padx=12, pady=12)
        detail_panel.pack(side="right", fill="y", padx=(12, 0))
        detail_panel.configure(width=540)
        detail_panel.pack_propagate(False)

        tk.Label(detail_panel, text="CHI TIẾT OCR", bg=UI_SURFACE, fg=UI_TEXT, font=ui_font(12, bold=True)).pack(anchor="w")
        tk.Label(detail_panel, text="Double-click vào một OCR bên trái để xem lại dữ liệu đã đọc.", bg=UI_SURFACE, fg=UI_MUTED, font=ui_font(11)).pack(anchor="w", pady=(2, 8))

        self.history_detail_host = tk.Frame(detail_panel, bg=UI_SURFACE)
        self.history_detail_host.pack(fill="both", expand=True)


        self.mapping_page = tk.Frame(content, bg=UI_BG, padx=self.main_padx, pady=self.main_pady)

        mapping_shell = card(self.mapping_page, padx=16 if self.compact_ui else 18, pady=14)

        mapping_shell.pack(fill="both", expand=True)

        mapping_top = tk.Frame(mapping_shell, bg=UI_SURFACE)

        mapping_top.pack(fill="x", pady=(0, 12))

        mapping_title = tk.Frame(mapping_top, bg=UI_SURFACE)

        mapping_title.pack(side="left", fill="x", expand=True)

        mapping_icon_title = self._ui_icon("07_sidebar_mapping.png", 24)
        mapping_title_row = tk.Frame(mapping_title, bg=UI_SURFACE)
        mapping_title_row.pack(anchor="w")
        image_label(mapping_title_row, mapping_icon_title, UI_SURFACE).pack(side="left", padx=(0, 8))
        tk.Label(mapping_title_row, text="MẪU MAPPING", bg=UI_SURFACE, fg=UI_TEXT, font=ui_font(14, bold=True)).pack(side="left")

        tk.Label(

            mapping_title,

            text="Các ánh xạ đã lưu từ phần xác nhận mapping cột.",

            bg=UI_SURFACE,

            fg=UI_MUTED,

            font=("Segoe UI", max(7, scale_px(8))),

        ).pack(anchor="w", pady=(2, 0))

        icon_button(mapping_top, "Làm mới", self._render_mapping_templates, "19_action_refresh.png", width=10, variant="soft").pack(side="right")

        mapping_body = tk.Frame(mapping_shell, bg=UI_SURFACE)

        mapping_body.pack(fill="both", expand=True)

        self.mapping_templates_canvas = tk.Canvas(mapping_body, bg=UI_SURFACE, highlightthickness=0, bd=0)

        mapping_scroll = ttk.Scrollbar(mapping_body, orient="vertical", command=self.mapping_templates_canvas.yview)

        self.mapping_templates_canvas.configure(yscrollcommand=mapping_scroll.set)

        mapping_scroll.pack(side="right", fill="y")

        self.mapping_templates_canvas.pack(side="left", fill="both", expand=True)

        self.mapping_templates_inner = tk.Frame(self.mapping_templates_canvas, bg=UI_SURFACE)

        self.mapping_templates_window = self.mapping_templates_canvas.create_window((0, 0), window=self.mapping_templates_inner, anchor="nw")

        self.mapping_templates_inner.bind(

            "<Configure>",

            lambda _e: self.mapping_templates_canvas.configure(scrollregion=self.mapping_templates_canvas.bbox("all")),

        )

        self.mapping_templates_canvas.bind(

            "<Configure>",

            lambda e: self.mapping_templates_canvas.itemconfigure(self.mapping_templates_window, width=e.width),

        )
        self.mapping_templates_canvas.bind("<MouseWheel>", self._on_mapping_mousewheel, add="+")
        self.mapping_templates_canvas.bind("<Button-4>", self._on_mapping_mousewheel, add="+")
        self.mapping_templates_canvas.bind("<Button-5>", self._on_mapping_mousewheel, add="+")

        self.excel_page = tk.Frame(content, bg=UI_BG, padx=self.main_padx, pady=self.main_pady)

        excel_shell = card(self.excel_page, padx=16 if self.compact_ui else 18, pady=14)

        excel_shell.pack(fill="both", expand=True)

        top_bar = tk.Frame(excel_shell, bg=UI_SURFACE)

        top_bar.pack(fill="x", pady=(0, 10))

        tabs = tk.Frame(top_bar, bg=UI_SURFACE)

        tabs.pack(side="left")

        self.excel_recent_mode = "recent"



        def make_tab(label, mode):

            btn = tk.Label(

                tabs,

                text=label,

                bg=UI_SURFACE,

                fg=UI_TEXT if mode == self.excel_recent_mode else UI_MUTED,

                font=ui_font(11, bold=(mode == self.excel_recent_mode)),

                padx=2,

                pady=4,

                cursor="hand2",

            )

            btn.pack(side="left", padx=(0, 18))

            btn.bind("<Button-1>", lambda _e, m=mode: self._set_excel_recent_mode(m))

            return btn



        self.excel_tab_recent = make_tab("Gần đây", "recent")

        self.excel_tab_pinned = make_tab("Đã ghim", "pinned")



        ui_button(top_bar, "Trang chủ", self.show_home_page, width=12, variant="soft").pack(side="right")



        list_card = tk.Frame(excel_shell, bg=UI_SURFACE)

        list_card.pack(fill="both", expand=True)

        recent_header = tk.Frame(list_card, bg=UI_SURFACE)

        recent_header.pack(fill="x", pady=(0, 8))

        tk.Label(recent_header, text="Tên", bg=UI_SURFACE, fg=UI_MUTED, font=ui_font(10)).pack(side="left", padx=(44, 0))

        tk.Label(recent_header, text="Ngày sửa đổi", bg=UI_SURFACE, fg=UI_MUTED, font=ui_font(10)).pack(side="right", padx=(0, 10))



        recent_wrap = tk.Frame(list_card, bg=UI_SURFACE)

        recent_wrap.pack(fill="both", expand=True)

        recent_scroll = ttk.Scrollbar(recent_wrap, orient="vertical")

        recent_scroll.pack(side="right", fill="y")

        canvas = tk.Canvas(recent_wrap, bg=UI_SURFACE, highlightthickness=0, bd=0, yscrollcommand=recent_scroll.set)

        canvas.pack(side="left", fill="both", expand=True)

        recent_scroll.config(command=canvas.yview)

        self.excel_recent_canvas = canvas
        self.excel_recent_canvas.bind("<MouseWheel>", self._on_recent_mousewheel, add="+")
        self.excel_recent_canvas.bind("<Button-4>", self._on_recent_mousewheel, add="+")
        self.excel_recent_canvas.bind("<Button-5>", self._on_recent_mousewheel, add="+")

        self.excel_recent_inner = tk.Frame(canvas, bg=UI_SURFACE)

        self.excel_recent_canvas_window = canvas.create_window((0, 0), window=self.excel_recent_inner, anchor="nw")



        def _sync_scrollregion(_e=None):

            try:

                canvas.configure(scrollregion=canvas.bbox("all"))

            except Exception:

                pass



        self.excel_recent_inner.bind("<Configure>", _sync_scrollregion)

        canvas.bind("<Configure>", lambda e: canvas.itemconfigure(self.excel_recent_canvas_window, width=e.width))



        hint = tk.Label(

            excel_shell,

            text="Danh sách gộp theo đường dẫn file thật. File trùng chỉ giữ một bản. Bấm đúp để mở lại.",

            bg=UI_SURFACE,

            fg=UI_MUTED,

            font=ui_font(10),

        )

        hint.pack(anchor="w", pady=(8, 0))

        self._render_excel_recent_rows()



        self.show_home_page()



    def save_key(self):

        save_env(
            self.api_key_var.get(),
            self.model_var.get(),
            getattr(self, "screen_profile_var", tk.StringVar(value="auto")).get(),
            getattr(self, "presence_server_var", tk.StringVar(value=DEFAULT_PRESENCE_SERVER_URL)).get(),
        )

        self._set_status("Đã lưu cài đặt vào file .env", "success")


    def _set_daily_summary_text(self, lines):

        try:

            widget = getattr(self, "daily_summary_text", None)

            if widget is None:

                return

            widget.delete("1.0", "end")

            text = "\n".join(lines or []).strip()

            if not text:

                text = "Chưa có dữ liệu tổng hợp."

            widget.insert("1.0", text + "\n")

        except Exception:

            pass


    def _refresh_daily_summary_panel(self, tables=None):

        try:

            tables_to_scan = tables if isinstance(tables, list) else (self.tables or [])

            lines = build_static_jacking_daily_summary_lines(tables_to_scan)

            self._set_daily_summary_text(lines)

        except Exception:

            try:

                self._set_daily_summary_text([])

            except Exception:

                pass








    def read_each_sheet_content(self):

        if not self.excel_path:

            messagebox.showwarning("Thiếu Excel", "Bạn chưa chọn file Excel.")

            return

        try:

            analysis = analyze_workbook_sheets(self.excel_path)

            out = last_run_dir()

            out.mkdir(exist_ok=True)

            (out / "excel_each_sheet_content.json").write_text(json.dumps(analysis, ensure_ascii=False, indent=2), encoding="utf-8")



            self.excel_info.delete("1.0", "end")

            lines = []

            lines.append("ĐỌC TỪNG SHEET TRONG FILE EXCEL\n")

            lines.append("=" * 60 + "\n")

            lines.append(f"File: {self.excel_path}\n\n")



            for sh in analysis["sheets"]:

                lines.append(f"Sheet: {sh.get('sheet')}\n")

                if sh.get("empty"):

                    lines.append("- Sheet trống\n\n")

                    continue



                ur = sh.get("used_range", {})

                lines.append(f"- Loại: {sh.get('sheet_type')}\n")

                lines.append(f"- Vùng: {ur.get('from')} → {ur.get('to')} | Header: {sh.get('header_rows')} | Dữ liệu: {sh.get('data_row_count')} dòng\n")

                lines.append(f"- TỔNG: {sh.get('total_rows')} | Công thức: {sh.get('formula_count')} ô | Merge: {sh.get('merged_ranges_count')}\n")



                cols = [f"{h['col']}:{short_header_name(h['name'], 24)}" for h in sh.get("headers", [])[:30]]

                if cols:

                    lines.append("- Cột: " + " | ".join(cols) + "\n")



                if sh.get("formula_samples"):

                    fs = sh.get("formula_samples", [])[:8]

                    lines.append("- Công thức mẫu: " + " | ".join([f"{f['cell']}={f['formula']}" for f in fs]) + "\n")



                lines.append("\n")



            lines.append("Log đầy đủ: last_run_v12\\excel_each_sheet_content.json\n")

            self.excel_info.insert("1.0", "".join(lines))

            self._set_status("Đã đọc từng sheet trong file Excel.", "success")

        except Exception:

            out = last_run_dir()

            out.mkdir(exist_ok=True)

            (out / "last_error_each_sheet.txt").write_text(traceback.format_exc(), encoding="utf-8")

            messagebox.showerror("Lỗi đọc từng sheet", "Có lỗi. Xem last_run_v12/last_error_each_sheet.txt")

            self._set_status("Lỗi đọc từng sheet.", "error")






    def _profile_workbook(self, excel_path):

        """

        Đọc toàn bộ workbook: tất cả sheet, header, STT, TỔNG, chuỗi STT, công thức tổng.

        Không ghi gì vào file.

        """

        wb = load_workbook(excel_path, data_only=False)

        profiles = []

        for ws in wb.worksheets:

            try:

                header_row = find_header_row_smart(ws)

                headers = get_headers_smart(ws, header_row)

                total_row = find_total_row(ws, header_row)

                no_col = find_no_column_smart(ws, headers, header_row, total_row) if total_row else None

                chains = []

                best = None

                if total_row and no_col:

                    chains = find_all_stt_chains(ws, no_col, header_row, total_row)

                    best = select_longest_stt_chain(ws, no_col, header_row, total_row)



                formula_cols = []

                if total_row:

                    try:

                        formula_cols = capture_total_sum_columns(

                            ws,

                            total_row,

                            best[0][0] if best else None,

                            best[-1][0] if best else None

                        )

                    except Exception:

                        formula_cols = capture_formula_columns(ws, total_row)



                profiles.append({

                    "file": str(excel_path),

                    "sheet": ws.title,

                    "max_row": ws.max_row,

                    "max_col": ws.max_column,

                    "header_row": header_row,

                    "headers": [{"col": get_column_letter(c), "index": c, "name": name} for c, name in headers],

                    "total_row": total_row,

                    "stt_col": get_column_letter(no_col) if no_col else None,

                    "stt_col_index": no_col,

                    "stt_chains": [

                        {

                            "from_row": ch[0][0],

                            "to_row": ch[-1][0],

                            "from_stt": ch[0][1],

                            "to_stt": ch[-1][1],

                            "length": len(ch),

                        }

                        for ch in chains

                    ],

                    "selected_chain": {

                        "from_row": best[0][0],

                        "to_row": best[-1][0],

                        "from_stt": best[0][1],

                        "to_stt": best[-1][1],

                        "length": len(best),

                    } if best else None,

                    "sum_columns": [get_column_letter(c) for c in formula_cols],

                })

            except Exception as e:

                profiles.append({

                    "file": str(excel_path),

                    "sheet": ws.title,

                    "error": repr(e),

                })

        return profiles



    def _display_profiles(self, profiles, title="Kết quả đọc Excel"):

        self.excel_info.delete("1.0", "end")

        lines = [title + "\n", "=" * 50 + "\n"]

        for p in profiles:

            lines.append(f"\nFile: {p.get('file')}\n")

            lines.append(f"Sheet: {p.get('sheet')}\n")

            if p.get("error"):

                lines.append(f"LỖI: {p.get('error')}\n")

                continue

            lines.append(

                f"Header: {p.get('header_row')} | TỔNG: {p.get('total_row')} | STT: {p.get('stt_col')}\n"

            )

            sc = p.get("selected_chain")

            if sc:

                lines.append(

                    f"STT chọn: {sc['from_stt']} → {sc['to_stt']} "

                    f"(dòng {sc['from_row']} → {sc['to_row']}) | Tiếp: {sc['to_stt'] + 1}\n"

                )

            if p.get("sum_columns"):

                lines.append("SUM: " + ", ".join(p.get("sum_columns")) + "\n")

            headers = p.get("headers", [])

            if headers:

                lines.append("Cột: " + " | ".join([f"{h['col']}:{short_header_name(h['name'], 28)}" for h in headers[:25]]) + "\n")

        self.excel_info.insert("1.0", "".join(lines))



    def scan_current_workbook(self):

        if not self.excel_path:

            messagebox.showwarning("Thiếu Excel", "Bạn chưa chọn file Excel.")

            return

        try:

            profiles = self._profile_workbook(self.excel_path)

            out = last_run_dir()

            out.mkdir(exist_ok=True)

            (out / "current_workbook_profiles.json").write_text(json.dumps(profiles, ensure_ascii=False, indent=2), encoding="utf-8")

            each_sheet = analyze_workbook_sheets(self.excel_path)

            (out / "excel_each_sheet_content.json").write_text(json.dumps(each_sheet, ensure_ascii=False, indent=2), encoding="utf-8")

            formula_logic = read_formula_logic_for_workbook(self.excel_path)

            (out / "excel_formula_logic.json").write_text(json.dumps(formula_logic, ensure_ascii=False, indent=2), encoding="utf-8")

            self._display_profiles(profiles, "Đã đọc toàn bộ workbook hiện tại")

            self._set_status("Đã đọc toàn bộ workbook.", "success")

        except Exception:

            out = last_run_dir()

            out.mkdir(exist_ok=True)

            (out / "last_error_scan_workbook.txt").write_text(traceback.format_exc(), encoding="utf-8")

            messagebox.showerror("Lỗi đọc workbook", "Có lỗi. Xem last_run_v12/last_error_scan_workbook.txt")

            self._set_status("Lỗi đọc workbook.", "error")









































    def reset_current_session(self):
        if not messagebox.askyesno("Đặt lại", "Đặt lại Excel, ảnh OCR, preview bảng và mapping hiện tại?"):
            return

        self.image_path = None
        self.excel_path = None
        self.tk_img = None
        self.workbook = None
        self.excel_folder = None
        self.header_row = None
        self.excel_headers = []
        self.tables = []
        self.current_workflow_id = None
        self.current_workflow_date = None
        self.current_workflow_label = None
        self.current_doc_kind = None
        self.image_paths = []
        self.preview_image_index = 0
        self.excel_recent_selected_key = None
        self.history_selected_entry = None

        try:
            self.sheet_var.set("")
            if hasattr(self, "sheet_combo") and self.sheet_combo is not None:
                self.sheet_combo["values"] = []
                try:
                    self.sheet_combo.set("")
                except Exception:
                    pass
        except Exception:
            pass

        try:
            self._original_image = None
            if hasattr(self, "img_label") and self.img_label is not None:
                self.img_label.config(
                    image="",
                    text="Kéo thả ảnh OCR vào đây\n\nHỗ trợ: .jpg, .png, .jpeg",
                )
            if hasattr(self, "preview_counter_var") and self.preview_counter_var is not None:
                self.preview_counter_var.set("")
        except Exception:
            pass

        try:
            if hasattr(self, "table_editor") and self.table_editor is not None:
                self.table_editor.set_tables([])
        except Exception:
            pass

        try:
            if hasattr(self, "mapping_editor") and self.mapping_editor is not None:
                self.mapping_editor.clear()
        except Exception:
            pass

        try:
            if hasattr(self, "excel_info") and self.excel_info is not None:
                self.excel_info.delete("1.0", "end")
                self.excel_info.insert("1.0", "Đã đặt lại. Chọn Excel và ảnh OCR để bắt đầu lại.\n")
        except Exception:
            pass

        try:
            self._set_daily_summary_text([])
        except Exception:
            pass

        try:
            self._render_history_detail(None)
        except Exception:
            pass

        try:
            self._sync_excel_recent_sidebar()
        except Exception:
            pass

        self._set_status("Đã đặt lại Excel, ảnh OCR, preview và mapping.", "success")


    def _clipboard_text_widget_has_focus(self):

        try:

            widget = self.root.focus_get()

        except Exception:

            return False

        if widget is None:

            return False

        try:

            cls = widget.winfo_class()

        except Exception:

            cls = ""

        return isinstance(widget, (tk.Entry, tk.Text, ttk.Entry)) or cls in {"Entry", "TEntry", "Text"}
















    def _merged_master_cell(self, ws, row, col):

        """

        Nếu ô đang ghi là MergedCell thì trả về ô góc trên-trái của vùng merge.

        Nếu không merge thì trả về chính ô đó.

        """

        cell = ws.cell(row, col)

        if cell.__class__.__name__ != "MergedCell":

            return cell

        for rng in ws.merged_cells.ranges:

            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:

                return ws.cell(rng.min_row, rng.min_col)

        return cell



    def _safe_set_cell_value(self, ws, row, col, value):

        """

        Ghi an toàn vào Excel.

        Tránh lỗi: MergedCell object attribute 'value' is read-only

        """

        cell = ws.cell(row, col)

        if cell.__class__.__name__ == "MergedCell":

            # Nếu ô đang nằm trong vùng merge, chỉ ghi khi nó là ô master.

            # Nếu không phải master thì bỏ qua để không làm vỡ form.

            master = self._merged_master_cell(ws, row, col)

            if master.__class__.__name__ == "MergedCell":

                return False

            master.value = value

            return True

        cell.value = value

        return True



    def _apply_rows_to_workbook(self, wb):

        """

        V20.5:

        - Excel bất kỳ: đọc header từ sheet đang chọn.

        - Nếu chưa có mapping thì tự map sau khi đã có excel_headers.

        - Ảnh quyết định cột nguồn, Excel quyết định cột đích/STT/TỔNG/SUM.

        """

        if not self.sheet_var.get():

            raise ValueError("Bạn chưa chọn sheet.")



        table = self.table_editor.get_current_table()

        if not table:

            raise ValueError("Chưa có dữ liệu từ ảnh.")



        ws = wb[self.sheet_var.get()]

        header_row = find_header_row_smart(ws)

        excel_headers = get_headers_smart(ws, header_row)

        self.header_row = header_row

        self.excel_headers = excel_headers



        mapping = self.mapping_editor.get_mapping()

        if not mapping:

            mapping = auto_mapping_to_excel_columns(table["columns"], excel_headers)

            try:

                auto_idx = auto_map_columns(table["columns"], excel_headers)

                auto_idx = ensure_no_column_in_mapping(table["columns"], auto_idx, excel_headers)

                self.mapping_editor.set_mapping(table["columns"], excel_headers, auto_idx)

            except Exception:

                pass



        if not mapping:

            raise ValueError("Chưa có mapping cột. Tool không tự map được vì Excel chưa có header rõ.")



        rows = table["rows"]



        total_row = find_total_row(ws, header_row)

        if not total_row:

            raise ValueError("Không tìm thấy dòng TỔNG/TOTAL trong Excel. Tool cần dòng TỔNG để biết chèn ở đâu.")



        no_col = find_no_column_smart(ws, excel_headers, header_row, total_row)

        if not no_col:

            raise ValueError("Không tìm thấy cột STT/No trong Excel.")



        chains = find_all_stt_chains(ws, no_col, header_row, total_row)

        best_chain = select_longest_stt_chain(ws, no_col, header_row, total_row)



        # Nếu không có chuỗi STT liên tục thì dùng fallback:

        # lấy dòng dữ liệu cuối trước TỔNG + số STT lớn nhất đang có.

        used_stt_fallback = False

        if best_chain:

            first_seq_row = best_chain[0][0]

            last_seq_row = best_chain[-1][0]

            last_no = best_chain[-1][1]

        else:

            used_stt_fallback = True

            last_seq_row = find_last_data_row_before_total(ws, header_row, total_row, mapping, no_col)

            if not last_seq_row:

                raise ValueError("Không tìm thấy dòng dữ liệu cuối trước dòng TỔNG. Kiểm tra lại sheet/header/dòng tổng.")

            first_seq_row = header_row + 1

            last_no = find_last_stt_number_loose(ws, no_col, header_row, total_row)



        # Cột cần SUM dựa trên dòng TỔNG mẫu trước khi sửa

        total_sum_cols_before = capture_total_sum_columns(ws, total_row, first_seq_row, last_seq_row)

        total_sum_cols_before = [c for c in total_sum_cols_before if c != no_col]



        # V22 FIX CHUẨN:

        # Không ghi vào vùng trống/merged có sẵn vì dễ phá form.

        # Luôn insert dòng mới ngay trước dòng TỔNG rồi mới nhập dữ liệu.

        garbage_count = 0



        insert_at = total_row

        ws.insert_rows(insert_at, amount=len(rows))



        total_row_after = total_row + len(rows)

        style_row = last_seq_row



        target_rows = list(range(insert_at, insert_at + len(rows)))



        for r_offset, row in enumerate(rows):

            dst_row = target_rows[r_offset]



            # Copy đúng format từ dòng dữ liệu mẫu

            copy_style_row(ws, style_row, dst_row, ws.max_column)

            try:

                copy_row_dimension(ws, style_row, dst_row)

            except Exception:

                pass



            # Copy công thức dòng mẫu nếu có

            try:

                apply_row_formulas_from_template(ws, style_row, dst_row)

            except Exception:

                pass



            # STT tự nối tiếp, không lấy từ ảnh

            self._safe_set_cell_value(ws, dst_row, no_col, last_no + r_offset + 1)



            # Ghi dữ liệu OCR theo mapping

            for src_idx, excel_col in enumerate(mapping):

                if excel_col is None:

                    continue

                if excel_col == no_col:

                    continue



                # Nếu ô đang có công thức thì giữ công thức

                if is_formula_value(ws.cell(dst_row, excel_col).value):

                    continue



                val = row[src_idx] if src_idx < len(row) else ""

                self._safe_set_cell_value(ws, dst_row, excel_col, convert_excel_value(val))



        sum_first_row = first_seq_row

        sum_last_row = target_rows[-1]

        set_total_formulas_by_template(

            ws,

            total_row_after,

            total_sum_cols_before,

            sum_first_row,

            sum_last_row

        )

        force_workbook_recalculate(wb)



        out = last_run_dir()

        out.mkdir(exist_ok=True)

        logic = {

            "rule": "V20.5 auto map sau khi đọc excel_headers",

            "header_row": header_row,

            "no_col": no_col,

            "total_row_before": total_row,

            "total_row_after": total_row_after,

            "selected_chain": {

                "from_row": first_seq_row,

                "to_row": last_seq_row,

                "from_stt": (best_chain[0][1] if best_chain else None),

                "to_stt": last_no,

                "length": (len(best_chain) if best_chain else 0),

                "fallback": used_stt_fallback,

            },

            "insert_at_row": insert_at,

            "new_stt_start": last_no + 1,

            "new_stt_end": last_no + len(rows),

            "sum_range_rows": [sum_first_row, sum_last_row],

            "sum_columns": total_sum_cols_before,

            "mapping": [

                {

                    "source": table["columns"][i] if i < len(table["columns"]) else "",

                    "excel_col": excel_col,

                    "excel_letter": get_column_letter(excel_col) if excel_col else None

                }

                for i, excel_col in enumerate(mapping)

            ]

        }

        (out / "v20_5_apply_logic.json").write_text(json.dumps(logic, ensure_ascii=False, indent=2), encoding="utf-8")



        return {

            "sheet": ws.title,

            "header_row": header_row,

            "first_stt_row": first_seq_row,

            "last_stt_row": last_seq_row,

            "last_stt_before": last_no,

            "start_fill_row": insert_at,

            "next_stt_start": last_no + 1,

            "rows_added": len(rows),

            "garbage_deleted_count": garbage_count,

            "total_row_after": total_row_after,

            "sum_first_row": sum_first_row,

            "sum_last_row": sum_last_row,

            "sum_columns_count": len(total_sum_cols_before),

            "used_stt_fallback": used_stt_fallback,

        }













from gk_pilepro.ui.gk_settings_ui import install_settings_ui

install_settings_ui(App)

from gk_pilepro.ui.gk_admin_ui import install_admin_ui

install_admin_ui(App)

from gk_pilepro.ui.gk_history_ui import install_history_ui

install_history_ui(App)

from gk_pilepro.ui.gk_excel_ui import install_excel_ui

install_excel_ui(App)

from gk_pilepro.ui.gk_ocr_ui import install_ocr_ui

install_ocr_ui(App)

from gk_pilepro.gk_overrides import install_app_overrides


install_app_overrides(App)


def _canvas_round_rect(canvas, x1, y1, x2, y2, radius=12, **kwargs):
    radius = max(1, int(radius))
    points = [
        x1 + radius, y1, x2 - radius, y1, x2, y1, x2, y1 + radius,
        x2, y2 - radius, x2, y2, x2 - radius, y2, x1 + radius, y2,
        x1, y2, x1, y2 - radius, x1, y1 + radius, x1, y1,
    ]
    return canvas.create_polygon(points, smooth=True, splinesteps=24, **kwargs)


def _cover_image(source, size):
    target_w, target_h = size
    img = source.convert("RGB")
    src_w, src_h = img.size
    scale = max(target_w / max(1, src_w), target_h / max(1, src_h))
    img = img.resize((max(1, int(src_w * scale)), max(1, int(src_h * scale))), Image.LANCZOS)
    left = max(0, (img.width - target_w) // 2)
    top = max(0, (img.height - target_h) // 2)
    return img.crop((left, top, left + target_w, top + target_h))


def _fit_logo_for_splash(path, max_size):
    logo = Image.open(path).convert("RGBA")
    alpha_bbox = logo.getchannel("A").getbbox()
    has_transparent_canvas = bool(alpha_bbox and alpha_bbox != (0, 0, logo.width, logo.height))
    # Older source logo images included the company title on a white canvas.
    # Keep only the emblem there; true transparent logo assets are already clean.
    if not has_transparent_canvas and logo.height > logo.width * 0.65:
        logo = logo.crop((0, 0, logo.width, int(logo.height * 0.78)))

    pixels = logo.load()
    width, height = logo.size
    visited = bytearray(width * height)
    stack = []

    def near_white(x, y):
        r, g, b, a = pixels[x, y]
        return a <= 8 or (r >= 168 and g >= 168 and b >= 168 and (max(r, g, b) - min(r, g, b) <= 42))

    for x in range(width):
        if near_white(x, 0):
            stack.append((x, 0))
        if near_white(x, height - 1):
            stack.append((x, height - 1))
    for y in range(height):
        if near_white(0, y):
            stack.append((0, y))
        if near_white(width - 1, y):
            stack.append((width - 1, y))

    while stack:
        x, y = stack.pop()
        idx = y * width + x
        if visited[idx] or not near_white(x, y):
            continue
        visited[idx] = 1
        r, g, b, _a = pixels[x, y]
        pixels[x, y] = (r, g, b, 0)
        if x > 0:
            stack.append((x - 1, y))
        if x < width - 1:
            stack.append((x + 1, y))
        if y > 0:
            stack.append((x, y - 1))
        if y < height - 1:
            stack.append((x, y + 1))

    alpha_bbox = logo.getchannel("A").getbbox()
    if alpha_bbox:
        logo = logo.crop(alpha_bbox)
    logo.thumbnail(max_size, Image.LANCZOS)
    return logo.filter(ImageFilter.SHARPEN)


def _load_startup_update_notice():
    try:
        value = str(os.getenv("STARTUP_UPDATE_NOTICE") or "").strip()
        if value:
            return value
    except Exception:
        pass
    try:
        path = app_data_path("startup_update_notice.txt")
        if path.exists():
            value = path.read_text(encoding="utf-8", errors="replace").strip()
            if value:
                return value.splitlines()[0].strip()
    except Exception:
        pass
    return ""


def _splash_font(size_px, bold=False):
    candidates = (
        "C:/Windows/Fonts/segoeuib.ttf" if bold else "C:/Windows/Fonts/segoeui.ttf",
        "C:/Windows/Fonts/arialbd.ttf" if bold else "C:/Windows/Fonts/arial.ttf",
        "segoeuib.ttf" if bold else "segoeui.ttf",
        "arialbd.ttf" if bold else "arial.ttf",
    )
    for name in candidates:
        try:
            return ImageFont.truetype(name, size_px)
        except Exception:
            pass
    return ImageFont.load_default()


def _draw_splash_gradient_text(draw, xy, text, font, top_color, bottom_color, stroke_fill, stroke_width=2):
    x, y = xy
    bbox = draw.textbbox((0, 0), text, font=font, stroke_width=stroke_width)
    tw = max(1, bbox[2] - bbox[0])
    th = max(1, bbox[3] - bbox[1])
    mask = Image.new("L", (tw + stroke_width * 4, th + stroke_width * 4), 0)
    mask_draw = ImageDraw.Draw(mask)
    mask_draw.text((stroke_width * 2 - bbox[0], stroke_width * 2 - bbox[1]), text, font=font, fill=255)

    gradient = Image.new("RGBA", mask.size, (0, 0, 0, 0))
    grad_px = gradient.load()
    for row in range(mask.height):
        t = row / max(1, mask.height - 1)
        r = int(top_color[0] * (1 - t) + bottom_color[0] * t)
        g = int(top_color[1] * (1 - t) + bottom_color[1] * t)
        b = int(top_color[2] * (1 - t) + bottom_color[2] * t)
        for col in range(mask.width):
            grad_px[col, row] = (r, g, b, 255)
    gradient.putalpha(mask)

    draw.text((x + 4, y + 5), text, font=font, fill=(0, 0, 0, 165), stroke_width=stroke_width, stroke_fill=(0, 0, 0, 150))
    draw.text((x, y), text, font=font, fill=stroke_fill, stroke_width=stroke_width, stroke_fill=stroke_fill)
    draw._image.alpha_composite(gradient, (x - stroke_width * 2, y - stroke_width * 2))


def _render_splash_overlay(w, h, progress, step_text, update_text):
    scale = 2
    W, H = w * scale, h * scale
    overlay = Image.new("RGBA", (W, H), (0, 0, 0, 0))
    draw = ImageDraw.Draw(overlay)
    S = scale

    title_font = _splash_font(max(34, min(52, int(w * 0.047))) * S, bold=True)
    part_a = "NỀN MÓNG "
    part_b = "GIA KHÁNH"
    a_bbox = draw.textbbox((0, 0), part_a, font=title_font, stroke_width=2 * S)
    b_bbox = draw.textbbox((0, 0), part_b, font=title_font, stroke_width=2 * S)
    title_w = (a_bbox[2] - a_bbox[0]) + (b_bbox[2] - b_bbox[0])
    title_y = int(h * 0.615) * S
    title_x = (W - title_w) // 2
    _draw_splash_gradient_text(
        draw,
        (title_x, title_y),
        part_a,
        title_font,
        (10, 30, 49),
        (2, 10, 20),
        (178, 125, 36, 255),
        stroke_width=2 * S,
    )
    _draw_splash_gradient_text(
        draw,
        (title_x + (a_bbox[2] - a_bbox[0]), title_y),
        part_b,
        title_font,
        (255, 216, 70),
        (178, 103, 4),
        (59, 35, 6, 255),
        stroke_width=2 * S,
    )

    panel_left = int(w * 0.145) * S
    panel_right = int(w * 0.855) * S
    panel_y = int(h * 0.755) * S
    panel_h = max(42, int(h * 0.082)) * S
    bevel = max(16, int((panel_h / S) * 0.42)) * S
    points = [
        (panel_left + bevel, panel_y - panel_h // 2),
        (panel_right - bevel, panel_y - panel_h // 2),
        (panel_right, panel_y),
        (panel_right - bevel, panel_y + panel_h // 2),
        (panel_left + bevel, panel_y + panel_h // 2),
        (panel_left, panel_y),
    ]
    draw.polygon(points, fill=(12, 17, 22, 232), outline=(126, 92, 35, 255))
    draw.line([(panel_left + bevel, panel_y + panel_h // 2 - 3 * S), (panel_right - bevel, panel_y + panel_h // 2 - 3 * S)], fill=(50, 36, 15, 180), width=S)

    bar_left = panel_left + int((panel_h / S) * 1.28) * S
    bar_right = panel_right - int((panel_h / S) * 1.28) * S
    bar_h = max(12, int(h * 0.020)) * S
    bar_top = panel_y - bar_h // 2
    radius = bar_h // 2
    draw.rounded_rectangle((bar_left, bar_top, bar_right, bar_top + bar_h), radius=radius, fill=(16, 22, 27, 255), outline=(165, 123, 42, 255), width=S)

    fill_w = int((bar_right - bar_left) * (max(0.0, min(23.0, float(progress))) / 23.0))
    if fill_w > 0:
        fill_right = min(bar_right - 2 * S, bar_left + fill_w)
        fill_box = (bar_left + 3 * S, bar_top + 3 * S, fill_right, bar_top + bar_h - 3 * S)
        if fill_box[2] > fill_box[0]:
            fill_mask = Image.new("L", overlay.size, 0)
            mask_draw = ImageDraw.Draw(fill_mask)
            mask_draw.rounded_rectangle(fill_box, radius=max(1, radius - 3 * S), fill=255)
            fill_layer = Image.new("RGBA", overlay.size, (0, 0, 0, 0))
            fill_draw = ImageDraw.Draw(fill_layer)
            for yy in range(fill_box[1], fill_box[3] + 1):
                t = (yy - fill_box[1]) / max(1, fill_box[3] - fill_box[1])
                col = (
                    int(255 * (1 - t) + 218 * t),
                    int(235 * (1 - t) + 139 * t),
                    int(103 * (1 - t) + 16 * t),
                    255,
                )
                fill_draw.line((fill_box[0], yy, fill_box[2], yy), fill=col, width=1)
            fill_layer.putalpha(fill_mask)
            overlay.alpha_composite(fill_layer)

        glow_x = max(bar_left + 8 * S, fill_right)
        glow = Image.new("RGBA", overlay.size, (0, 0, 0, 0))
        glow_draw = ImageDraw.Draw(glow)
        glow_draw.line((glow_x - 34 * S, panel_y, glow_x + 34 * S, panel_y), fill=(255, 238, 145, 210), width=3 * S)
        glow_draw.line((glow_x, panel_y - 17 * S, glow_x, panel_y + 17 * S), fill=(255, 248, 205, 230), width=2 * S)
        glow_draw.ellipse((glow_x - 5 * S, panel_y - 5 * S, glow_x + 5 * S, panel_y + 5 * S), fill=(255, 250, 217, 255))
        glow = glow.filter(ImageFilter.GaussianBlur(radius=0.45 * S))
        overlay.alpha_composite(glow)

    status_font = _splash_font(max(15, min(22, int(w * 0.019))) * S, bold=False)
    status_bbox = draw.textbbox((0, 0), step_text, font=status_font)
    status_x = (W - (status_bbox[2] - status_bbox[0])) // 2
    status_y = int(h * 0.848) * S
    draw.text((status_x, status_y), step_text, font=status_font, fill=(218, 165, 50, 245))

    update_text = str(update_text or "").strip()
    if update_text:
        update_font = _splash_font(max(11, min(15, int(w * 0.012))) * S, bold=True)
        pill_w = min(int(w * 0.42), max(int(w * 0.27), len(update_text) * max(6, int(w * 0.007)))) * S
        pill_h = max(28, int(h * 0.062)) * S
        pill_x = (W - pill_w) // 2
        pill_y = int(h * 0.900) * S
        draw.rounded_rectangle((pill_x, pill_y, pill_x + pill_w, pill_y + pill_h), radius=16 * S, fill=(13, 23, 36, 230), outline=(152, 124, 57, 255), width=S)
        dot = max(8, int((pill_h / S) * 0.28)) * S
        draw.ellipse((pill_x + 14 * S, pill_y + (pill_h - dot) // 2, pill_x + 14 * S + dot, pill_y + (pill_h + dot) // 2), fill=(248, 195, 41, 255))
        draw.text((pill_x + 14 * S + dot + 10 * S, pill_y + pill_h // 2), update_text, font=update_font, fill=(255, 233, 166, 255), anchor="lm")

    return overlay.resize((w, h), Image.LANCZOS)


def _draw_startup_splash(canvas, w, h, progress=0, step_text="Đang khởi động hệ thống...", update_text=""):
    canvas.delete("all")
    progress = max(0.0, min(23.0, float(progress or 0)))

    if getattr(canvas, "_splash_static_key", None) != (w, h):
        static_img = Image.new("RGBA", (w, h), (7, 17, 31, 255))
        bg_path = resource_path(*APP_SPLASH_BG_PNG.parts)
        if bg_path.exists():
            try:
                bg_img = _cover_image(Image.open(bg_path), (w, h)).convert("RGBA")
                bg_img = Image.alpha_composite(bg_img, Image.new("RGBA", (w, h), (5, 10, 18, 46)))
                shade = Image.new("RGBA", (w, h), (0, 0, 0, 0))
                shade_draw = ImageDraw.Draw(shade)
                shade_draw.rectangle((0, int(h * 0.68), w, h), fill=(0, 0, 0, 66))
                static_img = Image.alpha_composite(bg_img, shade)
            except Exception:
                pass

        logo_path = resource_path(*APP_SPLASH_LOGO_PNG.parts)
        if not logo_path.exists():
            logo_path = resource_path(*APP_LOGO_PNG.parts)
        if logo_path.exists():
            try:
                logo = _fit_logo_for_splash(logo_path, (int(w * 0.50), int(h * 0.50)))
                static_img.alpha_composite(logo, ((w - logo.width) // 2, int(h * 0.34) - logo.height // 2))
            except Exception:
                pass
        canvas._splash_static_key = (w, h)
        canvas._splash_static_photo = ImageTk.PhotoImage(static_img.convert("RGB"))

    canvas.create_image(0, 0, image=canvas._splash_static_photo, anchor="nw")

    try:
        overlay = _render_splash_overlay(w, h, progress, step_text, update_text)
        overlay_photo = ImageTk.PhotoImage(overlay)
        canvas._splash_overlay = overlay_photo
        canvas.create_image(0, 0, image=overlay_photo, anchor="nw")
    except Exception:
        pass


def _show_startup_splash(root, update_text=""):
    try:
        splash = tk.Toplevel(root)
        splash.overrideredirect(True)
        splash.configure(bg="#050b14")
        splash.attributes("-topmost", True)
        sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
        w = min(980, max(720, int(sw * 0.74)))
        h = int(w * 9 / 16)
        if h > int(sh * 0.78):
            h = int(sh * 0.78)
            w = int(h * 16 / 9)
        splash.geometry(f"{w}x{h}+{max(0, (sw - w) // 2)}+{max(0, (sh - h) // 2)}")
        canvas = tk.Canvas(splash, width=w, height=h, bg="#050b14", bd=0, highlightthickness=0)
        canvas.pack(fill="both", expand=True)
        splash._canvas = canvas
        splash._size = (w, h)
        splash._update_text = str(update_text or "").strip()
        splash._progress = 0.0
        _draw_startup_splash(canvas, w, h, 0, "Đang khởi động hệ thống...", splash._update_text)
        splash.update_idletasks()
        splash.update()
        return splash
    except Exception:
        return None


def _update_startup_splash(splash, progress, step_text):
    try:
        if splash is None or not splash.winfo_exists():
            return
        canvas = splash._canvas
        w, h = splash._size
        start = float(getattr(splash, "_progress", 0.0))
        target = max(0.0, min(23.0, float(progress or 0)))
        if target < start:
            start = target
        distance = abs(target - start)
        frames = max(8, min(22, int(distance * 2.2)))
        for idx in range(1, frames + 1):
            t = idx / frames
            eased = 1 - pow(1 - t, 3)
            value = start + (target - start) * eased
            _draw_startup_splash(canvas, w, h, value, step_text, getattr(splash, "_update_text", ""))
            splash._progress = value
            splash.update_idletasks()
            splash.update()
            time.sleep(0.012)
        splash._progress = target
    except Exception:
        pass


def _close_startup_splash(splash):
    try:
        if splash is not None and splash.winfo_exists():
            splash.destroy()
    except Exception:
        pass


def main():

    try:

        try:

            import ctypes

            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("GiaKhanh.App")

        except Exception:

            pass

        root = tk.Tk()

        try:
            root.withdraw()
        except Exception:
            pass

        if not is_admin_build():
            try:
                server_url = presence_server_url_from_env()
                if not check_presence_server_alive(server_url, timeout=0.1):
                    messagebox.showwarning(
                        "Server đang bảo trì",
                        "Server đang bảo trì.\n\nVui lòng chờ admin mở lại server để tiếp tục sử dụng ứng dụng.",
                    )
                    try:
                        root.destroy()
                    except Exception:
                        pass
                    return
            except Exception:
                try:
                    messagebox.showwarning(
                        "Server đang bảo trì",
                        "Server đang bảo trì.\n\nVui lòng chờ admin mở lại server để tiếp tục sử dụng ứng dụng.",
                    )
                    root.destroy()
                except Exception:
                    pass
                return

        splash = _show_startup_splash(root, _load_startup_update_notice())
        _update_startup_splash(splash, 3, "Đang khởi động hệ thống...")
        _update_startup_splash(splash, 8, "Đang khởi động hệ thống...")
        app = App(root)
        _update_startup_splash(splash, 18, "Đang khởi động hệ thống...")
        _update_startup_splash(splash, 23, "Đang khởi động hệ thống...")
        _close_startup_splash(splash)
        try:
            if not getattr(app, "member_locked", False):
                root.deiconify()
                root.lift()
        except Exception:
            pass

        root.mainloop()

    except Exception as exc:

        try:
            write_role_error_log("main", exc)

        except Exception:

            pass

        try:

            import traceback

            app_data_path("gk_pilepro_error.log").write_text(

                traceback.format_exc(),

                encoding="utf-8",

            )

        except Exception:

            pass

        try:

            messagebox.showerror("GK PilePro", f"Không mở được ứng dụng:\n{exc}")

        except Exception:

            pass



if __name__ == "__main__":
    if "--presence-server" in sys.argv:
        try:
            sys.argv = [sys.argv[0]] + [arg for arg in sys.argv[1:] if arg != "--presence-server"]
            import presence_server as _presence_server_main_mod
            _presence_server_main_mod.main()
        except SystemExit:
            raise
        except Exception:
            try:
                app_data_path("presence_server_error.log").write_text(traceback.format_exc(), encoding="utf-8")
            except Exception:
                pass
            raise
    else:
        main()



